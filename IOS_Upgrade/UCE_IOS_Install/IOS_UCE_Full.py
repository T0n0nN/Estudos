#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from netmiko import ConnectHandler
import os
import sys
import time
import getpass
import logging
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Tuple, List, Dict
from dotenv import load_dotenv
from tqdm import tqdm
import socket

# -------------------------
# Carregar .env (raiz do workspace e pasta atual)
# -------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(os.path.dirname(SCRIPT_DIR))
load_dotenv(os.path.join(ROOT_DIR, ".env"))
load_dotenv(os.path.join(SCRIPT_DIR, ".env"))

# -------------------------
#          Configuráveis
# -------------------------
# Caminho da planilha (padrão: "Switches IOS Full.xlsx" no mesmo diretório). Pode sobrescrever via UCE_SWITCHES_XLSX_FULL
XLSX_FILE_ENV   = os.getenv("UCE_SWITCHES_XLSX_FULL", "").strip()
# Limitar concorrência (reloads simultâneos em UCEs). Default 3
MAX_WORKERS     = max(1, min(3, int(os.getenv("IOS_MAX_WORKERS", "3"))))
CONNECT_TIMEOUT = int(os.getenv("IOS_CONNECT_TIMEOUT", "10"))   # segundos
RETRY_COUNT     = int(os.getenv("IOS_RETRY_COUNT", "2"))
RETRY_DELAY     = int(os.getenv("IOS_RETRY_DELAY", "5"))        # segundos entre tentativas
# Storage: exigir sdflash (ou variantes). Não permitir flash/bootflash para UCE.
STORAGE_PREF    = os.getenv("IOS_STORAGE", "sdflash:").strip().lower()
# Espera por reload e validação pós-boot
WAIT_RELOAD     = (os.getenv("IOS_WAIT_RELOAD", "1").strip().lower() in ("1", "true", "yes"))
DOWN_WAIT_SEC   = int(os.getenv("IOS_DOWN_WAIT_SEC", "600"))     # máx. 10 min p/ cair
UP_WAIT_SEC     = int(os.getenv("IOS_UP_WAIT_SEC", "1800"))      # máx. 30 min p/ voltar
POLL_INTERVAL   = int(os.getenv("IOS_POLL_INTERVAL", "10"))      # intervalo entre tentativas

# -------------------------
#          Logger (console)
# -------------------------

def setup_logger() -> logging.Logger:
    logger = logging.getLogger("IOSUCE-FULL")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(ch)
    return logger

# -------------------------
#       Utilidades SSH
# -------------------------

def _enable_if_needed(conn, secret: str | None, logger: logging.Logger | None = None):
    try:
        if secret:
            conn.enable()
        try:
            conn.send_command("terminal length 0", expect_string=r"#", delay_factor=2)
        except Exception:
            pass
    except Exception as e:
        if logger:
            logger.debug(f"Falha ao entrar em enable: {e}")

def _detect_storage(conn) -> str:
    try:
        out = conn.send_command("dir sdflash:", expect_string=r"#", delay_factor=2)
        if "Directory of" in out or "bytes" in (out or "").lower():
            return "sdflash:"
    except Exception:
        pass
    for cand in ("sdflash0:", "sdflash1:"):
        try:
            out = conn.send_command(f"dir {cand}", expect_string=r"#", delay_factor=2)
            if "Directory of" in out or "bytes" in (out or "").lower():
                return cand
        except Exception:
            pass
    # como fallback, retornar sdflash:
    return "sdflash:"

def _image_exists(conn, storage: str, image_name: str) -> bool:
    try:
        out = conn.send_command(f"dir {storage}{image_name}", expect_string=r"#", delay_factor=3)
        return image_name in out and ("No such file" not in out and "Error" not in out)
    except Exception:
        return False

# -------------------------
#      TFTP e MD5
# -------------------------

def tftp_copy_image(conn, tftp_server: str, image_name: str, storage: str, logger: logging.Logger) -> Tuple[bool, str]:
    """Copia via TFTP usando URL completa, tentando responder prompts automaticamente."""
    dest = f"{storage}{image_name}"
    cmd = f"copy tftp://{tftp_server}/{image_name} {dest}"
    out = conn.send_command_timing(cmd, delay_factor=8, max_loops=2000)
    for _ in range(100):
        low = (out or '').lower()
        if "address or name of remote host" in low:
            out += "\n" + conn.send_command_timing(tftp_server, delay_factor=8, max_loops=2000)
            continue
        if "source filename" in low:
            out += "\n" + conn.send_command_timing(image_name, delay_factor=8, max_loops=2000)
            continue
        if "destination filename" in low:
            out += "\n" + conn.send_command_timing("\n", delay_factor=8, max_loops=2000)
            continue
        if "overwrite" in low and ("[confirm]" in low or "[y/n]" in low):
            confirm = "\n" if "[confirm]" in low else "y"
            out += "\n" + conn.send_command_timing(confirm, delay_factor=8, max_loops=2000)
            continue
        if any(x in low for x in ["error", "timed out", "no space", "access violation", "not found", "%invalid"]):
            break
        more = ""
        try:
            more = conn.read_channel()
        except Exception:
            pass
        if more:
            out += more
            time.sleep(1)
            continue
        if any(k in low for k in ["bytes copied", "copied in", "bytes/sec", "transfer completed", "copied, secs"]):
            break
        time.sleep(1)
    logger.debug(f"Saída do copy: \n{out}")
    low = (out or '').lower()
    ok = any(k in low for k in ["bytes copied", "copied in", "bytes/sec", "transfer completed"]) and not any(k in low for k in ["error", "no space", "timed out", "access violation", "%invalid"])
    return ok, out

def verify_md5(conn, storage: str, image_name: str, logger: logging.Logger) -> Tuple[bool, str, str]:
    """Retorna (ok, md5_calc, output)."""
    cmd = f"verify /md5 {storage}{image_name}"
    out = conn.send_command_timing(cmd, delay_factor=12, max_loops=4000)
    import re
    md5_match = re.search(r"=\s*([a-fA-F0-9]{32})", out or "")
    if not md5_match:
        md5_match = re.search(r"\b([a-fA-F0-9]{32})\b", out or "")
    calc = md5_match.group(1).lower() if md5_match else ""
    ok = bool(calc)
    logger.debug(f"Saída do verify md5: \n{out}")
    return ok, calc, out

# -------------------------
#   Boot e Reload
# -------------------------

def _running_boot_lines(conn) -> List[str]:
    try:
        out = conn.send_command("show run | i ^boot system", expect_string=r"#", delay_factor=2)
    except Exception:
        out = ""
    return [ln.strip() for ln in (out or "").splitlines() if ln and ln.strip().startswith("boot system")]

def _set_boot_to_bin(conn, storage: str, bin_file: str, logger: logging.Logger) -> Tuple[bool, str]:
    try:
        conn.config_mode()
    except Exception:
        pass
    for ln in list(_running_boot_lines(conn)):
        try:
            conn.send_command_timing("no " + ln, delay_factor=3, max_loops=200)
        except Exception:
            pass
    boot_path = (storage if storage else "sdflash:") + bin_file
    try:
        conn.send_command_timing(f"boot system {boot_path}", delay_factor=3, max_loops=200)
    except Exception:
        pass
    try:
        if conn.check_config_mode():
            conn.exit_config_mode()
    except Exception:
        pass
    try:
        w = conn.send_command_timing("write memory", delay_factor=5, max_loops=200)
        wl = (w or "").lower()
        if "[confirm]" in wl:
            conn.send_command_timing("\n", delay_factor=3, max_loops=100)
        elif "[y/n]" in wl:
            conn.send_command_timing("y", delay_factor=3, max_loops=100)
        elif "[yes/no]" in wl:
            conn.send_command_timing("yes", delay_factor=3, max_loops=100)
    except Exception:
        pass
    try:
        sb = conn.send_command("show boot", expect_string=r"#", delay_factor=3)
    except Exception:
        sb = ""
    ok = bin_file.lower() in (sb or "").lower()
    return ok, ("OK" if ok else "não foi possível confirmar no show boot")

def _send_reload_with_confirms(conn) -> str:
    try:
        r = conn.send_command_timing("reload", delay_factor=2, max_loops=100)
    except Exception as e:
        try:
            conn.write_channel("reload\n")
        except Exception:
            pass
        return f"reload emitido (exceção tolerada: {e})"
    rl = (r or '').lower()
    try:
        if "[confirm]" in rl:
            conn.send_command_timing("\n", delay_factor=1, max_loops=20)
        elif "[y/n]" in rl:
            conn.send_command_timing("y", delay_factor=1, max_loops=20)
        elif "[yes/no]" in rl:
            conn.send_command_timing("yes", delay_factor=1, max_loops=20)
        if "save? [yes/no]" in rl or "save the configuration" in rl:
            conn.send_command_timing("no", delay_factor=1, max_loops=20)
    except Exception:
        pass
    return rl

def _tcp_can_connect(host: str, port: int = 22, timeout: int = 3) -> bool:
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except Exception:
        return False

def _await_reload_and_reconnect(device: dict, logger: logging.Logger) -> Tuple[bool, object, str]:
    host = device.get("host")
    start = time.time()
    saw_down = False
    while time.time() - start < DOWN_WAIT_SEC:
        if not _tcp_can_connect(host):
            saw_down = True
            break
        time.sleep(POLL_INTERVAL)
    if not saw_down:
        logger.debug(f"{host} ➜ SSH não ficou indisponível dentro do tempo; seguindo para reconexão.")
    start = time.time()
    while time.time() - start < UP_WAIT_SEC:
        if _tcp_can_connect(host):
            try:
                conn = ConnectHandler(**device)
                return True, conn, "Reconectado"
            except Exception:
                pass
        time.sleep(POLL_INTERVAL)
    return False, None, "Não reconectou dentro do tempo limite"

# -------------------------
#     Leitura da planilha "Switches IOS Full.xlsx"
# -------------------------

def _discover_xlsx_path() -> str:
    if XLSX_FILE_ENV and os.path.isfile(XLSX_FILE_ENV):
        return XLSX_FILE_ENV
    candidates = [
        os.path.join(SCRIPT_DIR, "Switches IOS Full.xlsx"),
        os.path.join(SCRIPT_DIR, "switches ios full.xlsx"),
        os.path.join(SCRIPT_DIR, "Switches_IOS_Full.xlsx"),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    return os.path.join(SCRIPT_DIR, "Switches IOS Full.xlsx")


def read_plan_xlsx(xlsx_path: str) -> Tuple[List[Dict[str, str]], Dict[str, str]]:
    def normalize_key(s: str) -> str:
        s = (s or "").strip().lower()
        s = s.replace("_", " ").replace("-", " ")
        s = " ".join(s.split())
        return s
    def canonical_key(s: str) -> str:
        return "".join(ch for ch in (s or "").lower() if ch.isalnum())
    def normalize_columns(cols: List[str]) -> Dict[str, str]:
        synonyms = {
            # essenciais
            "ip": "IP",
            "endereco ip": "IP",
            "address": "IP",
            # tftp
            "tftp": "TFTP",
            "tftp server": "TFTP",
            "tftp server ip": "TFTP",
            "tftp ip": "TFTP",
            "tftp_server": "TFTP",
            "tftp_server_ip": "TFTP",
            # imagem
            "image": "IMAGE",
            "image filename": "IMAGE",
            "image file": "IMAGE",
            "image_filename": "IMAGE",
            "imagem": "IMAGE",
            "arquivo": "IMAGE",
            "bin": "IMAGE",
            # md5s
            "md5": "MD5",
            "hash": "MD5",
            "cisco md5": "CISCO_MD5",
            "ciscomd5": "CISCO_MD5",
            # status
            "ios status": "IOS_STATUS",
            "ios_status": "IOS_STATUS",
            "status": "IOS_STATUS",
        }
        mapped: Dict[str, str] = {}
        for c in cols:
            if not isinstance(c, str):
                continue
            raw = c.strip()
            low = normalize_key(raw)
            if low in synonyms and synonyms[low] not in mapped:
                mapped[synonyms[low]] = raw
            else:
                if raw not in mapped:
                    mapped[raw] = raw
        # Fallbacks por chave canônica
        can2raw: Dict[str, str] = {}
        for c in cols:
            if isinstance(c, str) and c.strip():
                can2raw.setdefault(canonical_key(c), c.strip())
        if "IP" not in mapped:
            for k in ("ip", "address", "enderecoip", "endereco"):
                if k in can2raw:
                    mapped["IP"] = can2raw[k]; break
        if "TFTP" not in mapped:
            for k in ("tftpserverip", "tftpip", "tftpserver", "tftp", "tftp_server", "tftp_server_ip"):
                if k in can2raw:
                    mapped["TFTP"] = can2raw[k]; break
        if "IMAGE" not in mapped:
            for k in ("imagefilename", "imagefile", "image", "imagem", "bin", "arquivo", "filename"):
                if k in can2raw:
                    mapped["IMAGE"] = can2raw[k]; break
        if "MD5" not in mapped:
            for k in ("md5", "hash"):
                if k in can2raw:
                    mapped["MD5"] = can2raw[k]; break
        if "CISCO_MD5" not in mapped:
            for k in ("ciscomd5", "ciscomd5hash", "ciscohash", "cisco md5", "ciscomd5 "):
                if k in can2raw:
                    mapped["CISCO_MD5"] = can2raw[k]; break
        if "IOS_STATUS" not in mapped:
            for k in ("iosstatus", "ios_status", "status"):
                if k in can2raw:
                    mapped["IOS_STATUS"] = can2raw[k]; break
        return mapped
    try:
        import pandas as pd
        df = pd.read_excel(xlsx_path)
        colmap = normalize_columns(list(df.columns))
        required = ["IP", "TFTP", "IMAGE", "CISCO_MD5"]
        missing = [r for r in required if r not in colmap]
        if missing:
            raise ValueError(
                "Planilha precisa conter as colunas (IP, TFTP, IMAGE, Cisco MD5). Não reconhecidas: "
                + ", ".join(missing)
                + ". Cabeçalhos encontrados: "
                + ", ".join(map(str, df.columns))
            )
        df = df.fillna("")
        rows: List[Dict[str, str]] = []
        for i, row in df.iterrows():
            ip    = str(row[colmap["IP"]]).strip()
            tftp  = str(row[colmap["TFTP"]]).strip()
            image = str(row[colmap["IMAGE"]]).strip()
            md5   = str(row[colmap.get("MD5", "")] if "MD5" in colmap else "").strip().lower()
            cisco = str(row[colmap["CISCO_MD5"]]).strip().lower()
            iosst = str(row[colmap.get("IOS_STATUS", "")] if "IOS_STATUS" in colmap else "").strip()
            if not ip:
                continue
            rows.append({
                "_row": int(i) + 2,
                "IP": ip,
                "TFTP": tftp,
                "IMAGE": image,
                "MD5": md5,
                "CISCO_MD5": cisco,
                "IOS_STATUS": iosst,
                "Nome": str(row.get("Nome", "")).strip(),
            })
        return rows, colmap
    except ImportError:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        colmap = normalize_columns(headers)
        required = ["IP", "TFTP", "IMAGE", "CISCO_MD5"]
        missing = [r for r in required if r not in colmap]
        if missing:
            raise ValueError(
                "Planilha precisa conter as colunas (IP, TFTP, IMAGE, Cisco MD5). Não reconhecidas: "
                + ", ".join(missing)
                + ". Cabeçalhos encontrados: "
                + ", ".join(headers)
            )
        rows: List[Dict[str, str]] = []
        for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
            vals = [str(c.value).strip() if c.value is not None else "" for c in r]
            h2v = {headers[idx]: vals[idx] for idx in range(min(len(headers), len(vals)))}
            ip    = h2v.get(colmap["IP"], "").strip()
            tftp  = h2v.get(colmap["TFTP"], "").strip()
            image = h2v.get(colmap["IMAGE"], "").strip()
            md5   = (h2v.get(colmap.get("MD5", ""), "").strip()).lower() if "MD5" in colmap else ""
            cisco = (h2v.get(colmap["CISCO_MD5"], "").strip()).lower()
            iosst = h2v.get(colmap.get("IOS_STATUS", ""), "").strip() if "IOS_STATUS" in colmap else ""
            if not ip:
                continue
            rows.append({"_row": i, "IP": ip, "TFTP": tftp, "IMAGE": image, "MD5": md5, "CISCO_MD5": cisco, "IOS_STATUS": iosst, "Nome": h2v.get("Nome", "").strip()})
        return rows, colmap

# -------------------------
#  Atualização do MD5/Status na planilha
# -------------------------

def update_md5_in_xlsx(xlsx_path: str, updates: List[Tuple[int, str]], md5_header_name: str):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    md5_col_idx = None
    for idx, h in enumerate(headers, start=1):
        if h.strip().lower() == md5_header_name.strip().lower():
            md5_col_idx = idx
            break
    if md5_col_idx is None:
        for idx, h in enumerate(headers, start=1):
            if h.strip().lower() == "md5":
                md5_col_idx = idx
                break
    if md5_col_idx is None:
        md5_col_idx = len(headers) + 1
        ws.cell(row=1, column=md5_col_idx, value="MD5")
    for row_num, md5v in updates:
        ws.cell(row=row_num, column=md5_col_idx, value=(md5v or "").lower())
    wb.save(xlsx_path)

def update_status_in_xlsx(xlsx_path: str, updates: List[Tuple[int, str]], status_header_name: str):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    status_col_idx = None
    targets = {status_header_name.strip().lower(), "ios status", "ios_status", "status"}
    for idx, h in enumerate(headers, start=1):
        if h and h.strip().lower() in targets:
            status_col_idx = idx
            break
    if status_col_idx is None:
        status_col_idx = len(headers) + 1
        ws.cell(row=1, column=status_col_idx, value=(status_header_name or "IOS Status"))
    for row_num, val in updates:
        ws.cell(row=row_num, column=status_col_idx, value=(val or "").upper())
    wb.save(xlsx_path)

# -------------------------
#   Fase 1: Download (marca IOS Status imediatamente)
# -------------------------

def attempt_download(row: Dict[str, str], username: str, password: str, secret: str | None, logger: logging.Logger) -> tuple:
    nome = row.get("Nome") or row.get("IP") or "Unknown"
    ip   = row.get("IP")
    tftp = (row.get("TFTP") or "").strip()
    image= (row.get("IMAGE") or "").strip()

    device = {
        "device_type":  "cisco_ios",
        "host":         ip,
        "username":     username,
        "password":     password,
        "timeout":      CONNECT_TIMEOUT,
        "secret":       secret or "",
        "fast_cli":     False,
    }

    # Nova mensagem de início por device
    logger.info(f"Iniciando download no switch com o IP {ip}")

    if not image or not tftp:
        return nome, ip, row.get("_row"), False, "parâmetros inválidos (image/tftp)"

    for attempt in range(1, RETRY_COUNT + 1):
        try:
            conn = ConnectHandler(**device)
            _enable_if_needed(conn, secret, logger)
            allowed = {"sdflash:", "sdflash0:", "sdflash1:"}
            storage = STORAGE_PREF if STORAGE_PREF in allowed else _detect_storage(conn)
            if storage not in allowed:
                try:
                    conn.disconnect()
                except Exception:
                    pass
                return nome, ip, row.get("_row"), False, "sdflash indisponível"

            if _image_exists(conn, storage, image):
                logger.info(f"{nome} ({ip}) ➜ imagem já existe em {storage}{image}.")
                try:
                    conn.disconnect()
                except Exception:
                    pass
                return nome, ip, row.get("_row"), True, "já existia"

            logger.info(f"{nome} ({ip}) ➜ copiando {image} de {tftp} para {storage}...")
            okc, outc = tftp_copy_image(conn, tftp, image, storage, logger)
            try:
                conn.disconnect()
            except Exception:
                pass
            if okc:
                return nome, ip, row.get("_row"), True, "copiado"
            else:
                if attempt < RETRY_COUNT:
                    time.sleep(RETRY_DELAY)
                    continue
                return nome, ip, row.get("_row"), False, "erro no copy"
        except Exception as e:
            tb = traceback.format_exc()
            logger.debug(f"[{nome} - tentativa {attempt}] erro download: {e}\n{tb}")
            if attempt < RETRY_COUNT:
                time.sleep(RETRY_DELAY)
                continue
            return nome, ip, row.get("_row"), False, f"exceção: {e}"

# -------------------------
#   Fase 2: MD5 + Boot + Reload (para quem baixou OK)
# -------------------------

def post_download_pipeline(row: Dict[str, str], username: str, password: str, secret: str | None, logger: logging.Logger) -> tuple:
    nome = row.get("Nome") or row.get("IP") or "Unknown"
    ip   = row.get("IP")
    image= (row.get("IMAGE") or "").strip()
    c_md5= (row.get("CISCO_MD5") or "").strip().lower()

    device = {
        "device_type":  "cisco_ios",
        "host":         ip,
        "username":     username,
        "password":     password,
        "timeout":      CONNECT_TIMEOUT,
        "secret":       secret or "",
        "fast_cli":     False,
    }

    import re
    if not (c_md5 and re.fullmatch(r"[a-f0-9]{32}", c_md5)):
        return nome, ip, None, "NOK: Cisco MD5 ausente/inválido"

    for attempt in range(1, RETRY_COUNT + 1):
        try:
            conn = ConnectHandler(**device)
            _enable_if_needed(conn, secret, logger)
            allowed = {"sdflash:", "sdflash0:", "sdflash1:"}
            storage = STORAGE_PREF if STORAGE_PREF in allowed else _detect_storage(conn)
            if storage not in allowed:
                try:
                    conn.disconnect()
                except Exception:
                    pass
                return nome, ip, None, "NOK: sdflash não disponível"

            if not _image_exists(conn, storage, image):
                try:
                    conn.disconnect()
                except Exception:
                    pass
                return nome, ip, None, f"NOK: imagem ausente em {storage}{image}"

            okv, calc, _ = verify_md5(conn, storage, image, logger)
            if not okv or not calc:
                try:
                    conn.disconnect()
                except Exception:
                    pass
                return nome, ip, None, "NOK: erro ao calcular MD5"
            if calc != c_md5:
                logger.info(f"Switch {ip} MD5 Status NOK")
                try:
                    conn.disconnect()
                except Exception:
                    pass
                return nome, ip, calc, "NOK: MD5 != Cisco MD5"
            else:
                logger.info(f"Switch {ip} MD5 Status OK")

            # Mensagem antes de setar o boot
            logger.info(f"Setando boot sdflash {image} no switch {ip}")
            okboot, msgboot = _set_boot_to_bin(conn, storage, image, logger)
            if not okboot:
                try:
                    conn.disconnect()
                except Exception:
                    pass
                return nome, ip, calc, f"NOK: boot .bin ({msgboot})"
            else:
                logger.info("Arquivo setado com sucesso")

            # Mensagem de reload
            logger.info(f"Reiniciando switch {ip}")
            try:
                _send_reload_with_confirms(conn)
            except Exception:
                pass
            try:
                conn.disconnect()
            except Exception:
                pass

            if not WAIT_RELOAD:
                return nome, ip, calc, "OK: reload enviado"

            ok, new_conn, reason = _await_reload_and_reconnect(device, logger)
            if not ok or not new_conn:
                return nome, ip, calc, f"NOK: pós-reload {reason}"
            try:
                sb = new_conn.send_command("show boot", expect_string=r"#", delay_factor=3)
            except Exception:
                sb = ""
            try:
                new_conn.disconnect()
            except Exception:
                pass

            if image.lower() in (sb or "").lower():
                return nome, ip, calc, "OK"
            else:
                return nome, ip, calc, "NOK: show boot não refletiu o .bin"

        except Exception as e:
            tb = traceback.format_exc()
            logger.debug(f"[{nome} - tentativa {attempt}] erro pós-download: {e}\n{tb}")
            if attempt < RETRY_COUNT:
                time.sleep(RETRY_DELAY)
                continue
            return nome, ip, None, f"NOK: {e}"

# -------------------------
#             Main
# -------------------------

def main():
    # Credenciais via .env (opcional) ou prompt
    username = os.getenv("SSH_USERNAME") or input("👤 Usuário: ")
    password = os.getenv("SSH_PASSWORD") or getpass.getpass("🔒 Senha: ")
    enable_secret = os.getenv("SSH_ENABLE_SECRET") or None

    xlsx_path = _discover_xlsx_path()
    if not os.path.isfile(xlsx_path):
        print(f"❌ Planilha não encontrada: {xlsx_path}")
        sys.exit(1)

    try:
        rows, colmap = read_plan_xlsx(xlsx_path)
        if not rows:
            raise ValueError("Nenhum device encontrado na planilha.")
    except Exception as e:
        print(f"❌ {e}")
        sys.exit(1)

    # Teste rápido de credenciais
    first = rows[0]
    print(f"\n🔍 Testando credenciais em {first.get('IP')}...")
    device_test = {
        "device_type":  "cisco_ios",
        "host":         first.get("IP"),
        "username":     username,
        "password":     password,
        "timeout":      CONNECT_TIMEOUT,
        "secret":       enable_secret or "",
    }
    try:
        conn = ConnectHandler(**device_test)
        _enable_if_needed(conn, enable_secret, None)
        conn.send_command("show clock", expect_string=r"#", delay_factor=1)
        conn.disconnect()
        print("✅ Credenciais válidas. Iniciando processo completo...\n")
    except Exception:
        print("❌ Credenciais inválidas ou privilégios insuficientes. Abortando script.")
        sys.exit(1)

    logger = setup_logger()

    # -----------------
    # FASE 1: Download (marca IOS Status imediatamente)
    # -----------------
    print("▶️  Fase 1: Download e marcação de IOS Status (OK/NOK)")
    status_header_name = colmap.get("IOS_STATUS", "IOS Status")
    ok_rows: List[Dict[str, str]] = []
    nok_rows: List[Dict[str, str]] = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(attempt_download, r, username, password, enable_secret, logger): r for r in rows}
        for future in tqdm(as_completed(futures), total=len(futures), desc="⬇️  Download", ncols=100):
            nome, ip, row_num, okd, msg = future.result()
            # Mensagens de resultado do download por device
            if okd:
                tqdm.write(f"Switch {ip} IOS OK")
            else:
                tqdm.write(f"Switch {ip} IOS NOK, indo para o próximo")
            row_n = int(row_num or 0)
            # Atualiza IOS Status imediatamente
            try:
                update_status_in_xlsx(xlsx_path, [(row_n, "OK" if okd else "NOK")], status_header_name)
            except Exception as e:
                print(f"⚠️  Falha ao atualizar IOS Status para {ip}: {e}")
            orig_row = futures[future]
            if okd:
                ok_rows.append(orig_row)
            else:
                nok_rows.append(orig_row)

    # -----------------
    # FASE 1B: Retry único de download para NOK
    # -----------------
    if nok_rows:
        print("🔁 Retry: tentando novamente download para os marcados como NOK")
        still_nok: List[Dict[str, str]] = []
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(attempt_download, r, username, password, enable_secret, logger): r for r in nok_rows}
            for future in tqdm(as_completed(futures), total=len(futures), desc="↩️  Retry DL", ncols=100):
                nome, ip, row_num, okd, msg = future.result()
                if okd:
                    tqdm.write(f"Switch {ip} IOS OK (retry)")
                else:
                    tqdm.write(f"Switch {ip} IOS NOK (retry), indo para o próximo")
                row_n = int(row_num or 0)
                try:
                    update_status_in_xlsx(xlsx_path, [(row_n, "OK" if okd else "NOK")], status_header_name)
                except Exception as e:
                    print(f"⚠️  Falha ao atualizar IOS Status (retry) para {ip}: {e}")
                orig_row = futures[future]
                if okd:
                    ok_rows.append(orig_row)
                else:
                    still_nok.append(orig_row)
        nok_rows = still_nok

    # -----------------
    # FASE 2: Para os com download OK -> MD5 + boot + reload
    # -----------------
    print("⚙️  Fase 2: MD5 + comparação Cisco MD5 + set boot + reload")
    sucesso = 0
    falha  = 0
    md5_updates: List[Tuple[int, str]] = []

    if ok_rows:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(post_download_pipeline, r, username, password, enable_secret, logger): r for r in ok_rows}
            for future in tqdm(as_completed(futures), total=len(futures), desc="⚙️  Pós-DL", ncols=100):
                nome, ip, md5_calc, resultado = future.result()
                tqdm.write(f"{nome} ({ip}) ➜ {resultado}")
                orig_row = futures[future]
                row_n = int(orig_row.get("_row") or 0)
                if md5_calc:
                    md5_updates.append((row_n, md5_calc))
                if str(resultado).startswith("OK"):
                    sucesso += 1
                else:
                    falha += 1

    # Escrever MD5s ao final
    try:
        if md5_updates:
            md5_header_name = colmap.get("MD5", "MD5")
            md5_updates = [(r, m) for (r, m) in md5_updates if r and isinstance(r, int) and m]
            update_md5_in_xlsx(xlsx_path, md5_updates, md5_header_name)
            print(f"📝 Planilha atualizada: MD5 ({len(md5_updates)})")
    except Exception as e:
        print(f"⚠️  Falha ao escrever MD5 na planilha: {e}")

    print("\n📊 Concluído!")
    print(f"✅ Pós-download OK (pipeline final): {sucesso}")
    print(f"❌ Pós-download falhas (pipeline final): {falha}")
    if nok_rows:
        print(f"ℹ️  Dispositivos que permaneceram com download NOK: {len(nok_rows)}")


if __name__ == "__main__":
    main()
