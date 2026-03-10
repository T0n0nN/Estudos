#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from netmiko import ConnectHandler
import os
import sys
import time
import getpass
import logging
import traceback
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Tuple, List, Dict
from dotenv import load_dotenv
from tqdm import tqdm
import socket

# -------------------------
# Carregar .env (raiz do workspace e pasta atual)
# -------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)
load_dotenv(os.path.join(ROOT_DIR, ".env"))
load_dotenv(os.path.join(SCRIPT_DIR, ".env"))

# -------------------------
#          Configuráveis
# -------------------------
# Planilha com o plano de upgrade
DEFAULT_XLSX_PATH = os.path.join(SCRIPT_DIR, "Switches IOS Upgrade.xlsx")
XLSX_FILE_ENV     = os.getenv("IOS_FULL_XLSX", DEFAULT_XLSX_PATH).strip()
# MAX_WORKERS respeita env, mas nunca ultrapassa 4 (e no mínimo 1)
MAX_WORKERS       = min(4, max(1, int(os.getenv("IOS_MAX_WORKERS", "4"))))
CONNECT_TIMEOUT   = int(os.getenv("IOS_CONNECT_TIMEOUT", "10"))
RETRY_COUNT       = int(os.getenv("IOS_RETRY_COUNT", "2"))
RETRY_DELAY       = int(os.getenv("IOS_RETRY_DELAY", "5"))
# Storage preferido
STORAGE_PREF      = os.getenv("IOS_STORAGE", "flash:").strip().lower()
# Fail-fast global
FAIL_FAST         = (os.getenv("IOS_FAIL_FAST", "1").strip().lower() in ("1", "true", "yes"))
# Aguardar reload/validação
WAIT_RELOAD       = (os.getenv("IOS_WAIT_RELOAD", "1").strip().lower() in ("1", "true", "yes"))
DOWN_WAIT_SEC     = int(os.getenv("IOS_DOWN_WAIT_SEC", "600"))
UP_WAIT_SEC       = int(os.getenv("IOS_UP_WAIT_SEC", "1800"))
POLL_INTERVAL     = int(os.getenv("IOS_POLL_INTERVAL", "10"))

CANCEL_EVENT      = None  # definido em main

# -------------------------
# Logger
# -------------------------

def setup_logger() -> Tuple[logging.Logger, str]:
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_filename = f"resultado_instalacao_{ts}.log"

    logger = logging.getLogger("IOSFull")
    logger.setLevel(logging.DEBUG)

    # evitar duplicação de handlers
    logger.handlers.clear()

    fh = logging.FileHandler(log_filename, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s: %(message)s"))
    logger.addHandler(fh)

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(ch)

    return logger, log_filename

# -------------------------
# Utilidades
# -------------------------

def _enable_if_needed(conn, secret: str | None, logger: logging.Logger | None = None):
    try:
        if secret:
            conn.enable()
        try:
            conn.send_command("terminal length 0", expect_string=r"#", delay_factor=2)
        except Exception:
            pass
    except Exception as e:
        if logger:
            logger.debug(f"Falha ao entrar em enable: {e}")


def _detect_storage(conn) -> str:
    try:
        out = conn.send_command("dir flash:", expect_string=r"#", delay_factor=2)
        if "Directory of" in out or "bytes" in out.lower():
            return "flash:"
    except Exception:
        pass
    try:
        out = conn.send_command("dir bootflash:", expect_string=r"#", delay_factor=2)
        if "Directory of" in out or "bytes" in out.lower():
            return "bootflash:"
    except Exception:
        pass
    return "flash:"


def _image_exists(conn, storage: str, image_name: str) -> bool:
    try:
        out = conn.send_command(f"dir {storage}{image_name}", expect_string=r"#", delay_factor=3)
        return image_name in out and ("No such file" not in out and "Error" not in out)
    except Exception:
        return False


def _is_install_mode(conn) -> bool:
    try:
        cmds = [
            "show install summary | i [Oo]perating|Mode|mode",
            "show install summary",
            "show version | i Mode|INSTALL|BUNDLE",
            "show version",
        ]
        for cmd in cmds:
            try:
                out = conn.send_command(cmd, expect_string=r"#", delay_factor=2)
            except Exception:
                continue
            ol = (out or "").lower()
            if "install" in ol and "bundle" not in ol:
                return True
            if "bundle" in ol and "install" not in ol:
                return False
        return False
    except Exception:
        return False


def _running_boot_lines(conn) -> List[str]:
    try:
        out = conn.send_command("show run | i ^boot system", expect_string=r"#", delay_factor=2)
    except Exception:
        out = ""
    return [ln.strip() for ln in (out or "").splitlines() if ln and ln.strip().startswith("boot system")]


def _ensure_boot_points_to_packages(conn, storage: str, logger: logging.Logger) -> Tuple[bool, str]:
    try:
        sb = conn.send_command("show boot", expect_string=r"#", delay_factor=3)
    except Exception:
        sb = ""
    try:
        sr_lines = _running_boot_lines(conn)
    except Exception:
        sr_lines = []
    text = ((sb or "") + "\n" + "\n".join(sr_lines)).lower()
    if "packages.conf" in text:
        return True, "boot já aponta para packages.conf"

    try:
        conn.config_mode()
    except Exception:
        pass

    for ln in list(sr_lines):
        try:
            conn.send_command_timing("no " + ln, delay_factor=3, max_loops=200)
        except Exception:
            pass

    boot_path = (storage if storage else "bootflash:") + "packages.conf"
    try:
        conn.send_command_timing(f"boot system {boot_path}", delay_factor=3, max_loops=200)
    except Exception:
        pass

    try:
        if conn.check_config_mode():
            conn.exit_config_mode()
    except Exception:
        pass

    try:
        w = conn.send_command_timing("write memory", delay_factor=5, max_loops=200)
        wl = (w or "").lower()
        if "[confirm]" in wl:
            conn.send_command_timing("\n", delay_factor=3, max_loops=100)
        elif "[y/n]" in wl:
            conn.send_command_timing("y", delay_factor=3, max_loops=100)
        elif "[yes/no]" in wl:
            conn.send_command_timing("yes", delay_factor=3, max_loops=100)
    except Exception:
        pass

    try:
        sb2 = conn.send_command("show boot", expect_string=r"#", delay_factor=3)
    except Exception:
        sb2 = ""
    try:
        sr2 = conn.send_command("show run | i ^boot system", expect_string=r"#", delay_factor=2)
    except Exception:
        sr2 = ""
    ok = ("packages.conf" in (sb2 or "").lower()) or ("packages.conf" in (sr2 or "").lower())
    return ok, ("OK" if ok else "não foi possível confirmar no show boot")


def _run_long_operation(conn, cmd: str, max_seconds: int = 7200, logger: logging.Logger | None = None) -> str:
    out = conn.send_command_timing(cmd, delay_factor=12, max_loops=2000, strip_command=False, strip_prompt=False)
    start = time.time()
    while time.time() - start < max_seconds:
        low = (out or "").lower()
        if "[yes/no]" in low or " yes/no]" in low:
            out += "\n" + conn.send_command_timing("yes", delay_factor=12, max_loops=2000)
            continue
        if "[y/n]" in low or " y/n]" in low or ("proceed" in low and "[y/n]" in low):
            out += "\n" + conn.send_command_timing("y", delay_factor=12, max_loops=2000)
            continue
        if "[confirm]" in low:
            out += "\n" + conn.send_command_timing("\n", delay_factor=12, max_loops=2000)
            continue
        if "save? [yes/no]" in low or "save the configuration" in low:
            out += "\n" + conn.send_command_timing("no", delay_factor=12, max_loops=2000)
            continue
        if "press return to continue" in low or "press enter to continue" in low:
            out += "\n" + conn.send_command_timing("\n", delay_factor=12, max_loops=2000)
            continue
        try:
            more = conn.read_channel()
        except Exception:
            more = ""
        if more:
            out += more
            time.sleep(2)
            continue
        if any(k in low for k in [
            "install add operation successful",
            "finished install operation",
            "activating software",
            "switch will be reloaded",
            "system will be reloaded",
            "this operation will reload the system",
            "reload command is being issued",
            "chassis will be rebooted",
            "rebooting",
            "packages will be activated at next reload",
        ]):
            break
        time.sleep(3)
    if logger:
        try:
            logger.debug(f"Saída acumulada de '{cmd}':\n{out}")
        except Exception:
            pass
    return out


def _send_reload_with_confirms(conn) -> str:
    r = conn.send_command_timing("reload", delay_factor=5, max_loops=200)
    rl = (r or '').lower()
    if "[confirm]" in rl:
        rl += "\n" + conn.send_command_timing("\n", delay_factor=5, max_loops=200)
    if "[y/n]" in rl:
        rl += "\n" + conn.send_command_timing("y", delay_factor=5, max_loops=200)
    if "[yes/no]" in rl:
        rl += "\n" + conn.send_command_timing("yes", delay_factor=5, max_loops=200)
    if "save? [yes/no]" in rl or "save the configuration" in rl:
        rl += "\n" + conn.send_command_timing("no", delay_factor=5, max_loops=200)
    return rl


def _tcp_can_connect(host: str, port: int = 22, timeout: int = 3) -> bool:
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except Exception:
        return False


def _await_reload_and_reconnect(device: dict, logger: logging.Logger) -> Tuple[bool, object, str]:
    host = device.get("host")
    start = time.time()
    saw_down = False
    while time.time() - start < DOWN_WAIT_SEC:
        if not _tcp_can_connect(host):
            saw_down = True
            break
        time.sleep(POLL_INTERVAL)
    if not saw_down:
        logger.debug(f"{host} ➜ SSH não ficou indisponível dentro do tempo; seguindo para reconexão.")
    start = time.time()
    while time.time() - start < UP_WAIT_SEC:
        if _tcp_can_connect(host):
            try:
                conn = ConnectHandler(**device)
                return True, conn, "Reconectado"
            except Exception:
                pass
        time.sleep(POLL_INTERVAL)
    return False, None, "Não reconectou dentro do tempo limite"


def _committed_version_matches(conn, expected_prefix: str | None) -> Tuple[bool, str]:
    try:
        out = conn.send_command("show install summary", expect_string=r"#", delay_factor=3)
        lines = (out or "").splitlines()
        committed = []
        for ln in lines:
            if "IMG" in ln and " C " in ln:
                parts = ln.split()
                if parts:
                    committed.append(parts[-1])
        if not committed:
            return False, "Nenhuma versão 'C' encontrada"
        if expected_prefix:
            ok = any(v.startswith(expected_prefix) for v in committed)
            return ok, ", ".join(committed)
        return True, ", ".join(committed)
    except Exception as e:
        return False, str(e)


def _extract_version_prefix_from_image(name: str) -> str | None:
    import re
    m = re.search(r"(\d+\.\d+\.\d+)", name)
    return m.group(1) if m else None

# -------------------------
# TFTP e MD5
# -------------------------

def tftp_copy_image(conn, tftp_server: str, image_name: str, storage: str, logger: logging.Logger) -> Tuple[bool, str]:
    """Copia via TFTP usando a forma com URL completa para reduzir prompts."""
    dest = f"{storage}{image_name}"
    cmd = f"copy tftp://{tftp_server}/{image_name} {dest}"
    out = conn.send_command_timing(cmd, delay_factor=8, max_loops=1500)
    for _ in range(50):
        low = (out or '').lower()
        if "destination filename" in low or "overwrite" in low:
            out += "\n" + conn.send_command_timing("\n", delay_factor=8, max_loops=1500)
            continue
        if any(x in low for x in ["error", "timed out", "no space", "access violation", "not found", "%invalid"]):
            break
        # ler mais
        more = ""
        try:
            more = conn.read_channel()
        except Exception:
            pass
        if more:
            out += more
            time.sleep(1)
            continue
        # heurística de término (bytes copiados)
        if any(k in low for k in ["copied", "bytes copied", "bytes/sec", "bytes transferred"]):
            break
        time.sleep(2)
    logger.debug(f"Saída do copy: \n{out}")
    low = (out or '').lower()
    ok = any(k in low for k in ["copied", "bytes copied", "bytes/sec", "transfer completed"]) and not any(k in low for k in ["error", "no space", "timed out", "access violation", "%invalid"])
    return ok, out


def verify_md5(conn, storage: str, image_name: str, logger: logging.Logger) -> Tuple[bool, str, str]:
    """Retorna (ok, md5_calc, output)."""
    cmd = f"verify /md5 {storage}{image_name}"
    out = conn.send_command_timing(cmd, delay_factor=12, max_loops=4000)
    # capturar md5
    import re
    md5_match = re.search(r"=\s*([a-fA-F0-9]{32})", out or "")
    if not md5_match:
        # alguns IOS imprimem "Verified (md5) flash:file = <hash>"
        md5_match = re.search(r"\b([a-fA-F0-9]{32})\b", out or "")
    calc = md5_match.group(1).lower() if md5_match else ""
    ok = bool(calc)
    logger.debug(f"Saída do verify md5: \n{out}")
    return ok, calc, out

# -------------------------
# Leitura da planilha
# -------------------------

def read_plan_xlsx(xlsx_path: str) -> List[Dict[str, str]]:
    def normalize_key(s: str) -> str:
        s = (s or "").strip().lower()
        s = s.replace("_", " ").replace("-", " ")
        s = " ".join(s.split())
        return s

    def normalize_columns(cols: List[str]) -> Dict[str, str]:
        # mapa de sinônimos -> alvo (chaves normalizadas)
        base_synonyms = {
            "ip": "IP",
            "endereco": "IP",
            "endereço ip": "IP",
            "endereco ip": "IP",
            "address": "IP",
            "host": "IP",
            "tftp": "TFTP",
            "tftp server": "TFTP",
            "server": "TFTP",
            "tftp ip": "TFTP",
            "tftp server ip": "TFTP",
            "imagem": "IMAGE",
            "image": "IMAGE",
            "image filename": "IMAGE",
            "arquivo": "IMAGE",
            "bin": "IMAGE",
            "md5": "MD5",
            "md5sum": "MD5",
            "hash": "MD5",
        }
        synonyms = {normalize_key(k): v for k, v in base_synonyms.items()}
        out: Dict[str, str] = {}
        for c in cols:
            if not isinstance(c, str):
                continue
            key = c.strip()
            low = normalize_key(key)
            if low in synonyms:
                out[synonyms[low]] = key
            else:
                out[key] = key
        return out

    try:
        import pandas as pd
        df = pd.read_excel(xlsx_path)
        colmap = normalize_columns([c for c in df.columns])
        required = ["IP", "TFTP", "IMAGE", "MD5"]
        for r in required:
            if r not in colmap:
                raise ValueError(f"Planilha precisa conter as colunas: IP, TFTP, IMAGE, MD5. Encontrado: {', '.join(df.columns)}")
        df = df.fillna("")
        recs = []
        for _, row in df.iterrows():
            ip    = str(row[colmap["IP"]]).strip()
            tftp  = str(row[colmap["TFTP"]]).strip()
            image = str(row[colmap["IMAGE"]]).strip()
            md5   = str(row[colmap["MD5"]]).strip().lower()
            if not ip:
                continue
            recs.append({"IP": ip, "TFTP": tftp, "IMAGE": image, "MD5": md5, "Nome": str(row.get("Nome", "")).strip()})
        # validações básicas
        missing = [str(i+2) for i, r in enumerate(recs) if not r.get("IMAGE") or not r.get("TFTP") or not r.get("MD5")]
        if missing:
            raise ValueError("Linhas com IP mas sem TFTP/IMAGE/MD5: " + ", ".join(missing))
        return recs
    except ImportError:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # construir colmap simples
        m = normalize_columns(headers)
        required = ["IP", "TFTP", "IMAGE", "MD5"]
        for r in required:
            if r not in m:
                raise ValueError(f"Planilha precisa conter as colunas: IP, TFTP, IMAGE, MD5. Encontrado: {', '.join(headers)}")
        recs: List[Dict[str, str]] = []
        for r in ws.iter_rows(min_row=2):
            vals = [str(c.value).strip() if c.value is not None else "" for c in r]
            h2v = {headers[i]: vals[i] for i in range(min(len(headers), len(vals)))}
            ip    = h2v.get(m["IP"], "").strip()
            tftp  = h2v.get(m["TFTP"], "").strip()
            image = h2v.get(m["IMAGE"], "").strip()
            md5   = (h2v.get(m["MD5"], "").strip()).lower()
            if not ip:
                continue
            recs.append({"IP": ip, "TFTP": tftp, "IMAGE": image, "MD5": md5, "Nome": h2v.get("Nome", "").strip()})
        missing = [str(i+2) for i, r in enumerate(recs) if not r.get("IMAGE") or not r.get("TFTP") or not r.get("MD5")]
        if missing:
            raise ValueError("Linhas com IP mas sem TFTP/IMAGE/MD5: " + ", ".join(missing))
        return recs

# -------------------------
# Upgrade por dispositivo
# -------------------------

def do_upgrade(row: Dict[str, str], username: str, password: str, secret: str | None, logger: logging.Logger) -> tuple:
    nome = row.get("Nome") or row.get("IP") or "Unknown"
    ip   = row.get("IP")
    tftp = row.get("TFTP")
    image= row.get("IMAGE")
    md5  = (row.get("MD5") or "").lower()

    device = {
        "device_type":  "cisco_xe",
        "host":         ip,
        "username":     username,
        "password":     password,
        "timeout":      CONNECT_TIMEOUT,
        "secret":       secret or "",
        "fast_cli":     False,
    }

    expected_prefix = _extract_version_prefix_from_image(image)

    for attempt in range(1, RETRY_COUNT + 1):
        try:
            if CANCEL_EVENT and CANCEL_EVENT.is_set():
                return nome, ip, "CANCELADO (fail-fast)"
            # Conectar
            conn = ConnectHandler(**device)
            _enable_if_needed(conn, secret, logger)

            # Verificar INSTALL
            if not _is_install_mode(conn):
                try:
                    conn.disconnect()
                except Exception:
                    pass
                msg = "Não está em modo INSTALL. Abortando."
                logger.error(f"{nome} ({ip}) ➜ {msg}")
                return nome, ip, f"ERRO: {msg}"

            # Storage
            storage = STORAGE_PREF if STORAGE_PREF in ("flash:", "bootflash:") else _detect_storage(conn)

            # Download se necessário ou se md5 não confere
            need_copy = True
            if _image_exists(conn, storage, image):
                ok, calc, _ = verify_md5(conn, storage, image, logger)
                if ok and calc == md5:
                    need_copy = False
                else:
                    need_copy = True
            if need_copy:
                logger.info(f"{nome} ({ip}) ➜ copiando imagem via TFTP...")
                okc, outc = tftp_copy_image(conn, tftp, image, storage, logger)
                if not okc:
                    try:
                        conn.disconnect()
                    except Exception:
                        pass
                    logger.error(f"{nome} ({ip}) ➜ ERRO no copy TFTP")
                    return nome, ip, "ERRO: copy tftp"
                # Validar MD5
                okv, calc, _ = verify_md5(conn, storage, image, logger)
                if not okv or calc != md5:
                    try:
                        conn.disconnect()
                    except Exception:
                        pass
                    logger.error(f"{nome} ({ip}) ➜ MD5 divergente (calc={calc}, planilha={md5})")
                    return nome, ip, "ERRO: MD5 divergente"
            else:
                logger.info(f"{nome} ({ip}) ➜ imagem já presente e MD5 ok")

            # Garantir boot packages.conf
            okboot, msgboot = _ensure_boot_points_to_packages(conn, storage, logger)
            if not okboot:
                try:
                    conn.disconnect()
                except Exception:
                    pass
                logger.error(f"{nome} ({ip}) ➜ ERRO ao ajustar boot packages.conf: {msgboot}")
                return nome, ip, f"ERRO: boot packages.conf ({msgboot})"

            # Executar install one-liner
            cmd  = f"install add file {storage}{image} activate commit"
            out = _run_long_operation(conn, cmd, max_seconds=7200, logger=logger)
            lower = (out or "").lower()

            # Decidir reload
            will_reload = any(m in lower for m in [
                "switch will be reloaded", "system will be reloaded", "this operation will reload the system",
                "reload command is being issued", "chassis will be rebooted", "rebooting"
            ])

            if WAIT_RELOAD and will_reload:
                try:
                    conn.disconnect()
                except Exception:
                    pass
                logger.info(f"{nome} ({ip}) ➜ aguardando reload e validação...")
                ok, new_conn, reason = _await_reload_and_reconnect(device, logger)
                if not ok or not new_conn:
                    logger.error(f"{nome} ({ip}) ➜ ERRO pós-reload: {reason}")
                    return nome, ip, f"ERRO: {reason}"
                _enable_if_needed(new_conn, secret, logger)
                # Validar INSTALL e versão
                if not _is_install_mode(new_conn):
                    try:
                        new_conn.disconnect()
                    except Exception:
                        pass
                    return nome, ip, "ERRO: pós-reload não está em INSTALL"
                okv, info = _committed_version_matches(new_conn, expected_prefix)
                try:
                    new_conn.disconnect()
                except Exception:
                    pass
                if okv:
                    logger.info(f"{nome} ({ip}) ➜ SUCESSO (validado): {info}")
                    return nome, ip, "SUCESSO"
                else:
                    logger.error(f"{nome} ({ip}) ➜ ERRO validação versão: {info}")
                    return nome, ip, f"ERRO: validação versão ({info})"
            else:
                # Sem espera de reload, considerar enviado
                try:
                    conn.disconnect()
                except Exception:
                    pass
                if WAIT_RELOAD and not will_reload:
                    logger.error(f"{nome} ({ip}) ➜ Nenhum indicador de reload após install")
                    return nome, ip, "ERRO: reload não disparou"
                logger.info(f"{nome} ({ip}) ➜ SUCESSO (comando enviado)")
                return nome, ip, "SUCESSO"

        except Exception as e:
            tb = traceback.format_exc()
            logger.debug(f"[{nome} - tentativa {attempt}] erro: {e}\n{tb}")
            if attempt < RETRY_COUNT:
                time.sleep(RETRY_DELAY)
                continue
            else:
                logger.error(f"{nome} ({ip}) ➜ ERRO: {e}")
                return nome, ip, f"ERRO: {e}"

# -------------------------
# Main
# -------------------------

def main():
    global CANCEL_EVENT
    CANCEL_EVENT = threading.Event()

    username = os.getenv("SSH_USERNAME") or input("👤 Usuário: ")
    password = os.getenv("SSH_PASSWORD") or getpass.getpass("🔒 Senha: ")
    enable_secret = os.getenv("SSH_ENABLE_SECRET") or None

    xlsx_path = XLSX_FILE_ENV
    if not os.path.isfile(xlsx_path):
        print(f"❌ Planilha não encontrada: {xlsx_path}")
        sys.exit(1)

    try:
        rows = read_plan_xlsx(xlsx_path)
        if not rows:
            raise ValueError("Nenhum switch encontrado na planilha.")
    except Exception as e:
        print(f"❌ {e}")
        sys.exit(1)

    # Teste rápido de credenciais
    first = rows[0]
    print(f"\n🔍 Testando credenciais em {first.get('IP')}...")
    device_test = {
        "device_type":  "cisco_xe",
        "host":         first.get("IP"),
        "username":     username,
        "password":     password,
        "timeout":      CONNECT_TIMEOUT,
        "secret":       enable_secret or "",
    }
    try:
        conn = ConnectHandler(**device_test)
        _enable_if_needed(conn, enable_secret, None)
        conn.send_command("show clock", expect_string=r"#", delay_factor=1)
        conn.disconnect()
        print("✅ Credenciais válidas. Iniciando...")
    except Exception:
        print("❌ Credenciais inválidas ou privilégios insuficientes. Abortando script.")
        sys.exit(1)

    logger, log_filename = setup_logger()
    logger.info("▶️  Início do Upgrade Completo (TFTP + MD5 + INSTALL)")

    sucesso = 0
    falha  = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(do_upgrade, r, username, password, enable_secret, logger): r for r in rows}
        for future in tqdm(as_completed(futures), total=len(futures), desc="🔄 Atualizando switches", ncols=100):
            nome, ip, resultado = future.result()
            tqdm.write(f"{nome} ({ip}) ➜ {resultado}")
            if "SUCESSO" in resultado:
                sucesso += 1
            elif "ERRO" in resultado:
                falha += 1
            if FAIL_FAST and "ERRO" in resultado and CANCEL_EVENT is not None:
                CANCEL_EVENT.set()

    print("\n📊 Concluído!")
    print(f"✅ Sucesso: {sucesso}")
    print(f"❌ Falhas: {falha}")
    print(f"📝 Log salvo em: {log_filename}")


if __name__ == "__main__":
    import threading
    main()
