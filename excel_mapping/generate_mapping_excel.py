import argparse
import json
import textwrap
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _set_col_width(ws, col_letter: str, width: float) -> None:
    ws.column_dimensions[col_letter].width = width


def _cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col)


def _smart_wrap_text(value: str, *, width: int = 16, max_lines: int = 3) -> str:
    """Insert '\n' for long strings.

    For rotated text, Excel's wrap_text can still overflow visually.
    Explicit newlines keep long labels inside the cell.
    """
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text or "\n" in text:
        return value
    if len(text) <= width:
        return value
    if " " not in text:
        return value

    lines = textwrap.wrap(text, width=width, break_long_words=False, break_on_hyphens=False)
    if not lines:
        return value
    return "\n".join(lines[:max_lines])


def _load_mapping_data(data_path: Path) -> dict:
    raw = json.loads(data_path.read_text(encoding="utf-8"))

    # Backward compatible format (single sheet)
    if "sheets" not in raw:
        sheet_name = raw.get("sheet_name") or raw.get("name") or "Cable 11 and 12"
        return {
            "sheets": {
                sheet_name: raw,
            }
        }

    return raw


def _normalize_sheet_cfg(sheet_cfg: dict) -> dict:
    left_letters = list(sheet_cfg.get("left_letters") or [])
    right_letters = list(sheet_cfg.get("right_letters") or [])
    extra_letters = list(sheet_cfg.get("extra_letters") or [])
    port6_letters = list(sheet_cfg.get("port6_letters") or [])
    port_column_after = sheet_cfg.get("port_column_after") or None
    port_column_title = sheet_cfg.get("port_column_title") or "PORT"
    port6_total_across_tethers = bool(sheet_cfg.get("port6_total_across_tethers") or False)
    port6_port_labels = bool(sheet_cfg.get("port6_port_labels") or False)
    port6_port_label_text = sheet_cfg.get("port6_port_label_text") or "Port"
    title_left = sheet_cfg.get("title_left", "")
    title_right = sheet_cfg.get("title_right", "")
    title_extra = sheet_cfg.get("title_extra", "")

    return {
        "title_left": title_left,
        "title_right": title_right,
        "title_extra": title_extra,
        "left_letters": left_letters,
        "right_letters": right_letters,
        "extra_letters": extra_letters,
        "port6_letters": port6_letters,
        "port_column_after": port_column_after,
        "port_column_title": port_column_title,
        "port6_total_across_tethers": port6_total_across_tethers,
        "port6_port_labels": port6_port_labels,
        "port6_port_label_text": port6_port_label_text,
        "tether1": sheet_cfg.get("tether1", {}) or {},
        "tether2": sheet_cfg.get("tether2", {}) or {},
    }


def _build_sheet(ws, sheet_cfg: dict) -> None:
    left_letters = sheet_cfg["left_letters"]
    right_letters = sheet_cfg["right_letters"]
    extra_letters = sheet_cfg.get("extra_letters") or []
    port6_letters = set(sheet_cfg.get("port6_letters") or [])
    letters = left_letters + right_letters + list(extra_letters)

    port_column_after = sheet_cfg.get("port_column_after")
    port_column_title = sheet_cfg.get("port_column_title") or "PORT"
    port6_total_across_tethers = bool(sheet_cfg.get("port6_total_across_tethers") or False)
    port6_port_labels = bool(sheet_cfg.get("port6_port_labels") or False)
    port6_port_label_text = sheet_cfg.get("port6_port_label_text") or "Port"
    insert_after_idx: int | None = None
    if port_column_after and port_column_after in letters:
        insert_after_idx = letters.index(port_column_after)

    def _port_col() -> int | None:
        if insert_after_idx is None:
            return None
        # Insert after the 2-column pair of the chosen letter.
        return first_letter_col + ((insert_after_idx + 1) * 2)

    def _letter_cols(i: int) -> tuple[int, int]:
        shift = 1 if (insert_after_idx is not None and i > insert_after_idx) else 0
        col1 = first_letter_col + (i * 2) + shift
        return col1, col1 + 1

    ws.title = ws.title or "Mapping"

    # Layout constants
    # Columns:
    # 1: PORT 1
    # 2: PORT 2
    # 3: Tether label
    # 4..: letter columns
    port1_col = 1
    port2_col = 2
    tether_col = 3
    first_letter_col = 4

    header_row = 1
    letters_row = 2

    # Each tether section: 6 port rows + 1 small Tap row
    tether1_start = 3
    tether1_ports_end = 8
    tether1_tap_row = 9

    tether2_start = 10
    tether2_ports_end = 15
    tether2_tap_row = 16

    # Styles
    hair = Side(style="hair", color="000000")
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Fills (match the gray tones from the reference print)
    header_fill = PatternFill("solid", fgColor="BFBFBF")
    subheader_fill = PatternFill("solid", fgColor="D9D9D9")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_no_wrap = Alignment(horizontal="center", vertical="center", wrap_text=False)
    center_shrink = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)
    center_rot90 = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    center_rot90_shrink = Alignment(
        horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True, text_rotation=90
    )
    top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)

    def _content_alignment(value: Any, *, rotation: int | None = None) -> Alignment:
        """Keep text inside cells without changing the layout.

        - If the value already contains explicit line breaks, keep wrap_text on.
        - Otherwise, prefer shrink_to_fit to avoid vertical clipping.
        """
        if isinstance(value, str) and "\n" in value:
            if rotation == 90:
                return center_rot90
            return center
        if rotation == 90:
            return center_rot90_shrink
        return center_shrink

    font_name = "Times New Roman"
    header_font = Font(name=font_name, bold=True, size=12)
    tether_font = Font(name=font_name, bold=True, size=14)

    # Column widths
    _set_col_width(ws, "A", 8)
    _set_col_width(ws, "B", 8)
    _set_col_width(ws, "C", 9)
    # Each letter uses two subcolumns (Tap X / Tap Y)
    # Optionally, insert a single PORT column between two letters.
    total_cols = (len(letters) * 2) + (1 if _port_col() is not None else 0)
    port_col = _port_col()
    for i in range(total_cols):
        col_idx = first_letter_col + i
        if port_col is not None and col_idx == port_col:
            _set_col_width(ws, get_column_letter(col_idx), 5)
        else:
            _set_col_width(ws, get_column_letter(col_idx), 8)

    # Row heights
    ws.row_dimensions[header_row].height = 20
    ws.row_dimensions[letters_row].height = 18

    # Default port rows
    for r in range(tether1_start, tether2_tap_row + 1):
        ws.row_dimensions[r].height = 24
    ws.row_dimensions[tether1_tap_row].height = 18
    ws.row_dimensions[tether2_tap_row].height = 18

    # When using the "6 ports total" layout, ports 3 and 6 absorb the TAP rows.
    # If port6_port_labels is enabled, we show a small top band (port range label)
    # and a larger content area below, keeping all 6 blocks equal-height.
    if port6_total_across_tethers:
        tap_h = ws.row_dimensions[tether1_tap_row].height or 18
        if port6_port_labels:
            # Desired equal block height: label_h + content_h == label_h + content2_h + tap_h
            # Use a slightly smaller label band to gain space for switch names.
            label_h = 10
            content_h = 38
            content2_h = (content_h - tap_h)
            # Tether 1 blocks: (3..4), (5..6), (7..9)
            ws.row_dimensions[3].height = label_h
            ws.row_dimensions[4].height = content_h
            ws.row_dimensions[5].height = label_h
            ws.row_dimensions[6].height = content_h
            ws.row_dimensions[7].height = label_h
            ws.row_dimensions[8].height = content2_h
            # row 9 keeps tap_h

            # Tether 2 blocks: (10..11), (12..13), (14..16)
            ws.row_dimensions[10].height = label_h
            ws.row_dimensions[11].height = content_h
            ws.row_dimensions[12].height = label_h
            ws.row_dimensions[13].height = content_h
            ws.row_dimensions[14].height = label_h
            ws.row_dimensions[15].height = content2_h
            # row 16 keeps tap_h
        else:
            # Equalize blocks without the label band.
            h_short = 20
            h_tall = (2 * h_short + tap_h) / 2
            for r in range(tether1_start, tether1_start + 4):
                ws.row_dimensions[r].height = h_tall
            for r in range(tether1_start + 4, tether1_ports_end + 1):
                ws.row_dimensions[r].height = h_short
            for r in range(tether2_start, tether2_start + 4):
                ws.row_dimensions[r].height = h_tall
            for r in range(tether2_start + 4, tether2_ports_end + 1):
                ws.row_dimensions[r].height = h_short

    # Top headers
    ws.merge_cells(start_row=header_row, start_column=port1_col, end_row=header_row, end_column=port2_col)
    c = _cell(ws, header_row, port1_col)
    c.value = "PANEL PORT"
    c.font = tether_font
    c.fill = header_fill
    c.alignment = center

    # Cable headers (compute using actual column positions so inserts are included)
    has_left = len(left_letters) > 0
    has_right = len(right_letters) > 0
    has_extra = len(extra_letters) > 0

    left_start_col = _letter_cols(0)[0] if has_left else None
    left_end_col = _letter_cols(len(left_letters) - 1)[1] if has_left else None

    right_start_col = (left_end_col + 1) if has_right else None
    right_end_col = _letter_cols(len(left_letters) + len(right_letters) - 1)[1] if has_right else None

    extra_start_col = ((right_end_col + 1) if has_right else (left_end_col + 1 if has_left else first_letter_col)) if has_extra else None
    extra_end_col = _letter_cols(len(letters) - 1)[1] if has_extra else None

    if has_left and left_start_col is not None and left_end_col is not None:
        ws.merge_cells(start_row=header_row, start_column=left_start_col, end_row=header_row, end_column=left_end_col)
        c = _cell(ws, header_row, left_start_col)
        c.value = sheet_cfg.get("title_left") or ""
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    if has_right and right_start_col is not None and right_end_col is not None:
        ws.merge_cells(start_row=header_row, start_column=right_start_col, end_row=header_row, end_column=right_end_col)
        c = _cell(ws, header_row, right_start_col)
        c.value = sheet_cfg.get("title_right") or ""
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    if has_extra and extra_start_col is not None and extra_end_col is not None:
        ws.merge_cells(start_row=header_row, start_column=extra_start_col, end_row=header_row, end_column=extra_end_col)
        c = _cell(ws, header_row, extra_start_col)
        c.value = sheet_cfg.get("title_extra") or ""
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    # Subheaders: PORT 1 / PORT 2
    _cell(ws, letters_row, port1_col).value = "PORT 1"
    _cell(ws, letters_row, port2_col).value = "PORT 2"
    for col in (port1_col, port2_col):
        cc = _cell(ws, letters_row, col)
        cc.font = header_font
        cc.fill = subheader_fill
        cc.alignment = center

    # Letter headers (merged across the 2 subcolumns)
    for i, letter in enumerate(letters):
        col1, col2 = _letter_cols(i)
        ws.merge_cells(start_row=letters_row, start_column=col1, end_row=letters_row, end_column=col2)
        cc = _cell(ws, letters_row, col1)
        cc.value = letter
        cc.font = header_font
        cc.fill = subheader_fill
        cc.alignment = center

    # Optional PORT column header (between letters)
    if port_col is not None:
        pc = _cell(ws, letters_row, port_col)
        pc.value = port_column_title
        pc.font = header_font
        pc.fill = subheader_fill
        pc.alignment = center

    # Panel ports (24 total). In the reference print, the PORT column occupies
    # the full height of Tether 1 + Tether 2 (not repeated). To match that,
    # we create 6 tall merged blocks (one per port pair).
    port1_pairs = ["1/2", "3/4", "5/6", "7/8", "9/10", "11/12"]
    port2_pairs = ["13/14", "15/16", "17/18", "19/20", "21/22", "23/24"]

    port_span_start = tether1_start
    port_span_end = tether2_ports_end
    span_rows = port_span_end - port_span_start + 1
    base = span_rows // len(port1_pairs)
    rem = span_rows % len(port1_pairs)

    no_side = Side(style=None)

    cur = port_span_start
    for idx, (p1, p2) in enumerate(zip(port1_pairs, port2_pairs)):
        seg_len = base + (1 if idx >= (len(port1_pairs) - rem) and rem else 0)
        r_start = cur
        r_end = cur + seg_len - 1
        cur = r_end + 1

        # Merge tall cells
        ws.merge_cells(start_row=r_start, start_column=port1_col, end_row=r_end, end_column=port1_col)
        ws.merge_cells(start_row=r_start, start_column=port2_col, end_row=r_end, end_column=port2_col)

        c1 = _cell(ws, r_start, port1_col)
        c2 = _cell(ws, r_start, port2_col)
        c1.value = p1
        c2.value = p2
        c1.alignment = center
        c2.alignment = center

        # Remove internal horizontal borders inside each merged segment so it looks like 1 cell
        for r in range(r_start, r_end + 1):
            for col in (port1_col, port2_col):
                cc = _cell(ws, r, col)
                top_side = no_side if r != r_start else thin
                bottom_side = no_side if r != r_end else thin
                cc.border = Border(
                    left=cc.border.left,
                    right=cc.border.right,
                    top=top_side,
                    bottom=bottom_side,
                )

    # Tether labels
    ws.merge_cells(start_row=tether1_start, start_column=tether_col, end_row=tether1_tap_row, end_column=tether_col)
    c = _cell(ws, tether1_start, tether_col)
    c.value = "Tether 1"
    c.font = tether_font
    c.alignment = center_rot90

    ws.merge_cells(start_row=tether2_start, start_column=tether_col, end_row=tether2_tap_row, end_column=tether_col)
    c = _cell(ws, tether2_start, tether_col)
    c.value = "Tether 2"
    c.font = header_font
    c.alignment = center_rot90

    # Data cells: large merged block for switch name + small Tap row split in 2
    tap_font = Font(name=font_name, size=12)
    switch_font = Font(name=font_name, size=14)
    port_num_font = Font(name=font_name, size=14)

    def _split_two(value) -> tuple[str, str]:
        """Normalize a per-letter mapping into two strings (left/right subcolumn).

        Supported inputs:
        - "text" -> ("text", "")
        - ["a", "b"] -> ("a", "b")
        - {"left": "a", "right": "b"} -> ("a", "b")
        """
        if value is None:
            return "", ""
        if isinstance(value, str):
            v = value.strip()
            return _smart_wrap_text(v), ""
        if isinstance(value, (list, tuple)):
            left = str(value[0]).strip() if len(value) > 0 and value[0] is not None else ""
            right = str(value[1]).strip() if len(value) > 1 and value[1] is not None else ""
            return _smart_wrap_text(left), _smart_wrap_text(right)
        if isinstance(value, dict):
            left = str(value.get("left") or "").strip()
            right = str(value.get("right") or "").strip()
            return _smart_wrap_text(left), _smart_wrap_text(right)
        return str(value).strip(), ""

    def _split_six(value) -> list[str]:
        """Normalize a per-letter mapping into 6 strings (ports 1..6)."""
        if value is None:
            return [""] * 6
        if isinstance(value, str):
            v = value.strip()
            return [v] + ([""] * 5)
        if isinstance(value, dict):
            if "ports" in value and isinstance(value.get("ports"), (list, tuple)):
                seq = list(value.get("ports") or [])
            else:
                # Support a more readable dict format using port ranges as keys
                # (e.g., "01-02", "03-04", ..., "11-12"), matching the Excel header.
                pairs = [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10), (11, 12)]
                seq = []
                found_any = False
                for a, b in pairs:
                    candidates = [
                        f"{a}-{b}",
                        f"{a} - {b}",
                        f"{a}/{b}",
                        f"{a} / {b}",
                        f"{a:02d}-{b:02d}",
                        f"{a:02d} - {b:02d}",
                        f"{a:02d}/{b:02d}",
                        f"{a:02d} / {b:02d}",
                    ]
                    v = None
                    for k in candidates:
                        if k in value:
                            v = value.get(k)
                            found_any = True
                            break
                    seq.append(v)

                if not found_any:
                    seq = [value.get(str(i)) if str(i) in value else value.get(i) for i in range(1, 7)]
            out = [str(x).strip() if x is not None else "" for x in seq]
            return (out + ([""] * 6))[:6]
        if isinstance(value, (list, tuple)):
            out = [str(x).strip() if x is not None else "" for x in list(value)]
            return (out + ([""] * 6))[:6]
        v = str(value).strip()
        return [v] + ([""] * 5)

    def fill_tether_block(tether_key: str, ports_start: int, ports_end: int, tap_row: int) -> None:
        tether_map = sheet_cfg.get(tether_key, {})

        def _merge_two_rows(r1: int, r2: int, col1: int, col2: int, value: str) -> None:
            ws.merge_cells(start_row=r1, start_column=col1, end_row=r2, end_column=col2)
            cc = _cell(ws, r1, col1)
            cc.value = value
            cc.font = switch_font
            cc.alignment = _content_alignment(value)

        # Optional PORT numbering column (1..6)
        if port_col is not None:
            if port6_total_across_tethers:
                # Show 1..3 in Tether 1 and 4..6 in Tether 2, each spanning 2 rows.
                start_value = 1 if tether_key == "tether1" else 4
                # Absorb the TAP row (tap_row) into the last port block so we don't
                # show a TAP divider on the right-side new connectors.
                pairs = [
                    (ports_start + 0, ports_start + 1),
                    (ports_start + 2, ports_start + 3),
                    (ports_start + 4, tap_row),
                ]
                for idx, (r1, r2) in enumerate(pairs):
                    panel_num = start_value + idx
                    if port6_port_labels:
                        # Top band: show the label "Port"; below: show the panel number.
                        try:
                            ws.unmerge_cells(start_row=r1, start_column=port_col, end_row=r2, end_column=port_col)
                        except Exception:
                            pass
                        c_top = _cell(ws, r1, port_col)
                        # Force a clean plain-string write to avoid any leftover rich-text underline.
                        c_top.value = None
                        c_top.value = port6_port_label_text
                        c_top.font = switch_font
                        c_top.alignment = center_no_wrap

                        if r1 + 1 <= r2:
                            ws.merge_cells(start_row=r1 + 1, start_column=port_col, end_row=r2, end_column=port_col)
                        c_num = _cell(ws, r1 + 1, port_col) if (r1 + 1 <= r2) else c_top
                        c_num.value = panel_num
                        c_num.font = switch_font
                        c_num.alignment = center
                    else:
                        ws.merge_cells(start_row=r1, start_column=port_col, end_row=r2, end_column=port_col)
                        cc = _cell(ws, r1, port_col)
                        cc.value = panel_num
                        cc.font = port_num_font
                        cc.alignment = center_no_wrap
            else:
                for rr in range(ports_start, ports_end + 1):
                    cc = _cell(ws, rr, port_col)
                    cc.value = rr - ports_start + 1
                    cc.font = port_num_font
                    cc.alignment = center_no_wrap
                _cell(ws, tap_row, port_col).value = ""

        for i, letter in enumerate(letters):
            col1, col2 = _letter_cols(i)

            is_port6 = letter in port6_letters

            if is_port6:
                # New connectors: 6 total ports, no TAPs.
                # Default: each tether has 6 ports.
                # If port6_total_across_tethers=true: use 3 ports in Tether 1 and 3 ports in Tether 2 (total 6),
                # each port spanning 2 rows to keep full table height.
                if port6_total_across_tethers:
                    # Source the full 6-port list from tether1 when not present in tether2.
                    src = tether_map.get(letter)
                    if src is None and tether_key == "tether2":
                        src = (sheet_cfg.get("tether1") or {}).get(letter)
                    ports = _split_six(src)
                    ports = ports[:6]
                    ports = ports[:3] if tether_key == "tether1" else ports[3:6]
                    pairs = [
                        (ports_start + 0, ports_start + 1),
                        (ports_start + 2, ports_start + 3),
                        (ports_start + 4, tap_row),
                    ]
                    panel_start = 1 if tether_key == "tether1" else 4
                    for idx, (r1, r2) in enumerate(pairs):
                        panel_num = panel_start + idx
                        p_odd = ((panel_num - 1) * 2) + 1
                        p_even = p_odd + 1
                        if port6_port_labels:
                            # Top band: port range label (e.g., "1 - 2").
                            ws.merge_cells(start_row=r1, start_column=col1, end_row=r1, end_column=col2)
                            cl = _cell(ws, r1, col1)
                            cl.value = f"{p_odd} - {p_even}"
                            cl.font = switch_font
                            cl.alignment = center

                            # Content area below the label.
                            if r1 + 1 <= r2:
                                ws.merge_cells(start_row=r1 + 1, start_column=col1, end_row=r2, end_column=col2)
                                cd = _cell(ws, r1 + 1, col1)
                                v = ports[idx] if idx < len(ports) else ""
                                cd.value = v
                                cd.font = switch_font
                                cd.alignment = _content_alignment(v)
                        else:
                            _merge_two_rows(r1, r2, col1, col2, ports[idx] if idx < len(ports) else "")
                else:
                    ports = _split_six(tether_map.get(letter))
                    for rr in range(ports_start, ports_end + 1):
                        ws.merge_cells(start_row=rr, start_column=col1, end_row=rr, end_column=col2)
                        cc = _cell(ws, rr, col1)
                        v = ports[rr - ports_start] if 0 <= (rr - ports_start) < 6 else ""
                        cc.value = v
                        cc.font = switch_font
                        cc.alignment = _content_alignment(v)

                # No TAP row for port6 letters in total-across-tethers mode.
                if not port6_total_across_tethers:
                    ws.merge_cells(start_row=tap_row, start_column=col1, end_row=tap_row, end_column=col2)
                    tc = _cell(ws, tap_row, col1)
                    tc.value = ""
                    tc.font = tap_font
                    tc.alignment = center_no_wrap
                continue

            # Switch name area: keep the two subcolumns separate so the divider line
            # exists through the whole letter field (as in the print).
            ws.merge_cells(start_row=ports_start, start_column=col1, end_row=ports_end, end_column=col1)
            ws.merge_cells(start_row=ports_start, start_column=col2, end_row=ports_end, end_column=col2)

            left_label, right_label = _split_two(tether_map.get(letter))

            cc_left = _cell(ws, ports_start, col1)
            cc_left.value = left_label
            cc_left.font = switch_font
            cc_left.alignment = _content_alignment(left_label, rotation=90)

            cc_right = _cell(ws, ports_start, col2)
            cc_right.value = right_label
            cc_right.font = switch_font
            cc_right.alignment = _content_alignment(right_label, rotation=90)

            # Tap numbering:
            # - Left cable resets across left_letters
            # - Right cable resets across right_letters
            # - Extra letters: panel space without a specific cable number, but still keep TAP labels
            tc1 = _cell(ws, tap_row, col1)
            tc2 = _cell(ws, tap_row, col2)
            tc1.font = tap_font
            tc2.font = tap_font
            tc1.alignment = center_no_wrap
            tc2.alignment = center_no_wrap

            if i < len(left_letters):
                idx_in_group = i
            elif i < (len(left_letters) + len(right_letters)):
                idx_in_group = i - len(left_letters)
            else:
                idx_in_group = i - (len(left_letters) + len(right_letters))

            tap_a = (idx_in_group * 2) + 1
            tap_b = tap_a + 1
            tc1.value = f"TAP {tap_a}"
            tc2.value = f"TAP {tap_b}"

    fill_tether_block("tether1", tether1_start, tether1_ports_end, tether1_tap_row)
    fill_tether_block("tether2", tether2_start, tether2_ports_end, tether2_tap_row)

    # Apply borders & fills to the whole used range
    max_row = tether2_tap_row
    max_col = first_letter_col + (len(letters) * 2) - 1 + (1 if port_col is not None else 0)

    for r in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cc = _cell(ws, r, col)
            cc.border = border
            if r == header_row:
                cc.fill = header_fill
                cc.font = header_font
                cc.alignment = center
            elif r == letters_row:
                cc.fill = subheader_fill
                cc.font = header_font
                cc.alignment = center
            elif r in (tether1_tap_row, tether2_tap_row):
                cc.font = tap_font
                # Keep current alignment if explicitly set elsewhere.
                if cc.alignment is None or (cc.alignment.horizontal is None and cc.alignment.vertical is None):
                    cc.alignment = center_no_wrap
            else:
                cc.font = switch_font

    # Vertical border weights:
    # - Between Tap subcolumns: hairline
    # - Between letters (after each pair): medium
    # Keep the very outer edges for the outer frame step below.
    for r in range(1, max_row + 1):
        # Hairline between subcolumns within each letter
        for i in range(len(letters)):
            col1, col2 = _letter_cols(i)
            is_port6 = letters[i] in port6_letters
            if 1 < col1 <= max_col and 1 <= col2 < max_col:
                if is_port6:
                    c_left = _cell(ws, r, col1)
                    c_right = _cell(ws, r, col2)
                    c_left.border = Border(
                        left=c_left.border.left,
                        right=no_side,
                        top=c_left.border.top,
                        bottom=c_left.border.bottom,
                    )
                    c_right.border = Border(
                        left=no_side,
                        right=c_right.border.right,
                        top=c_right.border.top,
                        bottom=c_right.border.bottom,
                    )
                else:
                    c_left = _cell(ws, r, col1)
                    c_left.border = Border(
                        left=c_left.border.left,
                        right=hair,
                        top=c_left.border.top,
                        bottom=c_left.border.bottom,
                    )

            # Medium line after each letter pair (except the last column; outer frame handles it)
            if col2 < max_col:
                c_right = _cell(ws, r, col2)
                c_right.border = Border(
                    left=c_right.border.left,
                    right=medium,
                    top=c_right.border.top,
                    bottom=c_right.border.bottom,
                )

        # Give the PORT column a clear separation on both sides.
        if port_col is not None and 1 <= port_col <= max_col:
            pc = _cell(ws, r, port_col)
            pc.border = Border(
                left=medium,
                right=medium,
                top=pc.border.top,
                bottom=pc.border.bottom,
            )

    # Outer frame a bit thicker (optional but helps match print)
    thick = Side(style="medium", color="000000")
    for r in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if r in (1, max_row) or col in (1, max_col):
                cc = _cell(ws, r, col)
                cc.border = Border(
                    left=thick if col == 1 else thin,
                    right=thick if col == max_col else thin,
                    top=thick if r == 1 else thin,
                    bottom=thick if r == max_row else thin,
                )

    # Cable 15 right-side template tweak: remove the small internal separator line
    # right under the word "Port" (PANEL column only) when using port labels.
    if port_col is not None and port6_total_across_tethers and port6_port_labels:
        no_side = Side(style=None)
        panel_blocks = [(3, 4), (5, 6), (7, 9), (10, 11), (12, 13), (14, 16)]
        for r1, _r2 in panel_blocks:
            c_top = _cell(ws, r1, port_col)
            c_num = _cell(ws, r1 + 1, port_col)
            c_top.border = Border(
                left=c_top.border.left,
                right=c_top.border.right,
                top=c_top.border.top,
                bottom=no_side,
            )
            c_num.border = Border(
                left=c_num.border.left,
                right=c_num.border.right,
                top=no_side,
                bottom=c_num.border.bottom,
            )

    # NOTE: In port6_total_across_tethers mode we keep full table height.

    out_path.parent.mkdir(parents=True, exist_ok=True)
    # NOTE: workbook save happens at a higher-level function.


def _sheet_fingerprint(sheet_cfg: dict) -> str:
    cfg = _normalize_sheet_cfg(sheet_cfg)
    return json.dumps(cfg, sort_keys=True, ensure_ascii=False, separators=(",", ":"))


def _sheet_layout_fingerprint(sheet_cfg: dict) -> str:
    """Fingerprint only the layout-affecting config (not the data).

    This lets us update cell values in-place without destroying any manual
    formatting, as long as the structure (letters/columns/modes) hasn't changed.
    """
    cfg = _normalize_sheet_cfg(sheet_cfg)
    layout = {
        "left_letters": cfg.get("left_letters") or [],
        "right_letters": cfg.get("right_letters") or [],
        "extra_letters": cfg.get("extra_letters") or [],
        "port6_letters": sorted(list(cfg.get("port6_letters") or [])),
        "port6_total_across_tethers": bool(cfg.get("port6_total_across_tethers") or False),
        "port6_port_labels": bool(cfg.get("port6_port_labels") or False),
        "port6_port_label_text": cfg.get("port6_port_label_text") or "Port",
        "port6_port_labels_style": "range" if bool(cfg.get("port6_port_labels") or False) else "",
        "port_column_after": cfg.get("port_column_after"),
        "port_column_title": cfg.get("port_column_title"),
    }
    return json.dumps(layout, sort_keys=True, ensure_ascii=False, separators=(",", ":"))


_LAYOUT_FP_ROW = 1000
_LAYOUT_FP_COL = 1
_LAYOUT_FP_PREFIX = "__anylan_layout_fp__:"


def _read_layout_fp(ws) -> str | None:
    try:
        v = ws.cell(row=_LAYOUT_FP_ROW, column=_LAYOUT_FP_COL).value
    except Exception:
        return None
    if not isinstance(v, str):
        return None
    if not v.startswith(_LAYOUT_FP_PREFIX):
        return None
    return v[len(_LAYOUT_FP_PREFIX) :]


def _write_layout_fp(ws, fp: str) -> None:
    cc = ws.cell(row=_LAYOUT_FP_ROW, column=_LAYOUT_FP_COL)
    cc.value = _LAYOUT_FP_PREFIX + fp
    ws.row_dimensions[_LAYOUT_FP_ROW].hidden = True


def _parse_layout_fp(fp: str | None) -> dict[str, Any] | None:
    if not fp:
        return None
    try:
        obj = json.loads(fp)
    except Exception:
        return None
    if not isinstance(obj, dict):
        return None
    return obj


def _migrate_enable_port6_port_labels(ws, sheet_cfg: dict) -> None:
    """Adjust merges to support the port-range label layout without recreating the sheet."""
    sheet_cfg = _normalize_sheet_cfg(sheet_cfg)

    left_letters = list(sheet_cfg.get("left_letters") or [])
    right_letters = list(sheet_cfg.get("right_letters") or [])
    extra_letters = list(sheet_cfg.get("extra_letters") or [])
    letters = left_letters + right_letters + extra_letters

    port6_letters = set(sheet_cfg.get("port6_letters") or [])

    # Layout constants
    first_letter_col = 4

    tether1_start = 3
    tether1_tap_row = 9

    tether2_start = 10
    tether2_tap_row = 16

    # PORT/PANEL column insertion
    port_column_after = sheet_cfg.get("port_column_after")
    insert_after_idx = None
    if port_column_after and isinstance(port_column_after, str):
        try:
            insert_after_idx = letters.index(port_column_after)
        except ValueError:
            insert_after_idx = None

    def _port_col() -> int | None:
        if insert_after_idx is None:
            return None
        return first_letter_col + ((insert_after_idx + 1) * 2)

    def _letter_cols(i: int) -> tuple[int, int]:
        shift = 1 if (insert_after_idx is not None and i > insert_after_idx) else 0
        col1 = first_letter_col + (i * 2) + shift
        return col1, col1 + 1

    def _migrate_block(col1: int, col2: int, r1: int, r2: int) -> None:
        # Unmerge any previous shapes we might have created.
        for a1, b1, a2, b2 in (
            (r1, col1, r2, col2),
            (r1, col1, r2, col1),
            (r1, col2, r2, col2),
            (r1, col1, r1, col2),
            (r1 + 1, col1, r2, col2),
        ):
            if a1 > a2:
                continue
            try:
                ws.unmerge_cells(start_row=a1, start_column=b1, end_row=a2, end_column=b2)
            except Exception:
                pass

        # New shape: top label row across both subcolumns + content area below.
        try:
            ws.merge_cells(start_row=r1, start_column=col1, end_row=r1, end_column=col2)
        except Exception:
            pass
        if r1 + 1 <= r2:
            try:
                ws.merge_cells(start_row=r1 + 1, start_column=col1, end_row=r2, end_column=col2)
            except Exception:
                pass

    # Panel blocks (6 total) using the existing row structure.
    blocks_t1 = [(tether1_start + 0, tether1_start + 1), (tether1_start + 2, tether1_start + 3), (tether1_start + 4, tether1_tap_row)]
    blocks_t2 = [(tether2_start + 0, tether2_start + 1), (tether2_start + 2, tether2_start + 3), (tether2_start + 4, tether2_tap_row)]

    port_col = _port_col()
    if port_col is not None:
        # Convert PANEL column blocks to: top label + number below.
        all_blocks = blocks_t1 + blocks_t2
        for r1, r2 in all_blocks:
            try:
                ws.unmerge_cells(start_row=r1, start_column=port_col, end_row=r2, end_column=port_col)
            except Exception:
                pass
            if r1 + 1 <= r2:
                try:
                    ws.merge_cells(start_row=r1 + 1, start_column=port_col, end_row=r2, end_column=port_col)
                except Exception:
                    pass

    for i, letter in enumerate(letters):
        if letter not in port6_letters:
            continue
        col1, col2 = _letter_cols(i)
        for r1, r2 in blocks_t1:
            _migrate_block(col1, col2, r1, r2)
        for r1, r2 in blocks_t2:
            _migrate_block(col1, col2, r1, r2)


def _update_sheet_values(ws, sheet_cfg: dict) -> None:
    """Update only values, preserving all existing formatting."""
    sheet_cfg = _normalize_sheet_cfg(sheet_cfg)

    left_letters = list(sheet_cfg.get("left_letters") or [])
    right_letters = list(sheet_cfg.get("right_letters") or [])
    extra_letters = list(sheet_cfg.get("extra_letters") or [])
    letters = left_letters + right_letters + extra_letters

    port6_letters = set(sheet_cfg.get("port6_letters") or [])
    port6_total_across_tethers = bool(sheet_cfg.get("port6_total_across_tethers") or False)
    port6_port_labels = bool(sheet_cfg.get("port6_port_labels") or False)
    port6_port_label_text = sheet_cfg.get("port6_port_label_text") or "Port"

    top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_no_wrap = Alignment(horizontal="center", vertical="center", wrap_text=False)
    center_shrink = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)
    center_rot90 = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    center_rot90_shrink = Alignment(
        horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True, text_rotation=90
    )

    def _content_alignment(value: Any, *, rotation: int | None = None) -> Alignment:
        if isinstance(value, str) and "\n" in value:
            if rotation == 90:
                return center_rot90
            return center
        if rotation == 90:
            return center_rot90_shrink
        return center_shrink

    # Layout constants (must match _build_sheet)
    port1_col = 1
    port2_col = 2
    tether_col = 3
    first_letter_col = 4

    header_row = 1
    letters_row = 2

    tether1_start = 3
    tether1_ports_end = 8
    tether1_tap_row = 9

    tether2_start = 10
    tether2_ports_end = 15
    tether2_tap_row = 16

    # Row heights for the Cable 15 right-side template (only adjusts heights; keeps colors/borders/fonts).
    if port6_total_across_tethers and port6_port_labels:
        # Slightly smaller label band to gain space for switch names.
        label_h = 10
        content_h = 38
        tap_h = ws.row_dimensions[tether1_tap_row].height or 18
        content2_h = max(12, content_h - tap_h)

        # Tether 1 blocks: (3..4), (5..6), (7..9)
        ws.row_dimensions[3].height = label_h
        ws.row_dimensions[4].height = content_h
        ws.row_dimensions[5].height = label_h
        ws.row_dimensions[6].height = content_h
        ws.row_dimensions[7].height = label_h
        ws.row_dimensions[8].height = content2_h

        # Tether 2 blocks: (10..11), (12..13), (14..16)
        ws.row_dimensions[10].height = label_h
        ws.row_dimensions[11].height = content_h
        ws.row_dimensions[12].height = label_h
        ws.row_dimensions[13].height = content_h
        ws.row_dimensions[14].height = label_h
        ws.row_dimensions[15].height = content2_h

    # PORT column insertion
    port_column_after = sheet_cfg.get("port_column_after")
    port_column_title = sheet_cfg.get("port_column_title") or "PORT"
    insert_after_idx = None
    if port_column_after and isinstance(port_column_after, str):
        try:
            insert_after_idx = letters.index(port_column_after)
        except ValueError:
            insert_after_idx = None

    def _port_col() -> int | None:
        if insert_after_idx is None:
            return None
        return first_letter_col + ((insert_after_idx + 1) * 2)

    def _letter_cols(i: int) -> tuple[int, int]:
        shift = 1 if (insert_after_idx is not None and i > insert_after_idx) else 0
        col1 = first_letter_col + (i * 2) + shift
        return col1, col1 + 1

    port_col = _port_col()

    # Headers
    ws.cell(row=header_row, column=port1_col).value = "PANEL PORT"

    has_left = len(left_letters) > 0
    has_right = len(right_letters) > 0
    has_extra = len(extra_letters) > 0

    left_start_col = _letter_cols(0)[0] if has_left else None
    left_end_col = _letter_cols(len(left_letters) - 1)[1] if has_left else None

    right_start_col = (left_end_col + 1) if has_right else None
    right_end_col = _letter_cols(len(left_letters) + len(right_letters) - 1)[1] if has_right else None

    extra_start_col = (
        ((right_end_col + 1) if has_right else (left_end_col + 1 if has_left else first_letter_col))
        if has_extra
        else None
    )
    extra_end_col = _letter_cols(len(letters) - 1)[1] if has_extra else None

    if has_left and left_start_col is not None:
        ws.cell(row=header_row, column=left_start_col).value = sheet_cfg.get("title_left") or ""
    if has_right and right_start_col is not None:
        ws.cell(row=header_row, column=right_start_col).value = sheet_cfg.get("title_right") or ""
    if has_extra and extra_start_col is not None:
        ws.cell(row=header_row, column=extra_start_col).value = sheet_cfg.get("title_extra") or ""

    ws.cell(row=letters_row, column=port1_col).value = "PORT 1"
    ws.cell(row=letters_row, column=port2_col).value = "PORT 2"

    for i, letter in enumerate(letters):
        col1, _col2 = _letter_cols(i)
        ws.cell(row=letters_row, column=col1).value = letter

    if port_col is not None:
        ws.cell(row=letters_row, column=port_col).value = port_column_title

    # Panel ports values (top-left cells only; merged formatting remains)
    port1_pairs = ["1/2", "3/4", "5/6", "7/8", "9/10", "11/12"]
    port2_pairs = ["13/14", "15/16", "17/18", "19/20", "21/22", "23/24"]
    port_span_start = tether1_start
    port_span_end = tether2_ports_end
    span_rows = port_span_end - port_span_start + 1
    base = span_rows // len(port1_pairs)
    rem = span_rows % len(port1_pairs)
    cur = port_span_start
    for idx, (p1, p2) in enumerate(zip(port1_pairs, port2_pairs)):
        seg_len = base + (1 if idx >= (len(port1_pairs) - rem) and rem else 0)
        r_start = cur
        r_end = cur + seg_len - 1
        cur = r_end + 1
        ws.cell(row=r_start, column=port1_col).value = p1
        ws.cell(row=r_start, column=port2_col).value = p2

    # Tether labels
    ws.cell(row=tether1_start, column=tether_col).value = "Tether 1"
    ws.cell(row=tether2_start, column=tether_col).value = "Tether 2"

    def _split_two(value) -> tuple[str, str]:
        if value is None:
            return "", ""
        if isinstance(value, str):
            v = value.strip()
            return _smart_wrap_text(v), ""
        if isinstance(value, (list, tuple)):
            left = str(value[0]).strip() if len(value) > 0 and value[0] is not None else ""
            right = str(value[1]).strip() if len(value) > 1 and value[1] is not None else ""
            return _smart_wrap_text(left), _smart_wrap_text(right)
        if isinstance(value, dict):
            left = str(value.get("left") or "").strip()
            right = str(value.get("right") or "").strip()
            return _smart_wrap_text(left), _smart_wrap_text(right)
        return str(value).strip(), ""

    def _split_six(value) -> list[str]:
        if value is None:
            return [""] * 6
        if isinstance(value, str):
            v = value.strip()
            return [v] + ([""] * 5)
        if isinstance(value, dict):
            if "ports" in value and isinstance(value.get("ports"), (list, tuple)):
                seq = list(value.get("ports") or [])
            else:
                pairs = [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10), (11, 12)]
                seq = []
                found_any = False
                for a, b in pairs:
                    candidates = [
                        f"{a}-{b}",
                        f"{a} - {b}",
                        f"{a}/{b}",
                        f"{a} / {b}",
                        f"{a:02d}-{b:02d}",
                        f"{a:02d} - {b:02d}",
                        f"{a:02d}/{b:02d}",
                        f"{a:02d} / {b:02d}",
                    ]
                    v = None
                    for k in candidates:
                        if k in value:
                            v = value.get(k)
                            found_any = True
                            break
                    seq.append(v)

                if not found_any:
                    seq = [value.get(str(i)) if str(i) in value else value.get(i) for i in range(1, 7)]
            out = [str(x).strip() if x is not None else "" for x in seq]
            return (out + ([""] * 6))[:6]
        if isinstance(value, (list, tuple)):
            out = [str(x).strip() if x is not None else "" for x in list(value)]
            return (out + ([""] * 6))[:6]
        v = str(value).strip()
        return [v] + ([""] * 5)

    def _write_tether_values(tether_key: str, ports_start: int, ports_end: int, tap_row: int) -> None:
        tether_map = sheet_cfg.get(tether_key, {}) or {}

        # Optional PORT numbering column
        if port_col is not None:
            if port6_total_across_tethers:
                start_value = 1 if tether_key == "tether1" else 4
                pairs = [
                    (ports_start + 0, ports_start + 1),
                    (ports_start + 2, ports_start + 3),
                    (ports_start + 4, tap_row),
                ]
                for idx, (r1, _r2) in enumerate(pairs):
                    panel_num = start_value + idx
                    if port6_port_labels:
                        c_top = ws.cell(row=r1, column=port_col)
                        # Force a clean plain-string write to avoid any leftover rich-text underline.
                        c_top.value = None
                        c_top.value = port6_port_label_text
                        c_top.alignment = center_no_wrap
                        c_num = ws.cell(row=r1 + 1, column=port_col)
                        c_num.value = panel_num
                        c_num.alignment = center_no_wrap
                    else:
                        ws.cell(row=r1, column=port_col).value = panel_num
            else:
                for rr in range(ports_start, ports_end + 1):
                    ws.cell(row=rr, column=port_col).value = rr - ports_start + 1
                ws.cell(row=tap_row, column=port_col).value = ""

        for i, letter in enumerate(letters):
            col1, col2 = _letter_cols(i)
            is_port6 = letter in port6_letters

            if is_port6:
                if port6_total_across_tethers:
                    pairs = [
                        (ports_start + 0, ports_start + 1),
                        (ports_start + 2, ports_start + 3),
                        (ports_start + 4, tap_row),
                    ]
                    if port6_port_labels:
                        panel_start = 1 if tether_key == "tether1" else 4
                        # Optional: use JSON values as the content text per panel block.
                        src = tether_map.get(letter)
                        src_present = letter in tether_map
                        if src is None and tether_key == "tether2":
                            t1_map = (sheet_cfg.get("tether1") or {})
                            if letter in t1_map:
                                src_present = True
                            src = t1_map.get(letter)
                        ports = _split_six(src)
                        ports = ports[:3] if tether_key == "tether1" else ports[3:6]
                        for idx, (r1, _r2) in enumerate(pairs):
                            panel_num = panel_start + idx
                            p_odd = ((panel_num - 1) * 2) + 1
                            p_even = p_odd + 1
                            c_label = ws.cell(row=r1, column=col1)
                            c_label.value = f"{p_odd} - {p_even}"
                            c_label.alignment = center_no_wrap

                            # Content text (if any) goes into the merged area below.
                            c_content = ws.cell(row=r1 + 1, column=col1)
                            text = ports[idx] if idx < len(ports) else ""
                            if isinstance(text, str):
                                text = text.strip()
                            # If the letter exists in JSON, treat it as authoritative and
                            # sync content (including clearing) so removed values don't linger.
                            if src_present:
                                c_content.value = text if text else None
                            # Keep the content area centered so manual edits remain centered on regen.
                            c_content.alignment = _content_alignment(text)
                    else:
                        src = tether_map.get(letter)
                        if src is None and tether_key == "tether2":
                            src = (sheet_cfg.get("tether1") or {}).get(letter)
                        ports = _split_six(src)
                        ports = ports[:3] if tether_key == "tether1" else ports[3:6]
                        for idx, (r1, _r2) in enumerate(pairs):
                            ws.cell(row=r1, column=col1).value = ports[idx] if idx < len(ports) else ""
                else:
                    ports = _split_six(tether_map.get(letter))
                    for rr in range(ports_start, ports_end + 1):
                        v = ports[rr - ports_start] if 0 <= (rr - ports_start) < 6 else ""
                        c = ws.cell(row=rr, column=col1)
                        c.value = v
                        c.alignment = _content_alignment(v)
                    ws.cell(row=tap_row, column=col1).value = ""
                continue

            left_label, right_label = _split_two(tether_map.get(letter))
            c_left = ws.cell(row=ports_start, column=col1)
            c_left.value = left_label
            c_left.alignment = _content_alignment(left_label, rotation=90)

            c_right = ws.cell(row=ports_start, column=col2)
            c_right.value = right_label
            c_right.alignment = _content_alignment(right_label, rotation=90)

            # TAP numbering labels
            if i < len(left_letters):
                idx_in_group = i
            elif i < (len(left_letters) + len(right_letters)):
                idx_in_group = i - len(left_letters)
            else:
                idx_in_group = i - (len(left_letters) + len(right_letters))

            tap_a = (idx_in_group * 2) + 1
            tap_b = tap_a + 1
            ws.cell(row=tap_row, column=col1).value = f"TAP {tap_a}"
            ws.cell(row=tap_row, column=col2).value = f"TAP {tap_b}"

    _write_tether_values("tether1", tether1_start, tether1_ports_end, tether1_tap_row)
    _write_tether_values("tether2", tether2_start, tether2_ports_end, tether2_tap_row)

    # Remove the small internal separator line under "Port" (PANEL column only).
    if port_col is not None and port6_total_across_tethers and port6_port_labels:
        no_side = Side(style=None)
        panel_blocks = [(3, 4), (5, 6), (7, 9), (10, 11), (12, 13), (14, 16)]
        for r1, _r2 in panel_blocks:
            c_top = ws.cell(row=r1, column=port_col)
            c_num = ws.cell(row=r1 + 1, column=port_col)
            c_top.border = Border(
                left=c_top.border.left,
                right=c_top.border.right,
                top=c_top.border.top,
                bottom=no_side,
            )
            c_num.border = Border(
                left=c_num.border.left,
                right=c_num.border.right,
                top=no_side,
                bottom=c_num.border.bottom,
            )

    _write_layout_fp(ws, _sheet_layout_fingerprint(sheet_cfg))


def _compose_cable0910_with_cable(sheets: dict, *, include_cable_sheet: bool) -> tuple[dict, dict[str, list[str]]]:
    """Compose output sheets.

    If both 'Cable 09 and 10' and 'Cable' exist in JSON, we merge J–M from 'Cable'
    into 'Cable 09 and 10' as an extra header group, and (by default) do not
    generate a separate 'Cable' worksheet.

    Returns (output_sheets, dependencies), where dependencies maps output sheet
    names to the input sheet names they depend on.
    """
    if not isinstance(sheets, dict):
        return sheets, {}

    if "Cable 09 and 10" not in sheets or "Cable" not in sheets:
        return sheets, {name: [name] for name in sheets.keys()}

    base_cfg = dict(sheets.get("Cable 09 and 10") or {})
    cable_cfg = dict(sheets.get("Cable") or {})

    cable_letters = list(cable_cfg.get("left_letters") or []) + list(cable_cfg.get("right_letters") or [])
    title_extra = cable_cfg.get("title_left") or cable_cfg.get("title_right") or ""

    base_cfg["title_extra"] = title_extra
    base_cfg["extra_letters"] = cable_letters

    for tether_key in ("tether1", "tether2"):
        base_map = dict(base_cfg.get(tether_key) or {})
        cable_map = dict(cable_cfg.get(tether_key) or {})
        for letter in cable_letters:
            base_map[letter] = cable_map.get(letter, ["", ""])
        base_cfg[tether_key] = base_map

    out_sheets = dict(sheets)
    out_sheets["Cable 09 and 10"] = base_cfg
    deps: dict[str, list[str]] = {name: [name] for name in sheets.keys()}
    deps["Cable 09 and 10"] = ["Cable 09 and 10", "Cable"]

    if not include_cable_sheet:
        out_sheets.pop("Cable", None)
        deps.pop("Cable", None)

    return out_sheets, deps


def generate_xlsx(
    data_path: Path,
    out_path: Path,
    sheet: str | None = None,
    *,
    sheets_to_update: list[str] | None = None,
) -> None:
    data = _load_mapping_data(data_path)
    sheets: dict = data.get("sheets") or {}
    if not isinstance(sheets, dict) or not sheets:
        raise ValueError("mapping_data.json must contain a non-empty 'sheets' object")

    # Compose: keep JSON split (Cable 09/10 + Cable), but output a single Excel tab.
    include_cable_sheet = sheet == "Cable"
    sheets, deps = _compose_cable0910_with_cable(sheets, include_cable_sheet=include_cable_sheet)

    if sheet is not None and sheet not in sheets:
        available = ", ".join(sheets.keys())
        raise ValueError(f"Sheet '{sheet}' not found in JSON. Available: {available}")

    if sheets_to_update is not None:
        missing = [name for name in sheets_to_update if name not in sheets]
        if missing:
            available = ", ".join(sheets.keys())
            raise ValueError(f"Sheets not found in JSON: {', '.join(missing)}. Available: {available}")

    if out_path.exists():
        wb = load_workbook(out_path)
    else:
        wb = Workbook()

    # If we created a new workbook, remove the default sheet after creating real sheets.
    default_sheet = wb.active.title if wb.worksheets else None

    if sheet is not None:
        items = [(sheet, sheets[sheet])]
    elif sheets_to_update is not None:
        # If any dependent input changed, update the composed output sheet.
        to_update = set(sheets_to_update)
        if "Cable" in to_update and "Cable 09 and 10" in sheets:
            to_update.add("Cable 09 and 10")
        items = [(name, sheets[name]) for name in sheets.keys() if name in to_update]
    else:
        items = list(sheets.items())

    for sheet_name, sheet_cfg in items:
        cfg = _normalize_sheet_cfg(sheet_cfg)

        # Backward-compatible sheet rename migrations (preserve manual formatting).
        # If the JSON sheet key changes, try to rename an existing legacy tab
        # instead of creating a brand-new one.
        if sheet_name not in wb.sheetnames:
            legacy_candidates_by_new = {
                "Cable 15 and 16": ["Cable 15"],
            }
            for legacy_name in legacy_candidates_by_new.get(sheet_name, []):
                if legacy_name in wb.sheetnames and sheet_name not in wb.sheetnames:
                    wb[legacy_name].title = sheet_name
                    break

        desired_layout_fp = _sheet_layout_fingerprint(cfg)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            existing_layout_fp = _read_layout_fp(ws)
            if existing_layout_fp == desired_layout_fp:
                _update_sheet_values(ws, cfg)
                continue

            # Attempt safe in-place migration for small layout changes that do not
            # require recreating the whole sheet (preserves manual formatting).
            old_layout = _parse_layout_fp(existing_layout_fp)
            new_layout = _parse_layout_fp(desired_layout_fp)
            if old_layout is not None and new_layout is not None:
                ignore_keys = {"port6_port_labels", "port6_port_label_text"}
                ignore_keys.add("port6_port_labels_style")
                old_cmp = {k: v for k, v in old_layout.items() if k not in ignore_keys}
                new_cmp = {k: v for k, v in new_layout.items() if k not in ignore_keys}
                if old_cmp == new_cmp and bool(new_layout.get("port6_port_labels") or False):
                    _migrate_enable_port6_port_labels(ws, cfg)
                    _update_sheet_values(ws, cfg)
                    _write_layout_fp(ws, desired_layout_fp)
                    continue

            # Layout changed: rebuild the tab.
            idx = wb.sheetnames.index(sheet_name)
            del wb[sheet_name]
            ws = wb.create_sheet(sheet_name, idx)
        else:
            ws = wb.create_sheet(sheet_name)

        _build_sheet(ws, cfg)
        _write_layout_fp(ws, desired_layout_fp)

    # If we are composing Cable 09/10 + Cable, ensure any old 'Cable' worksheet
    # is removed from an existing workbook.
    if sheet != "Cable" and "Cable" not in sheets and "Cable" in wb.sheetnames:
        del wb["Cable"]

    if default_sheet in wb.sheetnames and default_sheet not in sheets:
        if len(wb.sheetnames) > 1:
            del wb[default_sheet]

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _alt_out_path(out_path: Path) -> Path:
    base = out_path.with_suffix("")
    suffix = out_path.suffix or ".xlsx"
    for i in range(1, 1000):
        candidate = Path(f"{base}_NEW" + ("" if i == 1 else f"_{i}") + suffix)
        if not candidate.exists():
            return candidate
    # Fallback (should never happen)
    return Path(f"{base}_NEW_{int(time.time())}" + suffix)


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--data", default="mapping_data.json", help="Path to mapping_data.json")
    p.add_argument("--out", default="AnyLAN_Mapping.xlsx", help="Output .xlsx path")
    p.add_argument(
        "--sheet",
        default=None,
        help="Generate/update only one sheet (must match a key under 'sheets' in the JSON).",
    )
    p.add_argument(
        "--watch",
        action="store_true",
        help="Regenerate the .xlsx whenever the JSON file changes (polling).",
    )
    p.add_argument(
        "--watch-all",
        action="store_true",
        help="When used with --watch and no --sheet, regenerate ALL sheets on every change (default is only changed sheets).",
    )
    p.add_argument("--poll", type=float, default=1.0, help="Watch polling interval in seconds")
    return p.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    base = Path(__file__).resolve().parent
    data_path = (base / args.data).resolve() if not Path(args.data).is_absolute() else Path(args.data)
    out_path = (base / args.out).resolve() if not Path(args.out).is_absolute() else Path(args.out)

    def run_once() -> None:
        generate_xlsx(data_path=data_path, out_path=out_path, sheet=args.sheet)
        print(f"Wrote: {out_path}")

    def run_once_safe() -> bool:
        try:
            run_once()
            return True
        except PermissionError as exc:
            print(
                f"WARNING: sem permissao para ler/gravar '{out_path}'. "
                "Provavelmente o arquivo esta aberto no Excel (ou bloqueado pelo OneDrive). "
                "Feche o arquivo e o watcher tentara novamente."
            )
            return False

    if not args.watch:
        try:
            run_once()
        except PermissionError:
            print(
                f"WARNING: sem permissao para ler/gravar '{out_path}'. "
                "Provavelmente o arquivo esta aberto no Excel (ou bloqueado pelo OneDrive). "
                "Feche o arquivo e rode novamente."
            )
            raise SystemExit(1)
        raise SystemExit(0)

    last_mtime = data_path.stat().st_mtime

    last_fingerprints: dict[str, str] | None = None
    if args.sheet is None and not args.watch_all:
        # Prime fingerprints so subsequent edits update only the changed tabs.
        try:
            initial = _load_mapping_data(data_path)
            last_fingerprints = {
                name: _sheet_fingerprint(cfg)
                for name, cfg in (initial.get("sheets") or {}).items()
                if isinstance(cfg, dict)
            }
        except Exception:
            last_fingerprints = None

    run_once_safe()
    while True:
        time.sleep(args.poll)
        try:
            mtime = data_path.stat().st_mtime
        except FileNotFoundError:
            continue
        if mtime != last_mtime:
            last_mtime = mtime
            if args.sheet is not None or args.watch_all or last_fingerprints is None:
                ok = run_once_safe()
                if args.sheet is None and not args.watch_all:
                    if ok:
                        try:
                            current = _load_mapping_data(data_path)
                            last_fingerprints = {
                                name: _sheet_fingerprint(cfg)
                                for name, cfg in (current.get("sheets") or {}).items()
                                if isinstance(cfg, dict)
                            }
                        except Exception:
                            last_fingerprints = None
                continue

            # Smart mode: update only changed sheets.
            try:
                current = _load_mapping_data(data_path)
                current_sheets: dict[str, Any] = current.get("sheets") or {}
                current_fps = {
                    name: _sheet_fingerprint(cfg)
                    for name, cfg in current_sheets.items()
                    if isinstance(cfg, dict)
                }
            except Exception:
                # If JSON is temporarily invalid while editing, skip this tick.
                continue

            changed = [
                name
                for name, fp in current_fps.items()
                if last_fingerprints.get(name) != fp
            ]
            if changed:
                try:
                    generate_xlsx(data_path=data_path, out_path=out_path, sheets_to_update=changed)
                    print(f"Wrote: {out_path} (updated: {', '.join(changed)})")
                    last_fingerprints = current_fps
                except PermissionError:
                    print(
                        f"WARNING: sem permissao para ler/gravar '{out_path}'. "
                        "Provavelmente o arquivo esta aberto no Excel (ou bloqueado pelo OneDrive). "
                        "Feche o arquivo e o watcher tentara novamente."
                    )
            else:
                last_fingerprints = current_fps
