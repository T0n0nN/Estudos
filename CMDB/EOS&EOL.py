import os
import json
import time
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
from datetime import datetime, date
from urllib.parse import quote
from argparse import ArgumentParser

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)
# Carregar .env raiz e, por compatibilidade, Cisco.env se existir
load_dotenv(os.path.join(ROOT_DIR, ".env"))
CISCO_ENV = os.path.join(SCRIPT_DIR, "Cisco.env")
if os.path.exists(CISCO_ENV):
    load_dotenv(dotenv_path=CISCO_ENV)

# Permite override por .env
EXCEL_PATH = os.path.join(ROOT_DIR, os.getenv("EOX_EXCEL", os.path.join("CMDB", "network_eox.xlsx")))
SHEET_NAME = os.getenv("EOX_SHEET_NAME", "--all-sheets")
CACHE_PATH = os.path.join(ROOT_DIR, os.getenv("EOX_CACHE", os.path.join("CMDB", "eox_cache.json")))


def _clean_env(v: str | None) -> str | None:
    if v is None:
        return None
    # Remove espaços, quebras de linha e aspas acidentais
    return v.strip().strip('"').strip("'")

CLIENT_ID = _clean_env(
    os.getenv("CISCO_CLIENT_ID")
    or os.getenv("CLIENT_ID")
    or os.getenv("CCO_CLIENT_ID")
)
CLIENT_SECRET = _clean_env(
    os.getenv("CISCO_CLIENT_SECRET")
    or os.getenv("CLIENT_SECRET")
    or os.getenv("CCO_CLIENT_SECRET")
)
SCOPE = os.getenv("CISCO_SCOPE") or os.getenv("SCOPE") or None
DEBUG_OAUTH = (os.getenv("CISCO_DEBUG_OAUTH") or "0").lower() in ("1", "true", "yes")
DEBUG_EOX = (os.getenv("CISCO_DEBUG_EOX") or "0").lower() in ("1", "true", "yes")

TOKEN_URL = "https://cloudsso.cisco.com/as/token.oauth2"
# EOX endpoints (tentar em ordem)
EOX_BASES = [
    "https://apix.cisco.com/supporttools/eox/rest/5",
    "https://api.cisco.com/supporttools/eox/rest/5",
]
# Preferir o novo provedor primeiro para evitar 401 do cloudsso
TOKEN_URLS = [
    "https://id.cisco.com/oauth2/default/v1/token",
    TOKEN_URL,
]

TIMEOUT = 30
SLEEP_BETWEEN_CALLS = 0.5
UA = "Goodyear-EoX-Client/1.0 (+requests)"
# Mensagens padrão quando não há datas na base Cisco
NO_EOX_EOL = "sem EOL na base Cisco"
NO_EOX_EOS = "sem EOS na base Cisco"


def _load_cache():
    try:
        if os.path.exists(CACHE_PATH):
            with open(CACHE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def _save_cache(cache):
    try:
        with open(CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _parse_date(value: str) -> str:
    if not value:
        return ""
    if isinstance(value, (int, float)):
        try:
            return datetime.utcfromtimestamp(value).strftime("%Y-%m-%d")
        except Exception:
            return ""
    s = str(value).strip()
    if s in ("N/A", "NA", "None", "null", ""):
        return ""
    fmts = [
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%d-%b-%Y",
        "%b %d, %Y",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    # Sometimes API returns dicts or extra text
    try:
        # Extract digits to try parse
        digits = "".join(ch if ch.isdigit() or ch in "-/" else "" for ch in s)
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(digits, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass
    except Exception:
        pass
    return ""


def _field_value(rec: dict, key: str) -> str:
    v = rec.get(key)
    if isinstance(v, dict):
        v = v.get("value")
    return v if isinstance(v, str) else (str(v) if v is not None else "")


def _http_err_msg(he: requests.HTTPError) -> str:
    code = None
    reason = ""
    body = ""
    try:
        if he.response is not None:
            code = he.response.status_code
            reason = getattr(he.response, "reason", "") or ""
            try:
                txt = he.response.text
                if txt:
                    body = txt[:120]
            except Exception:
                pass
    except Exception:
        pass
    parts = ["http"]
    if code is not None:
        parts.append(str(code))
    if reason:
        parts.append(reason)
    if body:
        parts.append(body)
    return " ".join(parts)


def get_token() -> str:
    if not CLIENT_ID or not CLIENT_SECRET:
        raise RuntimeError("Cisco API credentials not found in Cisco.env (CLIENT_ID/CLIENT_SECRET).")
    if DEBUG_OAUTH:
        tail = CLIENT_ID[-4:] if isinstance(CLIENT_ID, str) and len(CLIENT_ID) >= 4 else "????"
        print(f"OAuth: iniciando tentativa de token (client_id=***{tail}, scope={SCOPE or '-'})")

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }

    attempts_info: list[str] = []

    for url in TOKEN_URLS:
        # Tentativa 1: client_id/client_secret no corpo
        data1 = {
            "grant_type": "client_credentials",
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
        }
        if SCOPE:
            data1["scope"] = SCOPE
        try:
            resp1 = requests.post(url, data=data1, headers=headers, timeout=TIMEOUT)
            if DEBUG_OAUTH:
                print(f"OAuth: POST {url} (body creds) -> {resp1.status_code}")
            if resp1.status_code == 200:
                return resp1.json().get("access_token")
            attempts_info.append(f"{url} body={resp1.status_code} {resp1.text[:200] if resp1.text else ''}")
        except Exception as ex:
            attempts_info.append(f"{url} body=exception {ex}")
            if DEBUG_OAUTH:
                print(f"OAuth: exceção em {url} (body creds): {ex}")

        # Tentativa 2: Basic Auth + corpo com grant_type (+ escopo)
        data2 = {"grant_type": "client_credentials"}
        if SCOPE:
            data2["scope"] = SCOPE
        try:
            resp2 = requests.post(url, data=data2, headers=headers, auth=(CLIENT_ID, CLIENT_SECRET), timeout=TIMEOUT)
            if DEBUG_OAUTH:
                print(f"OAuth: POST {url} (basic auth) -> {resp2.status_code}")
            if resp2.status_code == 200:
                return resp2.json().get("access_token")
            attempts_info.append(f"{url} basic={resp2.status_code} {resp2.text[:200] if resp2.text else ''}")
        except Exception as ex:
            attempts_info.append(f"{url} basic=exception {ex}")
            if DEBUG_OAUTH:
                print(f"OAuth: exceção em {url} (basic auth): {ex}")

    # Falha geral: agregar detalhes sem expor segredos
    summary = " | ".join(attempts_info)
    raise requests.HTTPError(f"Falha no OAuth em todos endpoints. Detalhes: {summary}")


def eox_by_serial(session: requests.Session, token: str, serial: str) -> dict:
    ser = quote(serial, safe="")
    last_err = None
    for base in EOX_BASES:
        # 1) Bearer header
        url = f"{base}/EOXBySerialNumber/1/{ser}?format=json"
        headers = {"Authorization": f"Bearer {token}", "Accept": "application/json", "User-Agent": UA}
        resp = session.get(url, headers=headers, timeout=TIMEOUT)
        if DEBUG_EOX:
            print(f"EOX serial {serial} via {base} (hdr): {resp.status_code}")
        if resp.status_code == 404:
            last_err = "404"
        else:
            try:
                resp.raise_for_status()
                data = resp.json() or {}
                records = data.get("EOXRecord") or data.get("Records") or []
                if isinstance(records, dict):
                    records = [records]
                if records:
                    rec = records[0]
                    # Preferir chaves corretas do v5
                    last_support = _field_value(rec, "LastDateOfSupport")
                    eol_announce = _field_value(rec, "EOXExternalAnnouncementDate")
                    end_of_sale = _field_value(rec, "EndOfSaleDate")
                    # Fallbacks
                    if not eol_announce:
                        eol_announce = _field_value(rec, "EndOfLifeAnnouncementDate")
                    if not end_of_sale:
                        end_of_sale = _field_value(rec, "EndOfSaleDate")
                    eos = _parse_date(last_support)
                    eol_pref = _parse_date(eol_announce) or _parse_date(end_of_sale)
                    if eos or eol_pref:
                        return {"source": "api", "found": True, "EOS": eos, "EOL": eol_pref}
                    else:
                        # Registro existe, mas sem datas úteis -> tentar outra base ou PID
                        last_err = "sem_datas"
                else:
                    last_err = "vazio"
            except requests.HTTPError as he:
                last_err = _http_err_msg(he)
        # 2) Query param access_token fallback (sem header)
        url2 = f"{url}&access_token={token}"
        headers2 = {"Accept": "application/json", "User-Agent": UA}
        resp2 = session.get(url2, headers=headers2, timeout=TIMEOUT)
        if DEBUG_EOX:
            print(f"EOX serial {serial} via {base} (param): {resp2.status_code}")
        if resp2.status_code == 404:
            last_err = "404"
            continue
        try:
            resp2.raise_for_status()
            data = resp2.json() or {}
            records = data.get("EOXRecord") or data.get("Records") or []
            if isinstance(records, dict):
                records = [records]
            if records:
                rec = records[0]
                last_support = _field_value(rec, "LastDateOfSupport")
                eol_announce = _field_value(rec, "EOXExternalAnnouncementDate")
                end_of_sale = _field_value(rec, "EndOfSaleDate")
                if not eol_announce:
                    eol_announce = _field_value(rec, "EndOfLifeAnnouncementDate")
                eos = _parse_date(last_support)
                eol_pref = _parse_date(eol_announce) or _parse_date(end_of_sale)
                if eos or eol_pref:
                    return {"source": "api", "found": True, "EOS": eos, "EOL": eol_pref}
                else:
                    last_err = "sem_datas"
            else:
                last_err = "vazio"
        except requests.HTTPError as he2:
            last_err = _http_err_msg(he2)
            continue
    # Não conseguimos dados úteis por serial
    return {"source": "api", "found": False}


def eox_by_pid(session: requests.Session, token: str, pid: str) -> dict:
    prod = quote(pid, safe="")
    last_err = None
    for base in EOX_BASES:
        # 1) Bearer header
        url = f"{base}/EOXByProductID/1/{prod}?format=json"
        headers = {"Authorization": f"Bearer {token}", "Accept": "application/json", "User-Agent": UA}
        resp = session.get(url, headers=headers, timeout=TIMEOUT)
        if DEBUG_EOX:
            print(f"EOX pid {pid} via {base} (hdr): {resp.status_code}")
        if resp.status_code == 404:
            last_err = "404"
        else:
            try:
                resp.raise_for_status()
                data = resp.json() or {}
                records = data.get("EOXRecord") or data.get("Records") or []
                if isinstance(records, dict):
                    records = [records]
                if records:
                    rec = records[0]
                    last_support = _field_value(rec, "LastDateOfSupport")
                    eol_announce = _field_value(rec, "EOXExternalAnnouncementDate")
                    end_of_sale = _field_value(rec, "EndOfSaleDate")
                    if not eol_announce:
                        eol_announce = _field_value(rec, "EndOfLifeAnnouncementDate")
                    eos = _parse_date(last_support)
                    eol_pref = _parse_date(eol_announce) or _parse_date(end_of_sale)
                    if eos or eol_pref:
                        return {"source": "api", "found": True, "EOS": eos, "EOL": eol_pref}
                    else:
                        last_err = "sem_datas"
                else:
                    last_err = "vazio"
            except requests.HTTPError as he:
                last_err = _http_err_msg(he)
        # 2) Query param access_token fallback (sem header)
        url2 = f"{url}&access_token={token}"
        headers2 = {"Accept": "application/json", "User-Agent": UA}
        resp2 = session.get(url2, headers=headers2, timeout=TIMEOUT)
        if DEBUG_EOX:
            print(f"EOX pid {pid} via {base} (param): {resp2.status_code}")
        if resp2.status_code == 404:
            last_err = "404"
            continue
        try:
            resp2.raise_for_status()
            data = resp2.json() or {}
            records = data.get("EOXRecord") or data.get("Records") or []
            if isinstance(records, dict):
                records = [records]
            if records:
                rec = records[0]
                last_support = _field_value(rec, "LastDateOfSupport")
                eol_announce = _field_value(rec, "EOXExternalAnnouncementDate")
                end_of_sale = _field_value(rec, "EndOfSaleDate")
                if not eol_announce:
                    eol_announce = _field_value(rec, "EndOfLifeAnnouncementDate")
                eos = _parse_date(last_support)
                eol_pref = _parse_date(eol_announce) or _parse_date(end_of_sale)
                if eos or eol_pref:
                    return {"source": "api", "found": True, "EOS": eos, "EOL": eol_pref}
                else:
                    last_err = "sem_datas"
            else:
                last_err = "vazio"
        except requests.HTTPError as he2:
            last_err = _http_err_msg(he2)
            continue
    return {"source": "api", "found": False}


def update_excel_with_eox(sheet_name: str | None = None, refresh: bool | None = None):
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Planilha não encontrada: {EXCEL_PATH}")

    # Use parâmetro, depois variável de ambiente EOX_SHEET_NAME, senão usa padrão SHEET_NAME
    selected_sheet = sheet_name or os.getenv("EOX_SHEET_NAME") or SHEET_NAME

    # Novo: por padrão, fazemos refresh (ignora cache) a cada execução, a menos que override via parâmetro/env
    if refresh is None:
        env_refresh = os.getenv("EOX_REFRESH", "1")
        refresh = env_refresh.lower() in ("1", "true", "yes", "y")

    df = pd.read_excel(EXCEL_PATH, sheet_name=selected_sheet)
    for col in ("EOL", "EOS", "eox_source", "error"):
        if col not in df.columns:
            df[col] = ""

    cache = _load_cache()
    token = get_token()
    session = requests.Session()

    results = []
    idents = []  # (serial, pid) por linha para relatório de mudanças
    for idx, row in df.iterrows():
        serial = str(row.get("Serial") or "").strip()
        pid = str(row.get("PID") or "").strip()
        idents.append((serial, pid))
        if not serial and not pid:
            results.append(("", "", "", "serial/pid ausente"))
            continue

        # Cache lookup (serial, then pid)
        cached = None
        if serial:
            cached = cache.get(serial) or cache.get(f"ser:{serial}")
        if not cached and pid:
            cached = cache.get(f"pid:{pid}")

        # Se NÃO for refresh, reutiliza cache e evita chamada à API
        if (not refresh) and cached and isinstance(cached, dict):
            eos = cached.get("EOS", "")
            eol = cached.get("EOL", "")
            # Se viemos de chamadas Cisco anteriores e ficaram vazios, preencher mensagens padrão por campo
            if str(cached.get("source", "")).startswith("cisco_api_"):
                if not eol:
                    eol = NO_EOX_EOL
                if not eos:
                    eos = NO_EOX_EOS
            results.append((eol, eos, cached.get("source", "cache"), cached.get("error", "")))
            continue

        # Try by serial first (if available)
        tried_serial = False
        try:
            if serial:
                tried_serial = True
                info = eox_by_serial(session, token, serial)
                if info.get("found"):
                    eos = info.get("EOS", "")
                    eol = info.get("EOL", "")
                    # Preencher mensagens padrão se algum campo vier vazio
                    if not eol:
                        eol = NO_EOX_EOL
                    if not eos:
                        eos = NO_EOX_EOS
                    cache[serial] = {"EOS": eos, "EOL": eol, "source": "cisco_api_serial"}
                    results.append((eol, eos, "cisco_api_serial", ""))
                    time.sleep(SLEEP_BETWEEN_CALLS)
                    continue
                else:
                    # not found by serial, try PID next
                    pass
        except requests.HTTPError as he:
            # If HTTP error by serial, try PID as fallback
            serial_err = _http_err_msg(he)
            if pid:
                try:
                    info2 = eox_by_pid(session, token, pid)
                    if info2.get("found"):
                        eos = info2.get("EOS", "")
                        eol = info2.get("EOL", "")
                        if not eol:
                            eol = NO_EOX_EOL
                        if not eos:
                            eos = NO_EOX_EOS
                        cache[f"pid:{pid}"] = {"EOS": eos, "EOL": eol, "source": "cisco_api_pid"}
                        results.append((eol, eos, "cisco_api_pid", ""))
                    else:
                        cache[f"pid:{pid}"] = {"EOS": NO_EOX_EOS, "EOL": NO_EOX_EOL, "source": "cisco_api_pid", "error": "não encontrado (pid)"}
                        results.append((NO_EOX_EOL, NO_EOX_EOS, "cisco_api_pid", "não encontrado (pid)"))
                except requests.HTTPError as he2:
                    # Fallback ao cache se disponível durante refresh
                    if refresh and cached and isinstance(cached, dict):
                        eos = cached.get("EOS", "")
                        eol = cached.get("EOL", "")
                        if str(cached.get("source", "")).startswith("cisco_api_"):
                            if not eol:
                                eol = NO_EOX_EOL
                            if not eos:
                                eos = NO_EOX_EOS
                        results.append((eol, eos, cached.get("source", "cache"), cached.get("error", "")))
                    else:
                        results.append(("", "", "cisco_api_pid", _http_err_msg(he2)))
                except Exception as e:
                    if refresh and cached and isinstance(cached, dict):
                        eos = cached.get("EOS", "")
                        eol = cached.get("EOL", "")
                        if str(cached.get("source", "")).startswith("cisco_api_"):
                            if not eol:
                                eol = NO_EOX_EOL
                            if not eos:
                                eos = NO_EOX_EOS
                        results.append((eol, eos, cached.get("source", "cache"), cached.get("error", "")))
                    else:
                        results.append(("", "", "cisco_api_pid", str(e)))
            else:
                # Sem PID: durante refresh, preferir manter dado anterior se existe
                if refresh and cached and isinstance(cached, dict):
                    eos = cached.get("EOS", "")
                    eol = cached.get("EOL", "")
                    if str(cached.get("source", "")).startswith("cisco_api_"):
                        if not eol:
                            eol = NO_EOX_EOL
                        if not eos:
                            eos = NO_EOX_EOS
                    results.append((eol, eos, cached.get("source", "cache"), cached.get("error", "")))
                else:
                    results.append(("", "", "cisco_api_serial", serial_err))
            time.sleep(SLEEP_BETWEEN_CALLS)
            continue
        except Exception as e:
            if pid:
                try:
                    info2 = eox_by_pid(session, token, pid)
                    if info2.get("found"):
                        eos = info2.get("EOS", "")
                        eol = info2.get("EOL", "")
                        if not eol:
                            eol = NO_EOX_EOL
                        if not eos:
                            eos = NO_EOX_EOS
                        cache[f"pid:{pid}"] = {"EOS": eos, "EOL": eol, "source": "cisco_api_pid"}
                        results.append((eol, eos, "cisco_api_pid", ""))
                    else:
                        cache[f"pid:{pid}"] = {"EOS": NO_EOX_EOS, "EOL": NO_EOX_EOL, "source": "cisco_api_pid", "error": "não encontrado (pid)"}
                        results.append((NO_EOX_EOL, NO_EOX_EOS, "cisco_api_pid", "não encontrado (pid)"))
                except requests.HTTPError as he2:
                    if refresh and cached and isinstance(cached, dict):
                        eos = cached.get("EOS", "")
                        eol = cached.get("EOL", "")
                        if str(cached.get("source", "")).startswith("cisco_api_"):
                            if not eol:
                                eol = NO_EOX_EOL
                            if not eos:
                                eos = NO_EOX_EOS
                        results.append((eol, eos, cached.get("source", "cache"), cached.get("error", "")))
                    else:
                        results.append(("", "", "cisco_api_pid", _http_err_msg(he2)))
                except Exception as e2:
                    if refresh and cached and isinstance(cached, dict):
                        eos = cached.get("EOS", "")
                        eol = cached.get("EOL", "")
                        if str(cached.get("source", "")).startswith("cisco_api_"):
                            if not eol:
                                eol = NO_EOX_EOL
                            if not eos:
                                eos = NO_EOX_EOS
                        results.append((eol, eos, cached.get("source", "cache"), cached.get("error", "")))
                    else:
                        results.append(("", "", "cisco_api_pid", str(e2)))
            else:
                if refresh and cached and isinstance(cached, dict):
                    eos = cached.get("EOS", "")
                    eol = cached.get("EOL", "")
                    if str(cached.get("source", "")).startswith("cisco_api_"):
                        if not eol:
                            eol = NO_EOX_EOL
                        if not eos:
                            eos = NO_EOX_EOS
                    results.append((eol, eos, cached.get("source", "cache"), cached.get("error", "")))
                else:
                    results.append(("", "", "cisco_api_serial", str(e)))
            time.sleep(SLEEP_BETWEEN_CALLS)
            continue

        # If reached here: either serial not found or not tried; try PID if available
        if pid:
            try:
                info2 = eox_by_pid(session, token, pid)
                if info2.get("found"):
                    eos = info2.get("EOS", "")
                    eol = info2.get("EOL", "")
                    if not eol:
                        eol = NO_EOX_EOL
                    if not eos:
                        eos = NO_EOX_EOS
                    cache[f"pid:{pid}"] = {"EOS": eos, "EOL": eol, "source": "cisco_api_pid"}
                    results.append((eol, eos, "cisco_api_pid", ""))
                else:
                    cache[f"pid:{pid}"] = {"EOS": NO_EOX_EOS, "EOL": NO_EOX_EOL, "source": "cisco_api_pid", "error": "não encontrado (pid)"}
                    # Se já tentamos serial e não achou, propagar como não encontrado geral
                    err = "não encontrado (serial/pid)" if tried_serial else "não encontrado (pid)"
                    results.append((NO_EOX_EOL, NO_EOX_EOS, "cisco_api_pid", err))
            except requests.HTTPError as he2:
                if refresh and cached and isinstance(cached, dict):
                    eos = cached.get("EOS", "")
                    eol = cached.get("EOL", "")
                    if str(cached.get("source", "")).startswith("cisco_api_"):
                        if not eol:
                            eol = NO_EOX_EOL
                        if not eos:
                            eos = NO_EOX_EOS
                    results.append((eol, eos, cached.get("source", "cache"), cached.get("error", "")))
                else:
                    results.append(("", "", "cisco_api_pid", _http_err_msg(he2)))
            except Exception as e2:
                if refresh and cached and isinstance(cached, dict):
                    eos = cached.get("EOS", "")
                    eol = cached.get("EOL", "")
                    if str(cached.get("source", "")).startswith("cisco_api_"):
                        if not eol:
                            eol = NO_EOX_EOL
                        if not eos:
                            eos = NO_EOX_EOS
                    results.append((eol, eos, cached.get("source", "cache"), cached.get("error", "")))
                else:
                    results.append(("", "", "cisco_api_pid", str(e2)))
        else:
            # Sem PID para fallback
            err = "não encontrado (serial)" if tried_serial else "serial ausente"
            # Se tentamos serial e não achou, preencher mensagens padrão; senão deixar em branco
            if tried_serial:
                results.append((NO_EOX_EOL, NO_EOX_EOS, "cisco_api_serial", err))
            else:
                results.append(("", "", "cisco_api_serial", err))

        time.sleep(SLEEP_BETWEEN_CALLS)

    _save_cache(cache)

    # Write back to the same sheet using openpyxl (preserve other sheets)
    wb = load_workbook(EXCEL_PATH)
    if selected_sheet not in wb.sheetnames:
        raise KeyError(f"Aba '{selected_sheet}' não encontrada na planilha.")
    ws = wb[selected_sheet]

    # Map headers -> column index
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if isinstance(val, str):
            headers[val.strip()] = col_idx

    def ensure_column(col_name: str) -> int:
        if col_name in headers:
            return headers[col_name]
        new_idx = ws.max_column + 1
        ws.cell(row=1, column=new_idx, value=col_name)
        headers[col_name] = new_idx
        return new_idx

    col_EOL = ensure_column("EOL")
    col_EOS = ensure_column("EOS")
    col_SRC = ensure_column("eox_source")
    col_ERR = ensure_column("error")

    # Coletar mudanças antes de salvar
    changes = []
    for i, (eol, eos, source, err) in enumerate(results, start=2):
        old_eol = ws.cell(row=i, column=col_EOL).value
        old_eos = ws.cell(row=i, column=col_EOS).value
        old_eol_s = str(old_eol or "")
        old_eos_s = str(old_eos or "")
        new_eol_s = str(eol or "")
        new_eos_s = str(eos or "")
        # Só considerar mudança se já havia informação anterior em pelo menos um dos campos
        had_prior_info = (old_eol_s.strip() != "" or old_eos_s.strip() != "")
        if had_prior_info and (old_eol_s != new_eol_s or old_eos_s != new_eos_s):
            serial, pid = idents[i - 2] if i - 2 < len(idents) else ("", "")
            changes.append({
                "sheet": selected_sheet,
                "row": i,
                "serial": serial,
                "pid": pid,
                "old_eol": old_eol_s,
                "new_eol": new_eol_s,
                "old_eos": old_eos_s,
                "new_eos": new_eos_s,
            })
        ws.cell(row=i, column=col_EOL, value=eol)
        ws.cell(row=i, column=col_EOS, value=eos)
        ws.cell(row=i, column=col_SRC, value=source)
        ws.cell(row=i, column=col_ERR, value=err)

    # Colorização do EOS conforme regra
    GREEN = "00FF00"   # Verde (Standard Green)
    ORANGE = "FFC000"  # Laranja (Standard Orange)
    RED = "FF0000"     # Vermelho
    fill_green = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    fill_orange = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
    fill_red = PatternFill(start_color=RED, end_color=RED, fill_type="solid")

    today = date.today()
    current_year = today.year

    for i in range(2, 2 + len(results)):
        cell = ws.cell(row=i, column=col_EOS)
        val = cell.value
        val_s = str(val or "").strip()
        # Sem data declarada (mensagem padrão ou vazio) => verde
        if not val_s or val_s.lower().startswith("sem eos"):
            cell.fill = fill_green
            continue
        # Tenta parse YYYY-MM-DD
        eos_dt = None
        try:
            eos_dt = datetime.strptime(val_s, "%Y-%m-%d").date()
        except Exception:
            # Tentar outros formatos comuns
            for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%b %d, %Y"):
                try:
                    eos_dt = datetime.strptime(val_s, fmt).date()
                    break
                except Exception:
                    pass
        if eos_dt is None:
            # Não parseável: manter sem cor específica
            continue
        # Expirada => vermelho
        if eos_dt < today:
            cell.fill = fill_red
        # Mesmo ano atual (e futura) => laranja
        elif eos_dt.year == current_year:
            cell.fill = fill_orange
        # Ano seguinte ou posterior => verde
        else:
            cell.fill = fill_green

    wb.save(EXCEL_PATH)

    # Retorna resumo para o chamador (usado para --all-sheets)
    return {"sheet": selected_sheet, "changes": changes}


if __name__ == "__main__":
    parser = ArgumentParser(description="Atualiza EOL/EOS na planilha network_eox.xlsx")
    parser.add_argument("-s", "--sheet", dest="sheet", help="Nome da aba a processar (padrão: variável EOX_SHEET_NAME ou 'Água Branca')", default=None)

    # Opções para refresh do cache: por padrão, fazemos refresh (ignora cache). '--no-refresh' restaura comportamento antigo.
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--refresh", action="store_true", help="Ignorar cache e reconsultar Cisco para todos os registros (padrão).")
    group.add_argument("--no-refresh", action="store_true", help="Usar cache quando disponível e evitar reconsultas (comportamento antigo).")

    # Novo: processar todas as abas do arquivo
    parser.add_argument("--all-sheets", action="store_true", help="Processa todas as abas do arquivo Excel, ignorando --sheet e EOX_SHEET_NAME.")

    args = parser.parse_args()

    if getattr(args, "no_refresh", False):
        refresh_flag = False
    elif getattr(args, "refresh", False):
        refresh_flag = True
    else:
        refresh_flag = (os.getenv("EOX_REFRESH", "1").lower() in ("1", "true", "yes", "y"))

    # Execução por uma ou todas as abas
    summaries = []
    requested_all = (
        getattr(args, "all_sheets", False)
        or os.getenv("EOX_SHEET_NAME", "").strip().lower() == "--all-sheets"
        or str(SHEET_NAME).strip().lower() == "--all-sheets"
    )
    if requested_all:
        wb_names = load_workbook(EXCEL_PATH).sheetnames
        for nm in wb_names:
            try:
                summaries.append(update_excel_with_eox(nm, refresh_flag))
            except Exception as ex:
                print(f"[WARN] Falha ao processar aba '{nm}': {ex}")
    else:
        summaries.append(update_excel_with_eox(args.sheet, refresh_flag))

    # Relatório final de mudanças
    total_changes = 0
    for s in summaries:
        ch = s.get("changes", []) if isinstance(s, dict) else []
        if ch:
            print(f"Atualizações na aba '{s.get('sheet', '?')}': {len(ch)}")
            for rec in ch:
                parts = []
                if rec.get("serial"): parts.append(f"Serial={rec['serial']}")
                if rec.get("pid"): parts.append(f"PID={rec['pid']}")
                eol_changed = rec.get("old_eol") != rec.get("new_eol")
                eos_changed = rec.get("old_eos") != rec.get("new_eos")
                if eol_changed:
                    parts.append(f"EOL: '{rec['old_eol']}' -> '{rec['new_eol']}'")
                if eos_changed:
                    parts.append(f"EOS: '{rec['old_eos']}' -> '{rec['new_eos']}'")
                print(" - " + "; ".join(parts) if parts else " - (linha alterada)")
            total_changes += len(ch)
    if total_changes == 0:
        print("Nenhuma atualização encontrada. EOL/EOS mantiveram-se iguais.")
