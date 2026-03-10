#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from netmiko import ConnectHandler
import os
import sys
import time
import getpass
import logging
import traceback
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Tuple, List, Dict
from dotenv import load_dotenv
from tqdm import tqdm
import socket

# -------------------------
# Carregar .env (raiz do workspace e pasta atual)
# -------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(os.path.dirname(SCRIPT_DIR))
load_dotenv(os.path.join(ROOT_DIR, ".env"))
load_dotenv(os.path.join(SCRIPT_DIR, ".env"))

# -------------------------
#          Configuráveis
# -------------------------
# Caminho da planilha (padrão: "Switches IOS.xlsx" no mesmo diretório). Pode sobrescrever via UCE_SWITCHES_XLSX
XLSX_FILE_ENV   = os.getenv("UCE_SWITCHES_XLSX", "").strip()
# Limitar a 5 em paralelo no máximo
MAX_WORKERS     = min(5, max(1, int(os.getenv("IOS_MAX_WORKERS", "5"))))
CONNECT_TIMEOUT = int(os.getenv("IOS_CONNECT_TIMEOUT", "10"))   # segundos
RETRY_COUNT     = int(os.getenv("IOS_RETRY_COUNT", "2"))
RETRY_DELAY     = int(os.getenv("IOS_RETRY_DELAY", "5"))        # segundos entre tentativas
# Storage preferido: "auto" tenta detectar (sdflash:/flash:/bootflash:), ou force com "sdflash:", "flash:" ou "bootflash:"
STORAGE_PREF    = os.getenv("IOS_STORAGE", "sdflash:").strip().lower()
# Fail-fast: cancela operações pendentes após primeira falha
FAIL_FAST       = (os.getenv("IOS_FAIL_FAST", "1").strip().lower() in ("1", "true", "yes"))

# -------------------------
#          Logger
# -------------------------

def setup_logger() -> Tuple[logging.Logger, str]:
    # Console apenas, sem arquivo
    logger = logging.getLogger("IOSUCE-DL")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(ch)
    return logger, ""

# -------------------------
#       Utilidades SSH
# -------------------------

def _enable_if_needed(conn, secret: str | None, logger: logging.Logger | None = None):
    try:
        if secret:
            conn.enable()
        try:
            conn.send_command("terminal length 0", expect_string=r"#", delay_factor=2)
        except Exception:
            pass
    except Exception as e:
        if logger:
            logger.debug(f"Falha ao entrar em enable: {e}")

def _detect_storage(conn) -> str:
    try:
        out = conn.send_command("dir sdflash:", expect_string=r"#", delay_factor=2)
        if "Directory of" in out or "bytes" in (out or "").lower():
            return "sdflash:"
    except Exception:
        pass
    for cand in ("sdflash0:", "sdflash1:"):
        try:
            out = conn.send_command(f"dir {cand}", expect_string=r"#", delay_factor=2)
            if "Directory of" in out or "bytes" in (out or "").lower():
                return cand
        except Exception:
            pass
    try:
        out = conn.send_command("dir flash:", expect_string=r"#", delay_factor=2)
        if "Directory of" in out or "bytes" in (out or "").lower():
            return "flash:"
    except Exception:
        pass
    try:
        out = conn.send_command("dir bootflash:", expect_string=r"#", delay_factor=2)
        if "Directory of" in out or "bytes" in (out or "").lower():
            return "bootflash:"
    except Exception:
        pass
    return "sdflash:"

def _image_exists(conn, storage: str, image_name: str) -> bool:
    try:
        out = conn.send_command(f"dir {storage}{image_name}", expect_string=r"#", delay_factor=3)
        return image_name in out and ("No such file" not in out and "Error" not in out)
    except Exception:
        return False

# -------------------------
#      TFTP e MD5
# -------------------------

def tftp_copy_image(conn, tftp_server: str, image_name: str, storage: str, logger: logging.Logger) -> Tuple[bool, str]:
    """Copia via TFTP usando URL completa, tentando responder prompts automaticamente."""
    dest = f"{storage}{image_name}"
    cmd = f"copy tftp://{tftp_server}/{image_name} {dest}"
    out = conn.send_command_timing(cmd, delay_factor=8, max_loops=2000)
    for _ in range(100):
        low = (out or '').lower()
        if "address or name of remote host" in low:
            out += "\n" + conn.send_command_timing(tftp_server, delay_factor=8, max_loops=2000)
            continue
        if "source filename" in low:
            out += "\n" + conn.send_command_timing(image_name, delay_factor=8, max_loops=2000)
            continue
        if "destination filename" in low:
            out += "\n" + conn.send_command_timing("\n", delay_factor=8, max_loops=2000)
            continue
        if "overwrite" in low and ("[confirm]" in low or "[y/n]" in low):
            # confirmar overwrite
            confirm = "\n" if "[confirm]" in low else "y"
            out += "\n" + conn.send_command_timing(confirm, delay_factor=8, max_loops=2000)
            continue
        if any(x in low for x in ["error", "timed out", "no space", "access violation", "not found", "%invalid"]):
            break
        # ler mais
        more = ""
        try:
            more = conn.read_channel()
        except Exception:
            pass
        if more:
            out += more
            time.sleep(1)
            continue
        # heurística de término
        if any(k in low for k in ["bytes copied", "copied in", "bytes/sec", "transfer completed", "copied, secs"]):
            break
        time.sleep(1)
    logger.debug(f"Saída do copy: \n{out}")
    low = (out or '').lower()
    ok = any(k in low for k in ["bytes copied", "copied in", "bytes/sec", "transfer completed"]) and not any(k in low for k in ["error", "no space", "timed out", "access violation", "%invalid"])
    return ok, out

def verify_md5(conn, storage: str, image_name: str, logger: logging.Logger) -> Tuple[bool, str, str]:
    """Retorna (ok, md5_calc, output)."""
    cmd = f"verify /md5 {storage}{image_name}"
    out = conn.send_command_timing(cmd, delay_factor=12, max_loops=4000)
    import re
    md5_match = re.search(r"=\s*([a-fA-F0-9]{32})", out or "")
    if not md5_match:
        md5_match = re.search(r"\b([a-fA-F0-9]{32})\b", out or "")
    calc = md5_match.group(1).lower() if md5_match else ""
    ok = bool(calc)
    logger.debug(f"Saída do verify md5: \n{out}")
    return ok, calc, out

# -------------------------
#     Leitura da planilha
# -------------------------

def _discover_xlsx_path() -> str:
    if XLSX_FILE_ENV and os.path.isfile(XLSX_FILE_ENV):
        return XLSX_FILE_ENV
    candidates = [
        os.path.join(SCRIPT_DIR, "Switches IOS.xlsx"),
        os.path.join(SCRIPT_DIR, "switches ios.xlsx"),
        os.path.join(SCRIPT_DIR, "Switches_IOS.xlsx"),
        # adicionalmente aceitar nomes genéricos
        os.path.join(SCRIPT_DIR, "Switches.xlsx"),
        os.path.join(SCRIPT_DIR, "switches.xlsx"),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    # fallback
    return os.path.join(SCRIPT_DIR, "Switches IOS.xlsx")


def read_plan_xlsx(xlsx_path: str) -> Tuple[List[Dict[str, str]], Dict[str, str]]:
    def normalize_key(s: str) -> str:
        s = (s or "").strip().lower()
        s = s.replace("_", " ").replace("-", " ")
        s = " ".join(s.split())
        return s
    def canonical_key(s: str) -> str:
        # remove espaços, underscores, hifens e caracteres não alfanuméricos para comparação robusta
        return "".join(ch for ch in (s or "").lower() if ch.isalnum())
    def normalize_columns(cols: List[str]) -> Dict[str, str]:
        # mapa principal por sinônimos (após normalize_key)
        synonyms = {
            "ip": "IP",
            "endereco": "IP",
            "endereco ip": "IP",
            "address": "IP",
            "tftp": "TFTP",
            "tftp server": "TFTP",
            "tftp ip": "TFTP",
            "tftp server ip": "TFTP",
            "server": "TFTP",
            "imagem": "IMAGE",
            "image": "IMAGE",
            "image filename": "IMAGE",
            "image file": "IMAGE",
            "nome da imagem": "IMAGE",
            "arquivo": "IMAGE",
            "bin": "IMAGE",
            "md5": "MD5",
            "hash": "MD5",
            # status
            "ios status": "IOS_STATUS",
            "status": "IOS_STATUS",
            "ios_status": "IOS_STATUS",
        }
        mapped: Dict[str, str] = {}
        for c in cols:
            if not isinstance(c, str):
                continue
            raw = c.strip()
            low = normalize_key(raw)
            if low in synonyms and synonyms[low] not in mapped:
                mapped[synonyms[low]] = raw
            else:
                # manter referência ao nome bruto
                if raw not in mapped:
                    mapped[raw] = raw
        # Fallbacks por chave canônica (aceita tftp_server_ip, image_filename etc.)
        can2raw: Dict[str, str] = {}
        for c in cols:
            if isinstance(c, str) and c.strip():
                can2raw.setdefault(canonical_key(c), c.strip())
        # Preencher ausentes com heurística canônica
        if "IP" not in mapped:
            for k in ("ip", "address", "enderecoip", "endereco"):
                if k in can2raw:
                    mapped["IP"] = can2raw[k]
                    break
        if "TFTP" not in mapped:
            for k in ("tftpserverip", "tftpip", "tftpserver", "tftp"):
                if k in can2raw:
                    mapped["TFTP"] = can2raw[k]
                    break
        if "IMAGE" not in mapped:
            for k in ("imagefilename", "imagefile", "image", "imagem", "bin", "arquivo", "filename"):
                if k in can2raw:
                    mapped["IMAGE"] = can2raw[k]
                    break
        if "MD5" not in mapped:
            for k in ("md5", "hash"):
                if k in can2raw:
                    mapped["MD5"] = can2raw[k]
                    break
        if "IOS_STATUS" not in mapped:
            for k in ("iosstatus", "ios_status", "status"):
                if k in can2raw:
                    mapped["IOS_STATUS"] = can2raw[k]
                    break
        return mapped
    try:
        import pandas as pd
        df = pd.read_excel(xlsx_path)
        colmap = normalize_columns(list(df.columns))
        # MD5 agora é opcional
        required = ["IP", "TFTP", "IMAGE"]
        missing = [r for r in required if r not in colmap]
        if missing:
            raise ValueError(
                "Planilha precisa conter as colunas mínimas (IP, TFTP, IMAGE). Não reconhecidas: "
                + ", ".join(missing)
                + ". Cabeçalhos encontrados: "
                + ", ".join(map(str, df.columns))
            )
        df = df.fillna("")
        rows: List[Dict[str, str]] = []
        for i, row in df.iterrows():
            ip    = str(row[colmap["IP"]]).strip()
            tftp  = str(row[colmap["TFTP"]]).strip() if "TFTP" in colmap else ""
            image = str(row[colmap["IMAGE"]]).strip()
            md5   = str(row[colmap.get("MD5", "")] if "MD5" in colmap else "").strip().lower()
            if not ip:
                continue
            rows.append({
                "_row": int(i) + 2,  # linha real no Excel
                "IP": ip,
                "TFTP": tftp,
                "IMAGE": image,
                "MD5": md5,
                "Nome": str(row.get("Nome", "")).strip(),
            })
        return rows, colmap
    except ImportError:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        colmap = normalize_columns(headers)
        required = ["IP", "TFTP", "IMAGE"]
        missing = [r for r in required if r not in colmap]
        if missing:
            raise ValueError(
                "Planilha precisa conter as colunas mínimas (IP, TFTP, IMAGE). Não reconhecidas: "
                + ", ".join(missing)
                + ". Cabeçalhos encontrados: "
                + ", ".join(headers)
            )
        rows: List[Dict[str, str]] = []
        for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
            vals = [str(c.value).strip() if c.value is not None else "" for c in r]
            h2v = {headers[idx]: vals[idx] for idx in range(min(len(headers), len(vals)))}
            ip    = h2v.get(colmap["IP"], "").strip()
            tftp  = h2v.get(colmap["TFTP"], "").strip() if "TFTP" in colmap else ""
            image = h2v.get(colmap["IMAGE"], "").strip()
            md5   = (h2v.get(colmap.get("MD5", ""), "").strip()).lower() if "MD5" in colmap else ""
            if not ip:
                continue
            rows.append({"_row": i, "IP": ip, "TFTP": tftp, "IMAGE": image, "MD5": md5, "Nome": h2v.get("Nome", "").strip()})
        return rows, colmap

# -------------------------
#  Atualização do MD5/Status na planilha
# -------------------------

def update_md5_in_xlsx(xlsx_path: str, updates: List[Tuple[int, str]], md5_header_name: str):
    """updates: lista de tuplas (row_number, md5_hash) para escrever o hash na coluna MD5."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active
    # localizar coluna MD5 pela primeira linha
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    md5_col_idx = None
    for idx, h in enumerate(headers, start=1):
        if h.strip().lower() == md5_header_name.strip().lower():
            md5_col_idx = idx
            break
    if md5_col_idx is None:
        # se não achar por nome original, tentar por 'MD5'
        for idx, h in enumerate(headers, start=1):
            if h.strip().lower() == "md5":
                md5_col_idx = idx
                break
    if md5_col_idx is None:
        # criar coluna MD5 no final
        md5_col_idx = len(headers) + 1
        ws.cell(row=1, column=md5_col_idx, value="MD5")
    for row_num, md5v in updates:
        ws.cell(row=row_num, column=md5_col_idx, value=(md5v or "").lower())
    wb.save(xlsx_path)


def update_status_in_xlsx(xlsx_path: str, updates: List[Tuple[int, str]], status_header_name: str):
    """updates: lista de tuplas (row_number, status_value) para escrever na coluna IOS Status (OK/NOK)."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    status_col_idx = None
    # procurar pelo nome informado ou por variantes de 'IOS Status'
    targets = {status_header_name.strip().lower(), "ios status", "ios_status", "status"}
    for idx, h in enumerate(headers, start=1):
        if h and h.strip().lower() in targets:
            status_col_idx = idx
            break
    if status_col_idx is None:
        # criar coluna ao final com cabeçalho padrão
        status_col_idx = len(headers) + 1
        ws.cell(row=1, column=status_col_idx, value=(status_header_name or "IOS Status"))
    for row_num, val in updates:
        ws.cell(row=row_num, column=status_col_idx, value=(val or "").upper())
    wb.save(xlsx_path)

# -------------------------
#   Download por dispositivo
# -------------------------

def download_to_uce(row: Dict[str, str], username: str, password: str, secret: str | None, logger: logging.Logger) -> tuple:
    nome = row.get("Nome") or row.get("IP") or "Unknown"
    ip   = row.get("IP")
    tftp = row.get("TFTP")
    image= row.get("IMAGE")

    device = {
        "device_type":  "cisco_ios",
        "host":         ip,
        "username":     username,
        "password":     password,
        "timeout":      CONNECT_TIMEOUT,
        "secret":       secret or "",
        "fast_cli":     False,
    }

    for attempt in range(1, RETRY_COUNT + 1):
        try:
            import re
            conn = ConnectHandler(**device)
            _enable_if_needed(conn, secret, logger)
            # Preferir STORAGE_PREF quando válido; caso contrário detectar
            allowed = {"sdflash:", "sdflash0:", "sdflash1:", "flash:", "bootflash:"}
            storage = STORAGE_PREF if STORAGE_PREF in allowed else _detect_storage(conn)

            # Se imagem já existe, pular copy e lidar com MD5
            if _image_exists(conn, storage, image):
                logger.info(f"{nome} ({ip}) ➜ imagem já existe em {storage}{image}. Pulando copy.")
                md5_sheet = (row.get("MD5") or "").strip().lower()
                if md5_sheet and re.fullmatch(r"[a-f0-9]{32}", md5_sheet):
                    try:
                        conn.disconnect()
                    except Exception:
                        pass
                    # Não recalcular MD5; manter o da planilha
                    return nome, ip, None, "SUCESSO", row.get("_row"), True
                # MD5 ausente/ inválido: calcular
                okv, calc, outv = verify_md5(conn, storage, image, logger)
                try:
                    conn.disconnect()
                except Exception:
                    pass
                if not okv or not calc:
                    logger.error(f"{nome} ({ip}) ➜ ERRO ao calcular MD5 de imagem já existente")
                    return nome, ip, None, "ERRO: md5", row.get("_row"), True
                logger.info(f"{nome} ({ip}) ➜ MD5 calculado: {calc}")
                return nome, ip, calc, "SUCESSO", row.get("_row"), True

            # Caso contrário, copiar imagem e calcular MD5
            logger.info(f"{nome} ({ip}) ➜ copiando {image} de {tftp} para {storage}...")
            okc, outc = tftp_copy_image(conn, tftp, image, storage, logger)
            if not okc:
                try:
                    conn.disconnect()
                except Exception:
                    pass
                logger.error(f"{nome} ({ip}) ➜ ERRO no copy TFTP")
                return nome, ip, None, "ERRO: copy tftp", row.get("_row"), False

            okv, calc, outv = verify_md5(conn, storage, image, logger)
            try:
                conn.disconnect()
            except Exception:
                pass
            if not okv or not calc:
                logger.error(f"{nome} ({ip}) ➜ ERRO ao calcular MD5 após copy")
                return nome, ip, None, "ERRO: md5", row.get("_row"), False

            logger.info(f"{nome} ({ip}) ➜ MD5 calculado: {calc}")
            return nome, ip, calc, "SUCESSO", row.get("_row"), False

        except Exception as e:
            tb = traceback.format_exc()
            logger.debug(f"[{nome} - tentativa {attempt}] erro: {e}\n{tb}")
            if attempt < RETRY_COUNT:
                time.sleep(RETRY_DELAY)
                continue
            else:
                logger.error(f"{nome} ({ip}) ➜ ERRO: {e}")
                return nome, ip, None, f"ERRO: {e}", row.get("_row"), False

# -------------------------
#             Main
# -------------------------

def main():
    # Credenciais via .env (opcional) ou prompt
    username = os.getenv("SSH_USERNAME") or input("👤 Usuário: ")
    password = os.getenv("SSH_PASSWORD") or getpass.getpass("🔒 Senha: ")
    enable_secret = os.getenv("SSH_ENABLE_SECRET") or None

    xlsx_path = _discover_xlsx_path()
    if not os.path.isfile(xlsx_path):
        print(f"❌ Planilha não encontrada: {xlsx_path}")
        sys.exit(1)

    try:
        rows, colmap = read_plan_xlsx(xlsx_path)
        if not rows:
            raise ValueError("Nenhum device encontrado na planilha.")
    except Exception as e:
        print(f"❌ {e}")
        sys.exit(1)

    # Teste rápido de credenciais
    first = rows[0]
    print(f"\n🔍 Testando credenciais em {first.get('IP')}...")
    device_test = {
        "device_type":  "cisco_ios",
        "host":         first.get("IP"),
        "username":     username,
        "password":     password,
        "timeout":      CONNECT_TIMEOUT,
        "secret":       enable_secret or "",
    }
    try:
        conn = ConnectHandler(**device_test)
        _enable_if_needed(conn, enable_secret, None)
        conn.send_command("show clock", expect_string=r"#", delay_factor=1)
        conn.disconnect()
        print("✅ Credenciais válidas. Iniciando downloads...\n")
    except Exception:
        print("❌ Credenciais inválidas ou privilégios insuficientes. Abortando script.")
        sys.exit(1)

    logger, _ = setup_logger()
    logger.info("▶️  Início do download de imagens (TFTP) e preenchimento de MD5/Status")

    sucesso = 0
    falha  = 0
    md5_updates: List[Tuple[int, str]] = []
    status_updates: List[Tuple[int, str]] = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(download_to_uce, r, username, password, enable_secret, logger): r for r in rows}
        for future in tqdm(as_completed(futures), total=len(futures), desc="⬇️  Baixando imagens", ncols=100):
            nome, ip, md5_calc, resultado, row_num, existed = future.result()
            orig_row = futures[future]
            tqdm.write(f"{nome} ({ip}) ➜ {resultado}")
            row_n = int(row_num or 0)
            # Validar se a planilha já tinha MD5 válido
            import re
            md5_expected = (orig_row.get("MD5") or "").strip().lower()
            has_valid_md5_in_sheet = bool(md5_expected and re.fullmatch(r"[a-f0-9]{32}", md5_expected))
            has_valid_md5_calc = bool(md5_calc and re.fullmatch(r"[a-f0-9]{32}", md5_calc))

            # Determinar status:
            if isinstance(resultado, str) and resultado.startswith("ERRO"):
                status_val = "NOK"
            else:
                if existed:
                    # arquivo já existia: OK se já havia MD5 na planilha ou se calculamos agora
                    status_val = "OK" if (has_valid_md5_in_sheet or has_valid_md5_calc) else "NOK"
                else:
                    # arquivo copiado: OK se conseguimos calcular MD5
                    status_val = "OK" if has_valid_md5_calc else "NOK"

            if row_n:
                status_updates.append((row_n, status_val))
                # Atualizar MD5 no Excel apenas se calculamos agora (não sobrescrever valor já existente)
                if has_valid_md5_calc and not has_valid_md5_in_sheet:
                    md5_updates.append((row_n, md5_calc))

            if status_val == "OK":
                sucesso += 1
            else:
                falha += 1

    # Escrever MD5s e Status na planilha
    try:
        md5_header_name = colmap.get("MD5", "MD5")
        status_header_name = colmap.get("IOS_STATUS", "IOS Status")
        # limpar updates inválidos
        md5_updates = [(r, m) for (r, m) in md5_updates if r and isinstance(r, int) and m]
        status_updates = [(r, s) for (r, s) in status_updates if r and isinstance(r, int) and s]
        if md5_updates:
            update_md5_in_xlsx(xlsx_path, md5_updates, md5_header_name)
        if status_updates:
            update_status_in_xlsx(xlsx_path, status_updates, status_header_name)
        if md5_updates or status_updates:
            print(f"📝 Planilha atualizada: MD5 ({len(md5_updates)}) e IOS Status ({len(status_updates)})")
        else:
            print("ℹ️  Nenhuma atualização para escrever na planilha.")
    except Exception as e:
        print(f"⚠️  Falha ao escrever na planilha: {e}")

    print("\n📊 Concluído!")
    print(f"✅ Sucesso: {sucesso}")
    print(f"❌ Falhas: {falha}")


if __name__ == "__main__":
    main()
