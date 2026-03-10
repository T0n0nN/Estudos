#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inventário de Telecom – Descoberta e Coleta

Resumo do que o script faz:
- Pede um IP semente e credenciais uma única vez
- Conecta em equipamentos Cisco (IOS/IOS-XE/NX-OS/ASA/WLC) via SSH (Netmiko)
- Descobre vizinhos por CDP e/ou LLDP e expande a varredura (BFS)
- Coleta: Device Name, Country, Location, Device Type, Model, IP, Serial Number
- Trata HA (ASA/WLC) e Stack (Switches) – contabiliza todos os chassis
- Para APs: nunca acessa os APs diretamente se houver WLC; coleta via WLC
- Atualiza a planilha existente sem alterar formatação, evita duplicidades e
  atualiza a aba "Total" com as contagens por país

Dependências sugeridas (instalar no seu ambiente Python):
  pip install netmiko openpyxl python-dotenv

Observações importantes:
- O script tenta localizar o arquivo de Excel por múltiplos nomes informados
  no workspace (Inventario_Telecom.xlsx ou Inventário Telecom.xlsx).
- A ordenação por Device Name é feita por inserção controlada para minimizar
  impacto na formatação (evita resort global da aba).
- A descoberta é limitada por salvaguardas para não explodir a rede (MAX_NODES).
"""

from __future__ import annotations

import ipaddress
import json
import os
import re
import sys
import time
import socket
import getpass
import logging
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set, Tuple
# NEW: typing Any for safe annotations in EOX helpers
from typing import Any
from collections import deque
import unicodedata
import shutil
from datetime import datetime, date
# NEW: EOX HTTP client
try:
    import requests  # used for Cisco EOX API
except Exception:
    requests = None

try:
    from netmiko import ConnectHandler
except Exception:
    ConnectHandler = None  # tratamos em runtime

try:
    from openpyxl import load_workbook
    # Color fill for EOS column (same rule as EOS&EOL.py)
    from openpyxl.styles import PatternFill
except Exception:
    load_workbook = None
    PatternFill = None

try:
    from dotenv import load_dotenv
except Exception:
    def load_dotenv(*args, **kwargs):
        return False

# -------------------- Configurações --------------------
WORKDIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # .../ChatGPT

# Caminho EXATO do arquivo a ser preenchido (não alterar formatação)
TARGET_EXCEL_PATH = r"C:\Users\za68397\Goodyear\Americas IT - LA Stefanini - Network Administrator L3\Network LA\Inventario\2025\Inventario_Telecom.xlsx"

# Mantemos candidatos antigos apenas como fallback opcional (não preferido)
DEFAULT_EXCEL_CANDIDATES = [
    TARGET_EXCEL_PATH,
    # Mesma pasta oficial – tentar variações de nome sem/with acento e com possível digitação
    r"C:\Users\za68397\Goodyear\Americas IT - LA Stefanini - Network Administrator L3\Network LA\Inventario\2025\Invetario_Telecom.xlsx",
    r"C:\Users\za68397\Goodyear\Americas IT - LA Stefanini - Network Administrator L3\Network LA\Inventario\2025\Inventário_Telecom.xlsx",
    # Fallbacks no workspace local
    os.path.join(WORKDIR, "Inventario_Telecom.xlsx"),
    os.path.join(WORKDIR, "Inventário", "Inventário Telecom.xlsx"),
    os.path.join(WORKDIR, "Inventário", "Inventário_Telecom.xlsx"),
]

INVENTARIO_SHEET_TOTAL = "Total"

# Limites de varredura (ajuste conforme necessário)
MAX_NODES = 600  # número máximo de IPs únicos a visitar
COMMAND_TIMEOUT = 30

# Países/Abas válidas no Excel (exatamente como aparecem nas abas)
VALID_SHEETS = {
    "Brasil",
    "Argentina",
    "Chile",
    "Colombia",
    "Mexico",
    "Peru",
    "Guatemala",
    # Inclua novas abas/paises aqui se existir na planilha
}

# --------- MISSING CORE CONSTANTS / HELPERS (restored) ---------
# Mapeamento de hostname -> (Country, Location)
SITE_MAP: Dict[str, Tuple[str, str]] = {
    "ar-buenosaires": ("Argentina", "Buenos Aires"),
    "br-americana": ("Brasil", "Americana"),
    "br-epiacaba": ("Brasil", "Epiacaba"),
    "br-paulinia": ("Brasil", "Paulínia"),
    "br-gravatai": ("Brasil", "Gravatí"),
    "br-recife": ("Brasil", "Recife"),
    "br-saobernardo": ("Brasil", "São Bernardo"),
    "br-saopaulolahq": ("Brasil", "Água Branca"),
    "br-stbarbara": ("Brasil", "Sta Bárbara"),
    "cl-santiago": ("Chile", "Maipu"),
    "cl-maipu": ("Chile", "Maipu"),
    "co-cali": ("Colombia", "Cali"),
    "gt-guatemala": ("Guatemala", "Guatemala"),
    "mx-sanluispotossi": ("Mexico", "San Luis Potosi"),
    "mx-santafe": ("Mexico", "Santa Fé"),
    "mx-tultitlan": ("Mexico", "Tultitlan"),
    "pe-callao": ("Peru", "Callao"),
}

def _normalize_txt(s: str) -> str:
    if not s:
        return ""
    nf = unicodedata.normalize("NFD", s)
    return "".join(ch for ch in nf if not unicodedata.combining(ch)).lower()

SITE_MAP_NORMALIZED: Dict[str, Tuple[str, str]] = { _normalize_txt(k): v for k, v in SITE_MAP.items() }

# Palavras que caracterizam telefones IP no CDP/LLDP (para ignorar)
PHONE_HINTS = ["IP Phone", "Cisco IP Phone", "CP-", "SEP", "Phone"]
# Debug flags
NEIGHBOR_DEBUG = True
AP_SERIAL_DEBUG = True

# Regex helpers
RE_IP = re.compile(r"(\d+\.\d+\.\d+\.\d+)")
RE_CDP_IP = re.compile(r"IP address: (\d+\.\d+\.\d+\.\d+)")
RE_LLDP_MGMT = re.compile(r"Management Address: (\d+\.\d+\.\d+\.\d+)")
RE_SN = re.compile(r"SN:\s*([A-Z0-9]+)")
AP_NAME_RE = re.compile(r"^[A-Za-z0-9_.-]{3,}$")
AP_STOPWORDS = {"show", "ap", "name", "number", "rd", "cc"}

# Helper para extrair PID do 'show inventory'
def get_pid_from_inventory(inv_text: str) -> Optional[str]:
    if not inv_text:
        return None
    m = re.search(r"(?is)NAME:\s*\"Chassis\".*?PID:\s*([\w-]+)", inv_text)
    if m:
        return m.group(1)
    m = re.search(r"(?i)PID:\s*([\w-]+)", inv_text)
    return m.group(1) if m else None
# ---------------------------------------------------------------

# ------------ EOX/EOS (Cisco API) Config -------------
# Reaproveita .env raiz e CMDB/Cisco.env
_EOX_SCRIPT_DIR = os.path.join(WORKDIR, "CMDB")
load_dotenv(os.path.join(WORKDIR, ".env"))
if os.path.exists(os.path.join(_EOX_SCRIPT_DIR, "Cisco.env")):
    load_dotenv(os.path.join(_EOX_SCRIPT_DIR, "Cisco.env"))

# Cache compartilhado com EOS&EOL.py
EOX_CACHE_PATH = os.path.join(_EOX_SCRIPT_DIR, "eox_cache.json")
# OAuth endpoints e API bases (mesmos do EOS&EOL.py)
_EOX_TOKEN_URLS = [
    "https://id.cisco.com/oauth2/default/v1/token",
    "https://cloudsso.cisco.com/as/token.oauth2",
]
_EOX_BASES = [
    "https://apix.cisco.com/supporttools/eox/rest/5",
    "https://api.cisco.com/supporttools/eox/rest/5",
]
_EOX_TIMEOUT = 30
_EOX_SLEEP = 0.5
_EOX_UA = "Goodyear-EoX-Client/1.0 (+requests)"
# Mensagens padrão
_NO_EOX_EOL = "sem EOL na base Cisco"
_NO_EOX_EOS = "sem EOS na base Cisco"

# Credenciais (aceita múltiplos nomes de variáveis)
def _eox_clean_env(v: Optional[str]) -> Optional[str]:
    if v is None:
        return None
    return v.strip().strip('"').strip("'")

EOX_CLIENT_ID = _eox_clean_env(
    os.getenv("CISCO_CLIENT_ID") or os.getenv("CLIENT_ID") or os.getenv("CCO_CLIENT_ID")
)
EOX_CLIENT_SECRET = _eox_clean_env(
    os.getenv("CISCO_CLIENT_SECRET") or os.getenv("CLIENT_SECRET") or os.getenv("CCO_CLIENT_SECRET")
)
EOX_SCOPE = os.getenv("CISCO_SCOPE") or os.getenv("SCOPE") or None
EOX_REFRESH_DEFAULT = (os.getenv("EOX_REFRESH", "1").lower() in ("1", "true", "yes", "y"))


def _eox_load_cache() -> Dict[str, dict]:
    try:
        if os.path.exists(EOX_CACHE_PATH):
            with open(EOX_CACHE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def _eox_save_cache(cache: Dict[str, dict]):
    try:
        os.makedirs(os.path.dirname(EOX_CACHE_PATH), exist_ok=True)
        with open(EOX_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _eox_parse_date(value: str) -> str:
    if not value:
        return ""
    s = str(value).strip()
    if s in ("N/A", "NA", "None", "null", ""):
        return ""
    fmts = ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%Y", "%b %d, %Y"]
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    # best-effort: keep as-is if unparsable
    return ""


def _eox_get_token() -> str:
    if requests is None:
        raise RuntimeError("A biblioteca 'requests' não está instalada. Instale com: pip install requests")
    if not EOX_CLIENT_ID or not EOX_CLIENT_SECRET:
        raise RuntimeError("Cisco API credentials not found (CLIENT_ID/CLIENT_SECRET) em .env/Cisco.env.")
    headers = {"Accept": "application/json", "Content-Type": "application/x-www-form-urlencoded"}
    for url in _EOX_TOKEN_URLS:
        data1 = {"grant_type": "client_credentials", "client_id": EOX_CLIENT_ID, "client_secret": EOX_CLIENT_SECRET}
        if EOX_SCOPE:
            data1["scope"] = EOX_SCOPE
        try:
            r = requests.post(url, data=data1, headers=headers, timeout=_EOX_TIMEOUT)
            if r.status_code == 200:
                return (r.json() or {}).get("access_token")
        except Exception:
            pass
        # Fallback basic auth
        data2 = {"grant_type": "client_credentials"}
        if EOX_SCOPE:
            data2["scope"] = EOX_SCOPE
        try:
            r2 = requests.post(url, data=data2, headers=headers, auth=(EOX_CLIENT_ID, EOX_CLIENT_SECRET), timeout=_EOX_TIMEOUT)
            if r2.status_code == 200:
                return (r2.json() or {}).get("access_token")
        except Exception:
            pass
    raise RuntimeError("Falha ao obter token OAuth da Cisco")


def _eox_field(rec: dict, key: str) -> str:
    v = rec.get(key)
    if isinstance(v, dict):
        v = v.get("value")
    return v if isinstance(v, str) else (str(v) if v is not None else "")


def _eox_by_serial(session: Any, token: str, serial: str) -> dict:
    ser = requests.utils.quote(serial, safe="")
    for base in _EOX_BASES:
        url = f"{base}/EOXBySerialNumber/1/{ser}?format=json"
        hdr = {"Authorization": f"Bearer {token}", "Accept": "application/json", "User-Agent": _EOX_UA}
        resp = session.get(url, headers=hdr, timeout=_EOX_TIMEOUT)
        if resp.status_code == 404:
            # try next base
            pass
        else:
            try:
                resp.raise_for_status()
                data = resp.json() or {}
                recs = data.get("EOXRecord") or data.get("Records") or []
                if isinstance(recs, dict):
                    recs = [recs]
                if recs:
                    rec = recs[0]
                    eos = _eox_parse_date(_eox_field(rec, "LastDateOfSupport"))
                    eol_ann = _eox_parse_date(_eox_field(rec, "EOXExternalAnnouncementDate")) or _eox_parse_date(_eox_field(rec, "EndOfSaleDate"))
                    return {"found": bool(eos or eol_ann), "EOS": eos or _NO_EOX_EOS, "EOL": eol_ann or _NO_EOX_EOL}
            except Exception:
                pass
        # query param fallback
        url2 = f"{url}&access_token={token}"
        resp2 = session.get(url2, headers={"Accept": "application/json", "User-Agent": _EOX_UA}, timeout=_EOX_TIMEOUT)
        try:
            resp2.raise_for_status()
            data2 = resp2.json() or {}
            recs2 = data2.get("EOXRecord") or data2.get("Records") or []
            if isinstance(recs2, dict):
                recs2 = [recs2]
            if recs2:
                rec = recs2[0]
                eos = _eox_parse_date(_eox_field(rec, "LastDateOfSupport"))
                eol_ann = _eox_parse_date(_eox_field(rec, "EOXExternalAnnouncementDate")) or _eox_parse_date(_eox_field(rec, "EndOfSaleDate"))
                return {"found": bool(eos or eol_ann), "EOS": eos or _NO_EOX_EOS, "EOL": eol_ann or _NO_EOX_EOL}
        except Exception:
            pass
    return {"found": False}


def _eox_by_pid(session: Any, token: str, pid: str) -> dict:
    prod = requests.utils.quote(pid, safe="")
    for base in _EOX_BASES:
        url = f"{base}/EOXByProductID/1/{prod}?format=json"
        hdr = {"Authorization": f"Bearer {token}", "Accept": "application/json", "User-Agent": _EOX_UA}
        resp = session.get(url, headers=hdr, timeout=_EOX_TIMEOUT)
        if resp.status_code == 404:
            pass
        else:
            try:
                resp.raise_for_status()
                data = resp.json() or {}
                recs = data.get("EOXRecord") or data.get("Records") or []
                if isinstance(recs, dict):
                    recs = [recs]
                if recs:
                    rec = recs[0]
                    eos = _eox_parse_date(_eox_field(rec, "LastDateOfSupport"))
                    eol_ann = _eox_parse_date(_eox_field(rec, "EOXExternalAnnouncementDate")) or _eox_parse_date(_eox_field(rec, "EndOfSaleDate"))
                    return {"found": bool(eos or eol_ann), "EOS": eos or _NO_EOX_EOS, "EOL": eol_ann or _NO_EOX_EOL}
            except Exception:
                pass
        # fallback with access_token param
        url2 = f"{url}&access_token={token}"
        resp2 = session.get(url2, headers={"Accept": "application/json", "User-Agent": _EOX_UA}, timeout=_EOX_TIMEOUT)
        try:
            resp2.raise_for_status()
            data2 = resp2.json() or {}
            recs2 = data2.get("EOXRecord") or data2.get("Records") or []
            if isinstance(recs2, dict):
                recs2 = [recs2]
            if recs2:
                rec = recs2[0]
                eos = _eox_parse_date(_eox_field(rec, "LastDateOfSupport"))
                eol_ann = _eox_parse_date(_eox_field(rec, "EOXExternalAnnouncementDate")) or _eox_parse_date(_eox_field(rec, "EndOfSaleDate"))
                return {"found": bool(eos or eol_ann), "EOS": eos or _NO_EOX_EOS, "EOL": eol_ann or _NO_EOX_EOL}
        except Exception:
            pass
    return {"found": False}


def _collect_records_from_country_sheets(wb) -> List[Tuple[str,str,str,str,str,str,str]]:
    """Lê as abas de países e retorna lista de tuplas
    (Device Name, Country, Location, Device Type, Model, IP, Serial Number)
    """
    rows: List[Tuple[str,str,str,str,str,str,str]] = []
    for sheet in wb.sheetnames:
        if sheet not in VALID_SHEETS:
            continue
        ws = wb[sheet]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not any(r):
                continue
            # Esperado: A..G
            name = str(r[0] or "").strip()
            country = str(r[1] or "").strip()
            location = str(r[2] or "").strip()
            dtype = str(r[3] or "").strip()
            model = str(r[4] or "").strip()
            ip = str(r[5] or "").strip()
            serial = str(r[6] or "").strip().upper()
            if not name:
                continue
            rows.append((name, country, location, dtype, model, ip, serial))
    # Remover totalmente vazios (sem Serial e sem PID)
    out = []
    seen = set()
    for t in rows:
        name, country, location, dtype, model, ip, serial = t
        key = serial or (model + "|" + name)
        if key in seen:
            continue
        seen.add(key)
        out.append(t)
    return out


def _write_eox_sheets(wb, records: List[Tuple[str,str,str,str,str,str,str]], results: Dict[str, Tuple[str,str]]):
    """Cria/atualiza as abas 'End of Life' e 'End of Support'.
    Estrutura de colunas: A..I = Device Name | Country | Location | Device Type | Model | IP | Serial Number | End of Life | End of Support
    A coloração segue a mesma regra do EOS&EOL.py: aplicada na coluna 'End of Support'.
    """
    headers = [
        "Device Name", "Country", "Location", "Device Type", "Model", "IP", "Serial Number", "End of Life", "End of Support"
    ]
    for tab in ("End of Life", "End of Support"):
        if tab in wb.sheetnames:
            ws = wb[tab]
            # Limpa conteúdo mantendo formatação básica (apaga valores)
            max_rows = ws.max_row
            max_cols = max(ws.max_column, len(headers))
            for c in range(1, max_cols + 1):
                ws.cell(row=1, column=c, value=headers[c-1] if c <= len(headers) else None)
            for r in range(2, max_rows + 1):
                for c in range(1, max_cols + 1):
                    ws.cell(row=r, column=c, value=None)
        else:
            ws = wb.create_sheet(tab)
            for i, h in enumerate(headers, start=1):
                ws.cell(row=1, column=i, value=h)
        # Preencher linhas
        for idx, rec in enumerate(records, start=2):
            name, country, location, dtype, model, ip, serial = rec
            ws.cell(row=idx, column=1, value=name)
            ws.cell(row=idx, column=2, value=country)
            ws.cell(row=idx, column=3, value=location)
            ws.cell(row=idx, column=4, value=dtype)
            ws.cell(row=idx, column=5, value=model)
            ws.cell(row=idx, column=6, value=ip)
            ws.cell(row=idx, column=7, value=serial)
            eol, eos = results.get(serial or (model + "|" + name), ("", ""))
            ws.cell(row=idx, column=8, value=eol or _NO_EOX_EOL)
            ws.cell(row=idx, column=9, value=eos or _NO_EOX_EOS)
        # Colorir EOS
        if PatternFill is not None:
            GREEN = "00FF00"
            ORANGE = "FFC000"
            RED = "FF0000"
            fill_green = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
            fill_orange = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
            fill_red = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
            today = date.today()
            current_year = today.year
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=9)  # End of Support
                val_s = str(cell.value or "").strip()
                if not val_s or val_s.lower().startswith("sem eos"):
                    cell.fill = fill_green
                    continue
                # parse YYYY-MM-DD
                eos_dt = None
                try:
                    eos_dt = datetime.strptime(val_s, "%Y-%m-%d").date()
                except Exception:
                    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%b %d, %Y"):
                        try:
                            eos_dt = datetime.strptime(val_s, fmt).date()
                            break
                        except Exception:
                            pass
                if eos_dt is None:
                    continue
                if eos_dt < today:
                    cell.fill = fill_red
                elif eos_dt.year == current_year:
                    cell.fill = fill_orange
                else:
                    cell.fill = fill_green


def _apply_eox_to_country_sheets(wb, results: Dict[str, Tuple[str, str]]):
    """Preenche as colunas H (End of Life) e I (End of Support) em cada aba de país.
    Não cria novas abas. Mantém formatação existente. Aplica coloração na coluna I.
    """
    # Preparar fills e datas para coloração
    fill_green = fill_orange = fill_red = None
    today = date.today()
    current_year = today.year
    if PatternFill is not None:
        GREEN = "00FF00"
        ORANGE = "FFC000"
        RED = "FF0000"
        fill_green = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
        fill_orange = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
        fill_red = PatternFill(start_color=RED, end_color=RED, fill_type="solid")

    for sheet in wb.sheetnames:
        if sheet not in VALID_SHEETS:
            continue
        ws = wb[sheet]
        # Garantir cabeçalhos
        if (str(ws.cell(row=1, column=8).value or "").strip() or "End of Life") != "End of Life":
            ws.cell(row=1, column=8, value="End of Life")
        if (str(ws.cell(row=1, column=9).value or "").strip() or "End of Support") != "End of Support":
            ws.cell(row=1, column=9, value="End of Support")

        for row in range(2, ws.max_row + 1):
            name = str(ws.cell(row=row, column=1).value or "").strip()
            model = str(ws.cell(row=row, column=5).value or "").strip()
            serial = str(ws.cell(row=row, column=7).value or "").strip().upper()
            if not any([name, model, serial]):
                continue
            key = serial or (model + "|" + name)
            eol, eos = results.get(key, ("", ""))
            if not eol and not eos:
                continue
            ws.cell(row=row, column=8, value=eol or _NO_EOX_EOL)
            eos_cell = ws.cell(row=row, column=9, value=eos or _NO_EOX_EOS)

            # Coloração do EOS
            if PatternFill is not None:
                val_s = str(eos_cell.value or "").strip()
                if not val_s or val_s.lower().startswith("sem eos"):
                    eos_cell.fill = fill_green
                else:
                    eos_dt = None
                    try:
                        eos_dt = datetime.strptime(val_s, "%Y-%m-%d").date()
                    except Exception:
                        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%b %d, %Y"):
                            try:
                                eos_dt = datetime.strptime(val_s, fmt).date()
                                break
                            except Exception:
                                continue
                    if eos_dt:
                        if eos_dt < today:
                            eos_cell.fill = fill_red
                        elif eos_dt.year == current_year:
                            eos_cell.fill = fill_orange
                        else:
                            eos_cell.fill = fill_green


def update_eox_tabs(excel_path: str, refresh: Optional[bool] = None):
    """Atualiza EOL/EOS diretamente nas colunas H/I das abas de países do arquivo informado,
    consultando Cisco por Serial (preferência) e PID (fallback), com cache. Não cria novas abas.
    """
    if load_workbook is None:
        raise RuntimeError("A biblioteca 'openpyxl' não está instalada. Instale com: pip install openpyxl")
    if requests is None:
        raise RuntimeError("A biblioteca 'requests' não está instalada. Instale com: pip install requests")
    if refresh is None:
        refresh = EOX_REFRESH_DEFAULT

    if not is_excel_writable(excel_path):
        raise RuntimeError("Arquivo Excel indisponível para escrita (talvez aberto no Excel)")

    wb = load_workbook(excel_path)
    records = _collect_records_from_country_sheets(wb)

    cache = _eox_load_cache()
    token = _eox_get_token()
    session = requests.Session()

    results: Dict[str, Tuple[str, str]] = {}
    for name, country, location, dtype, model, ip, serial in records:
        key = serial or (model + "|" + name)
        cached = None
        if serial:
            cached = cache.get(serial) or cache.get(f"ser:{serial}")
        if not cached and model:
            cached = cache.get(f"pid:{model}")
        if (not refresh) and cached and isinstance(cached, dict):
            eol = cached.get("EOL", _NO_EOX_EOL)
            eos = cached.get("EOS", _NO_EOX_EOS)
            results[key] = (eol or _NO_EOX_EOL, eos or _NO_EOX_EOS)
            continue

        eol_val = ""
        eos_val = ""
        tried_serial = False
        if serial:
            tried_serial = True
            try:
                info = _eox_by_serial(session, token, serial)
                if info.get("found"):
                    eol_val = info.get("EOL", _NO_EOX_EOL) or _NO_EOX_EOL
                    eos_val = info.get("EOS", _NO_EOX_EOS) or _NO_EOX_EOS
                    cache[serial] = {"EOL": eol_val, "EOS": eos_val, "source": "cisco_api_serial"}
                    results[key] = (eol_val, eos_val)
                    time.sleep(_EOX_SLEEP)
                    continue
            except Exception:
                pass
        if model:
            try:
                info2 = _eox_by_pid(session, token, model)
                if info2.get("found"):
                    eol_val = info2.get("EOL", _NO_EOX_EOL) or _NO_EOX_EOL
                    eos_val = info2.get("EOS", _NO_EOX_EOS) or _NO_EOX_EOS
                    cache[f"pid:{model}"] = {"EOL": eol_val, "EOS": eos_val, "source": "cisco_api_pid"}
                else:
                    eol_val = _NO_EOX_EOL if tried_serial else ""
                    eos_val = _NO_EOX_EOS if tried_serial else ""
            except Exception:
                if cached and isinstance(cached, dict):
                    eol_val = cached.get("EOL", _NO_EOX_EOL)
                    eos_val = cached.get("EOS", _NO_EOX_EOS)
        else:
            if tried_serial:
                eol_val = _NO_EOX_EOL
                eos_val = _NO_EOX_EOS
        results[key] = (eol_val, eos_val)
        time.sleep(_EOX_SLEEP)

    _eox_save_cache(cache)

    # Remover abas indevidas criadas anteriormente por engano
    for tab in ("End of Life", "End of Support"):
        if tab in wb.sheetnames:
            try:
                wb.remove(wb[tab])
            except Exception:
                pass

    # Aplicar EOL/EOS diretamente nas abas de países
    _apply_eox_to_country_sheets(wb, results)
    wb.save(excel_path)


# -------------------- Restante do script (descoberta/coleta) --------------------

@dataclass
class Credentials:
    username: str
    password: str
    secret: Optional[str] = None

@dataclass
class DeviceRecord:
    device_name: str
    country: str
    location: str
    device_type: str  # Router | Switch | WLC | Firewall | Access Points
    model: str
    ip: str
    serial_number: str

@dataclass
class DiscoveryContext:
    creds: Credentials
    visited: Set[str] = field(default_factory=set)
    # fila BFS agora deque para O(1) popleft
    queue: deque = field(default_factory=deque)
    found_records: List[DeviceRecord] = field(default_factory=list)
    neighbors: Dict[str, Set[str]] = field(default_factory=dict)
    ap_candidates_by_site: Dict[str, Set[Tuple[str, str]]] = field(default_factory=dict)
    wlc_sites: Set[str] = field(default_factory=set)
    reach_cache: Dict[str, bool] = field(default_factory=dict)  # cache reachability

# Helper para chave de site

def site_key(country: str, location: str) -> str:
    return f"{country}|{location}"


# Classifica vizinho (CDP/LLDP) e extrai ip/nome
def _classify_neighbor_section(sec: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    # retorna (kind, name, ip) onde kind in {"WLC","AP",None}
    sec_l = sec.lower()
    # Nome
    name = None
    m = re.search(r"(?im)^\s*Device ID\s*:\s*(.+)$", sec)
    if m:
        name = m.group(1).strip()
    if not name:
        m = re.search(r"(?im)^\s*System Name\s*:\s*(.+)$", sec)
        if m:
            name = m.group(1).strip()
    if not name:
        # Fallback: some parsers split on 'Device ID:', making first line the name
        lines = [ln.strip() for ln in sec.splitlines() if ln.strip()]
        if lines:
            first = lines[0]
            if (":" not in first) and not first.lower().startswith(("entry address", "interface", "platform", "holdtime")):
                name = first
    # IP (CDP ou LLDP)
    ip = None
    m = RE_CDP_IP.search(sec)
    if m:
        ip = m.group(1)
    if not ip:
        m = RE_LLDP_MGMT.search(sec)
        if m:
            ip = m.group(1)
        else:
            m = RE_IP.search(sec)
            if m:
                ip = m.group(1)
    # Classificação
    kind = None
    if re.search(r"(?i)C9800|AIR-CT|WLC|Wireless\s*Controller|Controller", sec):
        kind = "WLC"
    elif re.search(r"(?i)Wlan\s*AP|AIR-AP|AIR-CAP|AIR-LAP|\bAP\b|C91(1|2|3)0|Catalyst\s*9[1-3]..\s*AP|Access\s*Point", sec):
        kind = "AP"
    return kind, name, ip


# Atualiza: retorna (neigh_ips, wlc_ips, ap_infos)
def get_neighbors(conn, device_type: str) -> Tuple[Set[str], Set[str], List[Tuple[str, str]]]:
    """Extrai IPs de vizinhos via CDP/LLDP; ignora telefones e não-Cisco.
    Retorna: (todos_ips, wlc_ips, [(ap_ip, ap_name)]).
    Instrumentado com logs NEIGHBOR quando NEIGHBOR_DEBUG = True.
    """
    neigh_ips: Set[str] = set()
    wlc_ips: Set[str] = set()
    ap_infos: List[Tuple[str, str]] = []

    def _nlog(msg: str):
        if NEIGHBOR_DEBUG:
            logging.info("NEIGHBOR %s", msg)

    def _consider(sec: str, kind: Optional[str], name: Optional[str], ip: Optional[str]):
        raw_ip = ip
        # Fallback opcional: resolver IP por DNS para APs quando não vier IP no CDP/LLDP
        # Ativado somente se INVENTARIO_AP_DNS=1. Mantém padrão atual quando desativado.
        try:
            if (not ip or not _is_in_10_network(ip)) and kind == "AP" and name and os.getenv("INVENTARIO_AP_DNS") == "1":
                cand_name = name.split()[0]
                try:
                    resolved = socket.gethostbyname(cand_name)
                    if _is_in_10_network(resolved):
                        ip = resolved
                        _nlog(f"dns_ap_ip name={name} ip={ip}")
                except Exception:
                    pass
        except Exception:
            pass
        if not ip:
            _nlog(f"skip:no_ip name={name}")
            return
        try:
            ipaddress.ip_address(ip)
        except Exception:
            _nlog(f"skip:bad_ip name={name} ip={raw_ip}")
            return
        if not _is_in_10_network(ip):
            _nlog(f"skip:out_of_10/8 ip={ip} name={name}")
            return
        sec_l = sec.lower()
        if any(h.lower() in sec_l for h in [s.lower() for s in PHONE_HINTS]):
            _nlog(f"skip:phone ip={ip} name={name}")
            return
        is_cisco = ("cisco" in sec_l) or (kind in ("WLC", "AP"))
        if not is_cisco:
            _nlog(f"skip:not_cisco ip={ip} name={name}")
            return
        caps_m = re.search(r"(?i)Capabilities\s*:\s*([^\n]+)", sec)
        if caps_m:
            caps_txt = caps_m.group(1).strip()
            # Aceita routers antigos (ex: 2921) que anunciam Host/IGMP
            if not re.search(r"(?i)Router|Switch|Bridge|IGMP|Host", caps_txt) and kind not in ("WLC", "AP"):
                _nlog(f"skip:capabilities ip={ip} name={name} caps={caps_txt}")
                return

        if ip not in neigh_ips:
            _nlog(f"add:neighbor ip={ip} name={name} kind={kind}")
        neigh_ips.add(ip)
        if kind == "WLC":
            if ip not in wlc_ips:
                _nlog(f"add:wlc ip={ip} name={name}")
            wlc_ips.add(ip)
        elif kind == "AP":
            ap_infos.append((ip, name or ""))
            _nlog(f"add:ap_candidate ip={ip} name={name}")

    out_cdp = safe_send(conn, "show cdp neighbors detail")
    if out_cdp:
        sections = re.split(r"-{5,}|Device ID:", out_cdp)
        for sec in sections:
            s = sec.strip()
            if not s:
                continue
            # heurísticas simples
            name = None
            m_name = re.search(r"(?im)^(?:Device ID|System Name)\s*:?:?\s*([\w.-]+)", s)
            if m_name:
                name = m_name.group(1).strip()
            ip = None
            m_ip = re.search(r"(?im)IP address:\s*(\d+\.\d+\.\d+\.\d+)", s)
            if not m_ip:
                m_ip = re.search(r"(?im)IPv4 Address\s*:?:?\s*(\d+\.\d+\.\d+\.\d+)", s) or \
                       re.search(r"(?im)Management Address\s*:?:?\s*(\d+\.\d+\.\d+\.\d+)", s) or \
                       re.search(r"(?im)Management IP Address\s*:?:?\s*(\d+\.\d+\.\d+\.\d+)", s)
            if m_ip:
                ip = m_ip.group(1)
            kind = None
            plat = re.search(r"(?im)Platform:\s*cisco\s*([^,\n]+)", s)
            if plat:
                plat_u = plat.group(1).upper()
                # Ampliado para detectar APs Catalyst 9100 (C9110/C9120/C9130/C9105/C9100)
                if re.search(r"AP|AIR-AP|C91(?:10|20|30)|C9105|C9100|CATALYST\s*9[1-3]\d{2}", plat_u):
                    kind = "AP"
                elif re.search(r"WLC|C9800|AIR-CT|CT1?\d{3,4}|V?WLC|VIRTUAL", plat_u):
                    kind = "WLC"
            if not kind:
                # Caso não identificado por Platform, tentar pela linha completa
                if re.search(r"(?i)Wireless\s*Controller|AIR-CT|C9800|vWLC|Virtual\s*WLC", s):
                    kind = "WLC"
            caps_line = re.search(r"(?i)Capabilities\s*:\s*([^\n]+)", s)
            if caps_line and not kind and re.search(r"Controller", caps_line.group(1)):
                kind = "WLC"
            _consider(s, kind, name, ip)

    out_lldp = safe_send(conn, "show lldp neighbors detail")
    if out_lldp:
        sections = re.split(r"\n\s*----+\n", out_lldp)
        for sec in sections:
            s = sec.strip()
            if not s:
                continue
            name = None
            m_name = re.search(r"(?im)^(?:System Name|Device ID)\s*:?:?\s*([\w.-]+)", s)
            if m_name:
                name = m_name.group(1).strip()
            ip = None
            m_ip = re.search(r"(?im)Management Address:\s*(\d+\.\d+\.\d+\.\d+)", s)
            if m_ip:
                ip = m_ip.group(1)
            kind = None
            plat = re.search(r"(?im)System Description:\s*(.*)", s)
            if plat:
                desc = plat.group(1).upper()
                # Ampliado para detectar APs Catalyst 9100
                if re.search(r"AP|AIR-AP|C91(?:10|20|30)|C9105|C9100|CATALYST\s*9[1-3]\d{2}", desc):
                    kind = "AP"
                elif re.search(r"WLC|C9800", desc):
                    kind = "WLC"
            _consider(s, kind, name, ip)

    try:
        if out_cdp:
            for m in re.finditer(r"Device ID:\s*([\w.-]+)[\s\S]*?IP address:\s*(\d+\.\d+\.\d+\.\d+)[\s\S]*?Platform:\s*cisco\s*([^,\n]+)\s*,\s*Capabilities:\s*([^\n]+)", out_cdp, re.IGNORECASE):
                name, ip, plat_s, caps_s = m.group(1), m.group(2), m.group(3), m.group(4)
                if ip not in neigh_ips and _is_in_10_network(ip):
                    _nlog(f"fallback:add ip={ip} name={name} platform={plat_s} caps={caps_s.strip()}")
                    neigh_ips.add(ip)
    except Exception as e:
        _nlog(f"fallback:error {e}")

    try:
        self_ips = set()
        show_ip_int_br = safe_send(conn, "show ip interface brief")
        for m in RE_IP.findall(show_ip_int_br):
            if _is_in_10_network(m):
                self_ips.add(m)
        if self_ips:
            before = len(neigh_ips)
            neigh_ips -= self_ips
            wlc_ips -= self_ips
            ap_infos[:] = [(ip, nm) for (ip, nm) in ap_infos if ip not in self_ips]
            removed = before - len(neigh_ips)
            if removed:
                _nlog(f"self_ip_removed count={removed} self_ips={','.join(sorted(self_ips))}")
    except Exception as e:
        _nlog(f"self_ip_check_error {e}")

    if NEIGHBOR_DEBUG:
        _nlog(f"summary total_neighbors={len(neigh_ips)} wlc={len(wlc_ips)} ap_candidates={len(ap_infos)}")

    return neigh_ips, wlc_ips, ap_infos


# -------------------- Utilitários de IP --------------------

def _is_in_10_network(ip_str: str) -> bool:
    """Retorna True se o IP pertencer à rede 10.0.0.0/8."""
    try:
        ip_obj = ipaddress.ip_address(ip_str)
        return ip_obj in ipaddress.ip_network("10.0.0.0/8")
    except Exception:
        return False


# -------------------- Utilitários --------------------

def resolve_excel_path() -> str:
    # Prioriza o caminho oficial solicitado
    if os.path.exists(TARGET_EXCEL_PATH):
        return TARGET_EXCEL_PATH
    # Fallback para compatibilidade, caso esteja trabalhando com cópia local
    for p in DEFAULT_EXCEL_CANDIDATES:
        if os.path.exists(p):
            return p
    raise FileNotFoundError(
        "Arquivo de Excel do inventário não encontrado. Esperado em: "
        + TARGET_EXCEL_PATH
    )


# Assinatura canônica: (mantida) prioridade Serial -> Nome
def signature_key(name: str, serial: str, ip: str = "") -> str:
    serial = (serial or "").strip().upper()
    if serial:
        return f"SN:{serial}"
    return f"NM:{(name or '').strip().lower()}"


def is_ip_reachable(host: str, timeout: float = 5.0) -> bool:
    """Verifica reachability apenas em 10.0.0.0/8 e somente porta 22 (SSH)."""
    try:
        ipaddress.ip_address(host)
    except ValueError:
        return False
    # Apenas rede 10/8 é considerada
    if not _is_in_10_network(host):
        return False
    # Somente SSH (22)
    try:
        with socket.create_connection((host, 22), timeout=timeout):
            return True
    except Exception:
        return False


def infer_site_from_hostname(hostname: str) -> Tuple[str, str]:
    # Substitui versão anterior adicionando normalização de acentos
    if not hostname:
        return ("", "")
    h_norm = _normalize_txt(hostname)
    # procura o maior prefixo correspondente (normalizado)
    candidates = sorted(SITE_MAP_NORMALIZED.keys(), key=len, reverse=True)
    for prefix in candidates:
        if h_norm.startswith(prefix):
            return SITE_MAP_NORMALIZED[prefix]
    parts = h_norm.split("-")
    if len(parts) >= 2:
        pref = f"{parts[0]}-{parts[1]}"
        if pref in SITE_MAP_NORMALIZED:
            return SITE_MAP_NORMALIZED[pref]
    return ("", "")


# -------------------- Conexão e Coleta --------------------

# Ajustes pós-login para evitar paginação e melhorar coleta
def post_login_tuning(conn, device_type: str):
    try:
        if device_type in ("cisco_ios", "cisco_nxos"):
            conn.send_command_timing("terminal length 0")
        elif device_type == "cisco_asa":
            conn.send_command_timing("terminal pager 0")
        elif device_type == "cisco_wlc" or device_type == "cisco_wlc_ssh":
            # AireOS/9800 (alguns aceitam este comando; se não, ignora)
            conn.send_command_timing("config paging disable")
    except Exception:
        pass


# Trata login interativo duplo do AireOS (família 5xxx): User -> User -> Password
def _ensure_aireos_cli_login(conn, username: str, password: str):
    try:
        # Dispara o prompt de login/prompt atual
        conn.write_channel("\n")
        login_pat = r"(?i)(?:User(?:name)?|Password)\s*:\s*$"
        prompt_pat = r"(?i)\(Cisco Controller\).*[>#]"
        combined = f"{login_pat}|{prompt_pat}"
        loops = 0
        while loops < 12:
            out = conn.read_until_pattern(pattern=combined, timeout=10) or ""
            # Já no prompt da controladora?
            if re.search(prompt_pat, out):
                break
            # Pede usuário (pode ocorrer 2 vezes)
            if re.search(r"(?i)User(?:name)?\s*:\s*$", out):
                conn.write_channel(username + "\n")
                loops += 1
                continue
            # Pede senha
            if re.search(r"(?i)Password\s*:\s*$", out):
                conn.write_channel(password + "\n")
                # Espera prompt depois da senha
                conn.read_until_pattern(pattern=prompt_pat, timeout=15)
                break
            # Se nada casou, tenta nova linha para avançar
            conn.write_channel("\n")
            loops += 1
    except Exception:
        # Ignora para não derrubar a sessão – coleta tratará se não houver prompt
        pass


# Detecta WLC 9800 (IOS-XE) mesmo quando a sessão é cisco_ios
def is_iosxe_wlc_9800(model: str, sh_ver: Optional[str] = None) -> bool:
    m = (model or "").upper()
    if "C9800" in m or re.search(r"\b9800\b", m):
        return True
    if sh_ver:
        sv = sh_ver.upper()
        if "C9800" in sv or "WIRELESS CONTROLLER" in sv:
            return True
    return False


# Utilitário: mapeia IP por AP a partir do summary

def parse_ap_summary_ips(conn) -> Dict[str, str]:
    ip_by_name: Dict[str, str] = {}
    summary = safe_send(conn, "show ap summary")
    for line in summary.splitlines():
        line = line.strip()
        if not line or line.lower().startswith("ap name") or line.startswith("-"):
            continue
        parts = line.split()
        if not parts:
            continue
        name = parts[0]
        if not name or "#" in name:
            continue
        if not AP_NAME_RE.match(name) or name.lower() in AP_STOPWORDS:
            continue
        m_ip = RE_IP.search(line)
        ip = m_ip.group(1) if m_ip else ""
        ip_by_name[name] = ip
    return ip_by_name


# Coleta APs em WLC 9800 (IOS-XE) incluindo IP

def collect_iosxe_wlc_aps(conn) -> List[Tuple[str, str, str, str]]:
    aps: List[Tuple[str, str, str, str]] = []  # (name, model, sn, ip)
    ip_map = parse_ap_summary_ips(conn)
    for ap in ip_map.keys():
        out = safe_send(conn, f"show ap name {ap} config general") or safe_send(conn, f"show ap config general {ap}")
        m_model = re.search(r"(?im)^AP\s*Model\s*[\.:]+\s*([\w-]+)\s*$", out)
        m_sn = re.search(r"(?im)^AP\s*Serial\s*Number\s*[\.:]+\s*([A-Z0-9]+)\s*$", out)
        ap_model = m_model.group(1) if m_model else ""
        ap_sn = m_sn.group(1) if m_sn else ""
        ap_ip = ip_map.get(ap, "")
        if not AP_NAME_RE.match(ap) or ap.lower() in AP_STOPWORDS or "#" in ap:
            continue
        if not ap_model and not ap_sn and not ap_ip:
            continue
        aps.append((ap, ap_model, ap_sn, ap_ip))
    return aps


def connect_any(ip: str, creds: Credentials):
    """Conecta via SSH (sem Telnet) usando drivers Netmiko conhecidos.
    WLC AireOS tem tratamento especial de login (cisco_wlc/cisco_wlc_ssh).
    Retorna (conn, device_type).
    """
    if ConnectHandler is None:
        raise RuntimeError(
            "A biblioteca 'netmiko' não está instalada. Instale com: pip install netmiko"
        )

    # Segurança: não tenta fora da rede 10/8
    if not _is_in_10_network(ip):
        raise RuntimeError(f"IP fora da rede 10/8, ignorado: {ip}")

    device_types = [
        "cisco_ios",      # IOS/IOS-XE (routers/switches)
        "cisco_nxos",     # NX-OS
        "cisco_asa",      # ASA
        "cisco_wlc_ssh",  # WLC AireOS via SSH com handler próprio do Netmiko
        "cisco_wlc",      # WLC (compatibilidade)
    ]
    last_exc = None
    for dt in device_types:
        try:
            conn = ConnectHandler(
                device_type=dt,
                host=ip,
                username=creds.username,
                password=creds.password,
                timeout=COMMAND_TIMEOUT,
                auth_timeout=COMMAND_TIMEOUT,
                banner_timeout=COMMAND_TIMEOUT,
                allow_agent=False,
                use_keys=False,
                global_delay_factor=1.0,
                fast_cli=False,
            )
            # Trata login interativo do AireOS, se necessário
            if dt in ("cisco_wlc", "cisco_wlc_ssh"):
                _ensure_aireos_cli_login(conn, creds.username, creds.password)
            # Desabilita paginação conforme plataforma
            post_login_tuning(conn, dt)
            logging.info("Conectado a %s via %s", ip, dt)
            return conn, dt
        except Exception as e:
            last_exc = e
            logging.info("Falha no método %s em %s: %s", dt, ip, e)
            continue

    # Sem Telnet fallback
    raise last_exc or RuntimeError(f"Falha ao conectar em {ip} por SSH")


def safe_send(conn, cmd: str) -> str:
    try:
        return conn.send_command(
            cmd,
            expect_string=r"[#>",

            strip_prompt=True,
            strip_command=True,
            read_timeout=COMMAND_TIMEOUT,
        )
    except Exception:
        try:
            return conn.send_command(cmd, read_timeout=COMMAND_TIMEOUT)
        except Exception as e:
            logging.debug("Erro ao executar comando %s: %s", cmd, e)
            return ""


# Helpers para reachability (cache) e backup - colocados antes do uso

def _is_ip_reachable_cached(ip: str, ctx: DiscoveryContext) -> bool:
    cached = ctx.reach_cache.get(ip)
    if cached is not None:
        return cached
    ok = is_ip_reachable(ip)
    ctx.reach_cache[ip] = ok
    return ok


def backup_excel(path: str):
    # Desativado: não criar arquivo de backup (requisito do usuário)
    return

# -------------------- Excel helpers (upsert to avoid duplicates) --------------------

def _index_sheet_rows(ws):
    """Create quick indexes to find existing rows by (name, ip) and by name."""
    by_name_ip: Dict[Tuple[str, str], int] = {}
    by_name: Dict[str, int] = {}
    for r in ws.iter_rows(min_row=2, values_only=False):
        row_idx = r[0].row
        name = str(r[0].value or "").strip()
        ip = str(r[5].value or "").strip()
        if not name:
            continue
        key_name = name.lower()
        if key_name not in by_name:
            by_name[key_name] = row_idx
        if ip:
            by_name_ip[(key_name, ip)] = row_idx
    return by_name_ip, by_name


def _upsert_record(ws, rec: DeviceRecord, existing_sigs: Set[str], indices):
    """Merge into existing row when appropriate, else insert new.
    Rules:
    - If exact signature (by Serial or Name) exists: skip.
    - Prefer matching by (Name+IP). Fallback to Name only for Access Point.
    - If a matching row has a different non-empty Serial and the record Serial is non-empty:
      - For non-AP devices (Switch/Router/WLC/Firewall), INSERT a new row (stack/HA members).
      - For AP, UPDATE the existing row (treat as enrichment/replacement) and do not add a new row.
    - When updating a row, fill blanks only; when inserting, write full record.
    """
    by_name_ip, by_name = indices
    sig = record_signature(rec)
    if sig in existing_sigs:
        return

    key_name = rec.device_name.lower()
    target_row = by_name_ip.get((key_name, rec.ip))
    if not target_row and rec.device_type == "Access Point":
        target_row = by_name.get(key_name)

    new_sn = (rec.serial_number or "").strip().upper()

    if target_row:
        # Compare SNs to decide whether to merge or insert
        cur_sn = str(ws.cell(row=target_row, column=7).value or "").strip().upper()
        if new_sn and cur_sn and new_sn != cur_sn and rec.device_type != "Access Point":
            # Different chassis in stack/HA -> insert a new row
            insert_record_grouped(ws, rec)
            new_row = ws.max_row
            existing_sigs.add(sig)
            if rec.ip:
                by_name_ip[(key_name, rec.ip)] = new_row
            # keep first by_name mapping as-is to preserve position
            return
        # Otherwise, enrich existing row (fill blanks only)
        def _get(col):
            return str(ws.cell(row=target_row, column=col).value or "").strip()
        if not _get(2) and rec.country:
            ws.cell(row=target_row, column=2, value=rec.country)
        if not _get(3) and rec.location:
            ws.cell(row=target_row, column=3, value=rec.location)
        if not _get(4) and rec.device_type:
            ws.cell(row=target_row, column=4, value=rec.device_type)
        if not _get(5) and rec.model:
            ws.cell(row=target_row, column=5, value=rec.model)
        if not _get(6) and rec.ip:
            ws.cell(row=target_row, column=6, value=rec.ip)
        if not _get(7) and new_sn:
            ws.cell(row=target_row, column=7, value=new_sn)
            existing_sigs.add(sig)  # now this signature exists on sheet
            if rec.ip:
                by_name_ip[(key_name, rec.ip)] = target_row
        # If we didn't set SN (e.g., rec SN empty), do not add the sig yet
        return

    # No matching row -> insert
    insert_record_grouped(ws, rec)
    existing_sigs.add(sig)
    new_row = ws.max_row
    if rec.ip:
        by_name_ip[(key_name, rec.ip)] = new_row
    if key_name not in by_name:
        by_name[key_name] = new_row


# -------------------- Excel helpers --------------------

def load_wb(path: str):
    if load_workbook is None:
        raise RuntimeError(
            "A biblioteca 'openpyxl' não está instalada. Instale com: pip install openpyxl"
        )
    return load_workbook(path)


def ensure_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    return wb[sheet_name]


def row_signature(row_values: List[str]) -> str:
    # chave para deduplicação: prioriza Serial; se não houver, usa Device Name
    name = str(row_values[0] or "").strip()
    serial = str(row_values[6] or "").strip()
    ip = str(row_values[5] or "").strip()
    return signature_key(name, serial, ip)


def record_signature(rec: DeviceRecord) -> str:
    return signature_key(rec.device_name, rec.serial_number, rec.ip)


def read_existing_signatures(ws) -> Set[str]:
    sigs: Set[str] = set()
    # Assumimos cabeçalho na linha 1
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        vals = list(r) + [None] * 7
        sigs.add(row_signature(vals))
    return sigs


def insert_record_sorted(ws, rec: DeviceRecord):
    """Insere mantendo ordenação por Device Name (coluna A) com impacto mínimo.
    Não reordena linhas existentes, apenas encontra a posição de inserção e insere 1 linha.
    """
    # Cabeçalho assumido na linha 1
    target_row = ws.max_row + 1
    names = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        nm = (row[0] or "").strip()
        names.append(nm)
    # encontra a primeira posição cujo nome é > rec.device_name
    insert_at = None
    for idx, nm in enumerate(names, start=2):
        if nm and nm.lower() > rec.device_name.lower():
            insert_at = idx
            break
    if insert_at is None:
        insert_at = target_row
    # Insere a linha
    if insert_at <= ws.max_row:
        ws.insert_rows(insert_at, amount=1)
    # Escreve os valores
    ws.cell(row=insert_at, column=1, value=rec.device_name)
    ws.cell(row=insert_at, column=2, value=rec.country)
    ws.cell(row=insert_at, column=3, value=rec.location)
    ws.cell(row=insert_at, column=4, value=rec.device_type)
    ws.cell(row=insert_at, column=5, value=rec.model)
    ws.cell(row=insert_at, column=6, value=rec.ip)
    ws.cell(row=insert_at, column=7, value=rec.serial_number)


# Inserção agrupada por tipo (mantém dispositivos do mesmo tipo juntos)

def insert_record_grouped(ws, rec: DeviceRecord):
    """Insere mantendo membros de Stack/HA consecutivos.
    Preferência:
    1) Inserir logo após a última linha com o mesmo Device Name (case-insensitive).
    2) Se não houver, inserir após a última linha do mesmo Device Type.
    """
    # 1) Procura última ocorrência do mesmo Device Name
    last_row_same_name = None
    rec_name_l = (rec.device_name or "").strip().lower()
    if rec_name_l:
        for row in range(2, ws.max_row + 1):
            nm = str(ws.cell(row=row, column=1).value or "").strip().lower()
            if nm == rec_name_l:
                last_row_same_name = row

    if last_row_same_name:
        insert_at = last_row_same_name + 1
    else:
        # 2) Caso não exista, insere após o último do mesmo tipo
        last_row_of_type = None
        for row in range(2, ws.max_row + 1):
            cell_type = str(ws.cell(row=row, column=4).value or "").strip()
            if cell_type == rec.device_type:
                last_row_of_type = row
        insert_at = (last_row_of_type + 1) if last_row_of_type else (ws.max_row + 1)

    if insert_at <= ws.max_row:
        ws.insert_rows(insert_at, amount=1)
    # Escreve os valores
    ws.cell(row=insert_at, column=1, value=rec.device_name)
    ws.cell(row=insert_at, column=2, value=rec.country)
    ws.cell(row=insert_at, column=3, value=rec.location)
    ws.cell(row=insert_at, column=4, value=rec.device_type)
    ws.cell(row=insert_at, column=5, value=rec.model)
    ws.cell(row=insert_at, column=6, value=rec.ip)
    ws.cell(row=insert_at, column=7, value=rec.serial_number)


def update_totals_sheet(wb):
    if INVENTARIO_SHEET_TOTAL not in wb.sheetnames:
        return
    ws = wb[INVENTARIO_SHEET_TOTAL]

    # Le as colunas do cabeçalho para achar índice por país
    header = [c.value if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # Mapa coluna por país
    col_by_country: Dict[str, int] = {}
    for idx, title in enumerate(header, start=1):
        t = str(title).strip()
        if t in VALID_SHEETS:
            col_by_country[t] = idx
    # Mapa row por categoria (Switches/Routers/WLC/Firewall/Access Points/ UCE's etc)
    row_by_device: Dict[str, int] = {}
    for r in ws.iter_rows(min_row=2, values_only=False):
        dev = str(r[0].value or "").strip()
        if dev:
            row_by_device[dev.lower()] = r[0].row

    # Contabiliza a partir das abas de países
    counts: Dict[str, Dict[str, int]] = {country: {} for country in VALID_SHEETS}
    for sheet in wb.sheetnames:
        if sheet not in VALID_SHEETS:
            continue
        w = wb[sheet]
        for row in w.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            dev_type = str(row[3] or "").strip()  # coluna D
            if not dev_type:
                continue
            counts[sheet][dev_type] = counts[sheet].get(dev_type, 0) + 1

    # Atualiza células – mantém TOTAL geral (coluna B) e por país
    for dev_type_lower, row_idx in row_by_device.items():
        dev_type_exact = ws.cell(row=row_idx, column=1).value
        total = 0
        for country, col_idx in col_by_country.items():
            val = counts.get(country, {}).get(dev_type_exact, 0)
            ws.cell(row=row_idx, column=col_idx, value=val)
            total += val
        # coluna B é Total
        ws.cell(row=row_idx, column=2, value=total)


# -------------------- Excel file availability --------------------

def is_excel_writable(path: str) -> bool:
    """Check if the Excel file can be opened for write (Windows/Excel lock safe)."""
    try:
        import msvcrt
        with open(path, "rb+") as f:
            try:
                # Try to acquire a non-blocking lock on 1 byte
                msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
                msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
                return True
            except OSError:
                return False
    except Exception:
        # Fallback: try open for write only
        try:
            with open(path, "rb+"):
                return True
        except Exception:
            return False


# -------------------- Sheet compaction (keep stack/HA contiguous) --------------------

def _compact_sheet_group_by_name(ws):
    """Rewrite data rows so that, within each Device Type group, all rows with the
    same Device Name (hostname) are contiguous. Only cell values are rewritten,
    preserving formatting.
    """
    # Read all non-empty rows (A..G)
    rows = []
    for r in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        if not r or not any(r):
            continue
        rows.append(list(r))
    if not rows:
        return

    # Device type priority (same used when inserting)
    type_priority = {
        "Access Point": 1,
        "Switch": 2,
        "Wireless Controller": 3,
        "Router": 4,
        "Firewall": 5,
    }

    # Partition by device type while preserving first appearance order
    def _ptype(v):
        return type_priority.get((v or "").strip(), 99)

    # Build order of types present, respecting priority
    types_present = sorted({(r[3] or "").strip() for r in rows}, key=lambda t: type_priority.get(t, 99))

    new_order = []
    for t in types_present:
        t_rows = [r for r in rows if (r[3] or "").strip() == t]
        # Group by device name preserving first occurrence
        seen = set()
        grouped = []
        for r in t_rows:
            name_l = (str(r[0] or "").strip().lower())
            if name_l not in seen:
                seen.add(name_l)
                # collect entire group for this name
                group = [rr for rr in t_rows if (str(rr[0] or "").strip().lower()) == name_l]
                grouped.extend(group)
        new_order.extend(grouped)

    # Write back values in place to preserve formatting
    # Clear all rows first
    max_data_rows = ws.max_row - 1
    for i in range(max_data_rows):
        row_idx = 2 + i
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col, value=None)
    # Fill with compacted data
    for i, r in enumerate(new_order):
        row_idx = 2 + i
        ws.cell(row=row_idx, column=1, value=r[0])
        ws.cell(row=row_idx, column=2, value=r[1])
        ws.cell(row=row_idx, column=3, value=r[2])
        ws.cell(row=row_idx, column=4, value=r[3])
        ws.cell(row=row_idx, column=5, value=r[4])
        ws.cell(row=row_idx, column=6, value=r[5])
        ws.cell(row=row_idx, column=7, value=r[6])


def _compact_sheet_by_location(ws):
    """Sort rows alphabetically by Location (column C), then by Device Name,
    while keeping rows of the same Device Name contiguous. Only values are
    rewritten; formatting is preserved.
    """
    # Read all non-empty rows (A..G)
    rows = []
    for r in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        if not r or not any(r):
            continue
        rows.append(list(r))
    if not rows:
        return

    def norm(s):
        s = str(s or "").strip().lower()
        return s if s else "~~~~"  # blanks go last

    # Stable sort by: Location (C), Device Name (A), Device Type (D), Serial (G)
    rows.sort(key=lambda r: (norm(r[2]), norm(r[0]), norm(r[3]), norm(r[6])))

    # Write back values in place to preserve formatting
    max_data_rows = ws.max_row - 1
    for i in range(max_data_rows):
        row_idx = 2 + i
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col, value=None)
    for i, r in enumerate(rows):
        row_idx = 2 + i
        ws.cell(row=row_idx, column=1, value=r[0])
        ws.cell(row=row_idx, column=2, value=r[1])
        ws.cell(row=row_idx, column=3, value=r[2])
        ws.cell(row=row_idx, column=4, value=r[3])
        ws.cell(row=row_idx, column=5, value=r[4])
        ws.cell(row=row_idx, column=6, value=r[5])
        ws.cell(row=row_idx, column=7, value=r[6])


def update_workbook(devices: List[DeviceRecord], excel_path: str):
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Planilha alvo não encontrada: {excel_path}")
    backup_excel(excel_path)
    wb = load_wb(excel_path)

    # Garante todas as abas citadas
    for country in VALID_SHEETS:
        ensure_sheet(wb, country)

    # Indexa assinaturas e linhas existentes por aba
    existing: Dict[str, Set[str]] = {}
    indices_by_sheet: Dict[str, Tuple[Dict[Tuple[str, str], int], Dict[str, int]]] = {}
    for sheet in VALID_SHEETS:
        ws = wb[sheet]
        existing[sheet] = read_existing_signatures(ws)
        indices_by_sheet[sheet] = _index_sheet_rows(ws)

    # Ordena para inserir agrupado por tipo e manter HA/stack juntos
    type_priority = {
        "Access Point": 1,
        "Switch": 2,
        "Wireless Controller": 3,
        "Router": 4,
        "Firewall": 5,
    }
    devices_sorted = sorted(
        devices,
        key=lambda r: (type_priority.get(r.device_type, 99), r.device_name.lower(), r.serial_number)
    )

    # Upsert devices
    for rec in devices_sorted:
        country = rec.country or ""
        if not country or country not in VALID_SHEETS:
            continue
        ws = wb[country]
        _upsert_record(ws, rec, existing[country], indices_by_sheet[country])

    # Compaction pass per country sheet to sort alphabetically by Location
    for sheet in VALID_SHEETS:
        _compact_sheet_by_location(wb[sheet])

    # Atualiza a aba Total
    update_totals_sheet(wb)

    wb.save(excel_path)


def prompt_credentials_once() -> Credentials:
    print("Informe as credenciais com privilégios (uma única vez):")
    username = input("Username: ").strip()
    password = getpass.getpass("Password: ")
    # Não solicitar enable secret: usuário já é privilege 15
    return Credentials(username=username, password=password, secret=None)


def bfs_discovery(seed_ip: str, creds: Credentials) -> List[DeviceRecord]:
    """Percorre a topologia via CDP/LLDP (BFS) limitando a rede 10/8 e MAX_NODES.
    Após coletar todos os devices (Switch/Router/WLC/Firewall) coleta APs via WLC;
    se um site não tem WLC mas APs apareceram como vizinhos, tenta acesso direto.
    """
    ctx = DiscoveryContext(creds=creds)
    ctx.queue.append(seed_ip)
    seen_sigs: Set[str] = set()
    while ctx.queue and len(ctx.visited) < MAX_NODES:
        ip = ctx.queue.popleft()
        try:
            process_device(ip, ctx, seen_sigs)
        except Exception as e:
            logging.info("Falha ao processar %s: %s", ip, e)
            continue
    _collect_direct_aps_without_wlc(ctx, seen_sigs)
    return ctx.found_records


class SummaryCounter:
    def __init__(self):
        self.counts = {"Switch":0,"Router":0,"Firewall":0,"Wireless Controller":0,"Access Point":0}
    def add(self, rec: DeviceRecord):
        self.counts.setdefault(rec.device_type,0)
        self.counts[rec.device_type]+=1
    def report(self):
        parts = []
        for k in ["Switch","Router","Firewall","Wireless Controller","Access Point"]:
            parts.append(f"{k}s: {self.counts.get(k,0)}")
        return ", ".join(parts)


def main():
    start_time = time.time()
    logging.getLogger().handlers.clear()
    logging.basicConfig(level=logging.WARNING, format="%(message)s")  # suppress info noise
    try:
        excel_path = resolve_excel_path()
    except Exception as e:
        print(str(e)); sys.exit(2)

    # Loop until file is writable
    while not is_excel_writable(excel_path):
       
        print("A planilha está aberta. Feche o arquivo e pressione Enter para continuar...")
        input()
    seed_ip = input("IP inicial (seed) para descoberta via CDP/LLDP: ").strip()
    try:
        ipaddress.ip_address(seed_ip)
    except Exception:
        print("IP inválido."); sys.exit(2)
    creds = prompt_credentials_once()
    print("Iniciando descoberta...")
    counter = SummaryCounter()
    devices = bfs_discovery(seed_ip, creds)
    for d in devices:
        counter.add(d)
    print(f"Descoberta concluída. {len(devices)} dispositivos: {counter.report()}")
    try:
        update_workbook(devices, excel_path)
        print("Planilha atualizada.")
        # NOVO: Atualiza abas 'End of Life' e 'End of Support' automaticamente
        try:
            update_eox_tabs(excel_path)
            print("Abas 'End of Life' e 'End of Support' atualizadas.")
        except Exception as eox_ex:
            print(f"Aviso: falha ao atualizar EOL/EOS: {eox_ex}")
    except Exception as e:
        print(f"Erro ao atualizar a planilha: {e}"); sys.exit(2)
    elapsed = time.time() - start_time
    print(f"Duração total: {elapsed:.1f}s")


# Para testes isolados (sem Excel, sem coleta real)
def test():
    class FakeConn:
        def send_command(self, cmd, **kwargs):
            if cmd == "show cdp neighbors detail":
                return CDP_NEIGHBORS_SAMPLE
            elif cmd == "show lldp neighbors detail":
                return LLDP_NEIGHBORS_SAMPLE
            return ""

    conn = FakeConn()
    device_type = "cisco_ios"
    neigh, wlc, ap = get_neighbors(conn, device_type)
    print("Vizinhos:", neigh)
    print("WLCs:", wlc)
    print(" APs:", ap)

# CDP sample output (for testing)
CDP_NEIGHBORS_SAMPLE = """
-------------------------
Device ID: Switch1
IP address: 10.0.0.1
Platform: cisco WS-C2960X-48TS-L
Capabilities: Switch IGMP
-------------------------
Device ID: WLC1
IP address: 10.0.0.2
Platform: cisco C9800-40
Capabilities: Controller
-------------------------
Device ID: AP1
IP address: 10.0.0.3
Platform: cisco AIR-AP2802I
Capabilities: Access Point
-------------------------
"""

# LLDP sample output (for testing)
LLDP_NEIGHBORS_SAMPLE = """
-------------------------
System Name: Switch1
Management Address: 10.0.0.1
System Description: Cisco IOS Software, C2960X Software (C2960X-UNIVERSALK9-M), Version 15.2(2)E1, RELEASE SOFTWARE (fc1)
-------------------------
System Name: WLC1
Management Address: 10.0.0.2
System Description: Cisco Wireless LAN Controller Software
-------------------------
System Name: AP1
Management Address: 10.0.0.3
System Description: Cisco Aironet 2800 Series Access Point
-------------------------
"""

def process_device(ip: str, ctx: DiscoveryContext, seen_sigs: Set[str]):
    if ip in ctx.visited:
        return
    ctx.visited.add(ip)
    if not _is_ip_reachable_cached(ip, ctx):
        logging.info("IP não alcançável (ou fora de 10/8): %s", ip)
        return
    try:
        conn, dt = connect_any(ip, ctx.creds)
    except Exception as e:
        logging.info("Falha ao conectar em %s: %s", ip, e)
        return
    # Print simple real-time progress line
    try:
        prompt = conn.find_prompt()
        host_hint = prompt.strip('#>') if prompt else ip
    except Exception:
        host_hint = ip
    print(f"Conectado: {host_hint} ({ip})")
    hostname = ""
    country = ""
    location = ""
    try:
        neigh_ips: Set[str] = set()
        wlc_ips: Set[str] = set()
        ap_infos: List[Tuple[str, str]] = []
        try:
            neigh_ips, wlc_ips, ap_infos = get_neighbors(conn, dt)
        except Exception:
            pass
        if dt in ("cisco_wlc", "cisco_wlc_ssh"):
            wlc_host, model, serials, peer_ip, aps = parse_wlc_info_and_aps(conn)
            hostname = wlc_host
            country, location = infer_site_from_hostname(hostname)
            if country and location:
                ctx.wlc_sites.add(site_key(country, location))
            for _, sn in serials:
                rec = DeviceRecord(device_name=hostname or ip, country=country, location=location,
                                   device_type="Wireless Controller", model=model or "WLC", ip=ip,
                                   serial_number=(sn or "").upper())
                sig = record_signature(rec)
                if sig not in seen_sigs:
                    ctx.found_records.append(rec); seen_sigs.add(sig)
            for ap_name, ap_model, ap_sn, ap_ip in aps:
                ap_rec = DeviceRecord(device_name=ap_name, country=country, location=location,
                                      device_type="Access Point", model=ap_model or "", ip=ap_ip or "",
                                      serial_number=(ap_sn or "").upper())
                sig = record_signature(ap_rec)
                if sig not in seen_sigs:
                    ctx.found_records.append(ap_rec); seen_sigs.add(sig)
            if peer_ip:
                _enqueue(ctx, peer_ip)
            # NOVO: descobrir outros WLCs do grupo de mobilidade / HA
            try:
                _enqueue_additional_wlc_peers(conn, ctx, ip)
            except Exception:
                pass
        elif dt == "cisco_asa":
            hostname, model, serials, peer_ip = parse_asa_device_info(conn)
            country, location = infer_site_from_hostname(hostname)
            for _, sn in serials:
                rec = DeviceRecord(device_name=hostname or ip, country=country, location=location,
                                   device_type="Firewall", model=model or "ASA", ip=ip,
                                   serial_number=(sn or "").upper())
                sig = record_signature(rec)
                if sig not in seen_sigs:
                    ctx.found_records.append(rec); seen_sigs.add(sig)
            if peer_ip:
                _enqueue(ctx, peer_ip)
        else:
            sh_ver = ""
            try:
                sh_ver = safe_send(conn, "show version")
            except Exception:
                pass
            hostname, model, serials = parse_ios_device_info(conn)
            if is_iosxe_wlc_9800(model, sh_ver):
                wlc_host, model_w, serials_w, peer_ip, _ignored = parse_wlc_info_and_aps(conn)
                aps = collect_iosxe_wlc_aps(conn)
                aps = refine_ap_serials_iosxe(conn, aps)  # enrich missing AP serial/model (IOS-XE 9800)
                hostname = wlc_host or hostname
                country, location = infer_site_from_hostname(hostname)
                if country and location:
                    ctx.wlc_sites.add(site_key(country, location))
                for _, sn in serials_w:
                    rec = DeviceRecord(device_name=hostname or ip, country=country, location=location,
                                       device_type="Wireless Controller", model=model_w or model or "WLC", ip=ip,
                                       serial_number=(sn or "").upper())
                    sig = record_signature(rec)
                    if sig not in seen_sigs:
                        ctx.found_records.append(rec); seen_sigs.add(sig)
                for ap_name, ap_model, ap_sn, ap_ip in aps:
                    ap_rec = DeviceRecord(device_name=ap_name, country=country, location=location,
                                          device_type="Access Point", model=ap_model or "", ip=ap_ip or "",
                                          serial_number=(ap_sn or "").upper())
                    sig = record_signature(ap_rec)
                    if sig not in seen_sigs:
                        ctx.found_records.append(ap_rec); seen_sigs.add(sig)
                if peer_ip:
                    _enqueue(ctx, peer_ip)
                # NOVO: peers adicionais (mobilidade) também em IOS-XE
                try:
                    _enqueue_additional_wlc_peers(conn, ctx, ip)
                except Exception:
                    pass
            else:
                country, location = infer_site_from_hostname(hostname)
                dev_kind = _classify_ios_nxos_device_type(model, dt)
                # Ajuste mínimo: hostname terminando com -rtr indica Router
                try:
                    if dev_kind == 'Switch' and hostname and hostname.lower().endswith('-rtr'):
                        dev_kind = 'Router'
                except Exception:
                    pass
                for _, sn in serials:
                    rec = DeviceRecord(device_name=hostname or ip, country=country, location=location,
                                       device_type=dev_kind, model=model or "", ip=ip,
                                       serial_number=(sn or "").upper())
                    sig = record_signature(rec)
                    if sig not in seen_sigs:
                        ctx.found_records.append(rec); seen_sigs.add(sig)
        if ap_infos:
            sk = site_key(country, location) if (country and location) else f"UNKNOWN|{hostname or ip}"
            s = ctx.ap_candidates_by_site.setdefault(sk, set())
            for ap_ip, ap_name in ap_infos:
                s.add((ap_ip, ap_name))
        ap_ips = {ip for (ip, _nm) in ap_infos}
        for n in neigh_ips:
            if n in ap_ips:  # evita processar AP via SSH nesta fase
                continue
            _enqueue(ctx, n)
        for w in wlc_ips:
            _enqueue(ctx, w)
    except Exception as e:
        logging.info("Erro ao processar %s: %s", ip, e)
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass


def refine_ap_serials_iosxe(conn, aps: List[Tuple[str,str,str,str]]) -> List[Tuple[str,str,str,str]]:
    """Tenta complementar serial/model de APs IOS-XE (9800) quando ausente.
    aps: [(name, model, sn, ip)]
    """
    out_list = []
    for name, model, sn, ip in aps:
        if sn and model:
            out_list.append((name, model, sn, ip)); continue
        # comandos extras de inventário/config
        try_cmds = [
            f"show ap name {name} inventory",
            f"show ap name {name} config general",
        ]
        new_model, new_sn = model, sn
        for cmd in try_cmds:
            if new_sn and new_model:
                break
            txt = safe_send(conn, cmd)
            if not txt:
                continue
            if not new_sn:
                m = re.search(r"(?im)AP\s*Serial\s*Number\s*[\.:]+\s*([A-Z0-9]+)", txt) or \
                    re.search(r"(?im)Serial\s*Number\s*[\.:]+\s*([A-Z0-9]+)", txt) or \
                    re.search(r"(?im)System\s*Serial\s*Number\s*[\.:]+\s*([A-Z0-9]+)", txt) or \
                    re.search(r"(?im)SN\s*[: ]+([A-Z0-9]+)", txt)
                if m:
                    new_sn = m.group(1)
                    if AP_SERIAL_DEBUG:
                        logging.info("AP_SERIAL_FETCH iosxe name=%s sn=%s via=%s", name, new_sn, cmd)
            if not new_model:
                m2 = re.search(r"(?im)AP\s*Model\s*[\.:]+\s*([\w-]+)", txt) or \
                     re.search(r"(?im)PID\s*[: ]+([\w-]+)", txt) or \
                     re.search(r"(?im)Model\s*Number\s*[\.:]+\s*([\w-]+)", txt)
                if m2:
                    new_model = m2.group(1)
        out_list.append((name, new_model or model, (new_sn or sn), ip))
    return out_list


def parse_wlc_info_and_aps(conn) -> Tuple[str, str, List[Tuple[str, str]], Optional[str], List[Tuple[str,str,str,str]]]:
    """Coleta info de WLC (AireOS ou 9800) e lista de APs.
    Retorna: (hostname, model, [(label, sn)], peer_mgmt_ip, [(ap_name, ap_model, ap_sn, ap_ip)])
    """
    hostname = ""; model = "WLC"; serials: List[Tuple[str,str]] = []; peer_ip: Optional[str] = None
    sysinfo = safe_send(conn, "show sysinfo") or safe_send(conn, "show system info")
    if sysinfo:
        m = re.search(r"(?im)^(?:System\s*Name|System\s*name)\s*[\.:]+\s*(.+)$", sysinfo)
        if m: hostname = m.group(1).strip()
        m = re.search(r"(?im)^(?:Model|Product/Model|Product)\s*Number\s*[\.:]+\s*([\w-]+)$", sysinfo)
        if m: model = m.group(1).strip()
        m = re.search(r"(?im)^(?:System|Chassis)?\s*Serial\s*Number\s*[\.:]+\s*([A-Z0-9]+)$", sysinfo)
        if m: serials.append(("Chassis", m.group(1)))
    inv_all = safe_send(conn, "show inventory")
    pid = get_pid_from_inventory(inv_all or "")
    if pid: model = pid
    if inv_all:
        # NEW: coletar TODOS os SNs de chassis (Active/Standby) em 9800 (ex.: "Chassis 1" e "Chassis 2")
        found_any = False
        for m in re.finditer(r"(?is)NAME:\s*\"Chassis(?:\s*\d+)?\"[\s\S]*?SN:\s*([A-Z0-9]+)", inv_all):
            sn = m.group(1)
            if sn and not any(sn == s for _, s in serials):
                serials.append(("Chassis", sn)); found_any = True
        if not found_any:
            # Compatibilidade com saída antiga contendo apenas "Chassis"
            m_sn_chassis = re.search(r"(?is)NAME:\s*\"Chassis\"[\s\S]*?SN:\s*([A-Z0-9]+)", inv_all)
            if m_sn_chassis:
                sn = m_sn_chassis.group(1)
                if ("Chassis", sn) not in serials:
                    serials.append(("Chassis", sn)); found_any = True
        if not found_any:
            # Fallback adicional: linhas com PID...SN... associadas a 9800/WLC
            for m in re.finditer(r"(?im)PID:\s*([A-Z0-9+\-\/]+).*?SN:\s*([A-Z0-9]+)", inv_all):
                pid_txt, sn = m.group(1), m.group(2)
                pid_u = (pid_txt or "").upper()
                if "C9800" in pid_u or "WLC" in pid_u:
                    if not any(sn == s for _, s in serials):
                        serials.append(("Chassis", sn)); found_any = True
        if not serials:
            m_any_sn = re.search(r"(?i)\bSN\b\s*[: ]+([A-Z0-9]+)", inv_all)
            if m_any_sn:
                serials.append(("Chassis", m_any_sn.group(1)))
    # Redundancy peers (AireOS/IOS-XE)
    red = safe_send(conn, "show redundancy summary") or safe_send(conn, "show ha status") or safe_send(conn, "show sso")
    if red:
        m = re.search(r"(?i)Peer.*?IP\s*[:=]\s*(\d+\.\d+\.\d+\.\d+)", red)
        if m: peer_ip = m.group(1)
    # AireOS 5xxx: Peer Redundancy Management IP Address
    if not peer_ip:
        red_det = safe_send(conn, "show redundancy detail")
        if red_det:
            m = re.search(r"(?i)Peer\s+Redundancy\s+Management\s+IP\s+Address\.*\s*(\d+\.\d+\.\d+\.\d+)", red_det)
            if m:
                peer_ip = m.group(1)
    # APs AireOS/IOS-XE inventory
    aps: List[Tuple[str,str,str,str]] = []
    inv = safe_send(conn, "show ap inventory all")
    ip_map = parse_ap_summary_ips(conn)
    if inv:
        blocks = re.split(r"(?m)^Cisco\s*AP\s*Name\s*[\.:]+\s*", inv)
        for b in blocks:
            b = b.strip();
            if not b: continue
            lines = [ln for ln in b.splitlines() if ln.strip()]
            if not lines: continue
            ap_name = lines[0].strip()
            if not AP_NAME_RE.match(ap_name) or ap_name.lower() in AP_STOPWORDS or "#" in ap_name:
                continue
            m_model = re.search(r"(?im)^(?:PID|Product\s*ID|Model\s*Number)\s*[\.:]+\s*([\w-]+)$", b)
            m_sn = re.search(r"(?im)^(?:Serial\s*Number|SN)\s*[\.:]+\s*([A-Z0-9]+)$", b)
            ap_model = m_model.group(1) if m_model else ""
            ap_sn = m_sn.group(1) if m_sn else ""
            ap_ip = ip_map.get(ap_name, "")
            if ap_model or ap_sn or ap_ip:
                aps.append((ap_name, ap_model, ap_sn, ap_ip))
    if not aps:
        # fallback summary
        summary = safe_send(conn, "show ap summary")
        for line in summary.splitlines():
            parts = line.split()
            if not parts: continue
            ap_name = parts[0]
            if not AP_NAME_RE.match(ap_name) or ap_name.lower() in AP_STOPWORDS or "#" in ap_name:
                continue
            cfg = safe_send(conn, f"show ap config general {ap_name}")
            m_model = re.search(r"(?im)^\s*AP\s*Model\s*[\.:]+\s*([\w-]+)", cfg)
            m_sn = re.search(r"(?im)^\s*AP\s*Serial\s*Number\s*[\.:]+\s*([A-Z0-9]+)", cfg)
            ap_model = m_model.group(1) if m_model else ""
            ap_sn = m_sn.group(1) if m_sn else ""
            ap_ip = (RE_IP.search(line).group(1) if RE_IP.search(line) else ip_map.get(ap_name, ""))
            if ap_model or ap_sn or ap_ip:
                aps.append((ap_name, ap_model, ap_sn, ap_ip))
    # Passo extra: tentar serial onde faltando (AireOS)
    enriched = []
    for ap_name, ap_model, ap_sn, ap_ip in aps:
        if ap_sn and ap_model:
            enriched.append((ap_name, ap_model, ap_sn, ap_ip)); continue
        # comandos específicos AireOS/IOS-XE
        for cmd in [f"show ap inventory {ap_name}", f"show ap config general {ap_name}"]:
            if ap_sn and ap_model:
                break
            txt = safe_send(conn, cmd)
            if not txt: continue
            if not ap_sn:
                m_sn2 = re.search(r"(?im)(?:Serial\s*Number|SN)\s*[\.:]+\s*([A-Z0-9]+)", txt)
                if m_sn2:
                    ap_sn = m_sn2.group(1)
                    if AP_SERIAL_DEBUG:
                        logging.info("AP_SERIAL_FETCH aireos name=%s sn=%s via=%s", ap_name, ap_sn, cmd)
            if not ap_model:
                m_m2 = re.search(r"(?im)(?:PID|Product\s*ID|Model\s*Number)\s*[\.:]+\s*([\w-]+)", txt)
                if m_m2:
                    ap_model = m_m2.group(1)
        enriched.append((ap_name, ap_model, ap_sn, ap_ip))
    aps = enriched
    return hostname, model, serials or [("Chassis", "")], peer_ip, aps


def parse_ios_device_info(conn) -> Tuple[str, str, List[Tuple[str, str]]]:
    """Retorna (hostname, model, [(member_label, serial)]) para IOS/NX-OS.
    Filtra apenas chassis / membros de stack.
    """
    hostname = ""; model = ""; serials: List[Tuple[str,str]] = []
    sh_ver = safe_send(conn, "show version")
    # hostname
    run_host = safe_send(conn, "show run | i hostname")
    m = re.search(r"(?im)hostname\s+([\w._-]+)", run_host)
    if m: hostname = m.group(1)
    if not hostname:
        try:
            prompt = conn.find_prompt(); hostname = prompt.strip('#>')
        except Exception: hostname = ""
    # model hints
    m = re.search(r"(?i)Model number\s*:\s*([\w-]+)", sh_ver)
    if m: model = m.group(1)
    if not model:
        m = re.search(r"(?i)\bcisco\s+([A-Z0-9\-\/]+)", sh_ver)
        if m: model = m.group(1)
    inv = safe_send(conn, "show inventory")
    if inv:
        entries = re.split(r"(?m)^NAME:\s*\"", inv)
        BAD = ["stackport","stack port","stackadapter","power","powersupply","power supply","fan","module","transceiver","sfp","gbic","line card","supervisor","sup","rps"]
        for e in entries:
            if not e.strip(): continue
            label = e.split("\"",1)[0].strip()
            m_sn = re.search(r"SN:\s*([A-Z0-9]+)", e)
            if not m_sn: continue
            lbl_l = label.lower()
            if any(b in lbl_l for b in BAD):
                continue
            is_switch_member = re.fullmatch(r"(?i)switch\s*\d+(?:\s*chassis)?", label) is not None
            is_chassis = re.search(r"(?i)\bchassis\b", label) is not None
            if not (is_switch_member or is_chassis):
                continue
            serials.append((label, m_sn.group(1)))
        if not serials:
            m_sn = re.search(r"(?i)NAME: \"Chassis\"[\s\S]*?SN:\s*([A-Z0-9]+)", inv)
            if m_sn: serials.append(("Chassis", m_sn.group(1)))
        pid = get_pid_from_inventory(inv or "")
        if pid: model = pid
        # Fallback adicional: alguns chassis em VSS/HA (ex.: 4500/6500) exibem múltiplas linhas
        # "PID: ... , VID: ..., SN: ..." sem blocos NAME distintos. Captura SNs extras.
        if len(serials) <= 1:
            for m in re.finditer(r"(?im)PID:\s*([A-Z0-9+\-\/]+).*?SN:\s*([A-Z0-9]+)", inv):
                pid_txt, sn = m.group(1), m.group(2)
                pid_u = (pid_txt or "").upper()
                # Filtra para chassis (evita SUP/line card: geralmente WS-X/ASR1000-RP etc.)
                if not (pid_u.startswith("WS-C") or pid_u.startswith("C9") or pid_u.startswith("N9") or pid_u.startswith("N3")):
                    continue
                if not any(sn == s for _, s in serials):
                    serials.append(("Chassis", sn))
    if not serials:
        for m in re.finditer(r"(?i)System serial number\s*:\s*([A-Z0-9]+)", sh_ver):
            serials.append(("Chassis", m.group(1)))
    if not serials:
        m = re.search(r"(?i)Processor board ID\s*([A-Z0-9]+)", sh_ver)
        if m: serials.append(("Chassis", m.group(1)))
    if not model or not serials:
        lic = safe_send(conn, "show license udi")
        if not serials:
            m = re.search(r"(?i)PID:\s*[\w\-\/]+[^\n]*SN:\s*([A-Z0-9]+)", lic)
            if m: serials.append(("Chassis", m.group(1)))
        if not model:
            m = re.search(r"(?i)PID:\s*([A-Z0-9\-\/]+)", lic)
            if m: model = m.group(1)
    return hostname, model, serials or [("Chassis", "")]


def parse_asa_device_info(conn) -> Tuple[str, str, List[Tuple[str, str]], Optional[str]]:
    hostname = ""; model = "ASA"; serials: List[Tuple[str,str]] = []; peer_ip: Optional[str] = None
    sh_ver = safe_send(conn, "show version")
    run_host = safe_send(conn, "show running-config hostname")
    m = re.search(r"(?im)^\s*hostname\s+([\w._-]+)", run_host)
    if m: hostname = m.group(1)
    if not hostname:
        try:
            prompt = conn.find_prompt(); hostname = prompt.strip('#>')
        except Exception: pass
    m = re.search(r"(?im)^Hardware:\s*([\w-]+)", sh_ver)
    if m: model = m.group(1)
    inv = safe_send(conn, "show inventory")
    m = re.search(r"(?i)Chassis\s*,\s*SN:\s*([A-Z0-9]+)", inv)
    if m: serials.append(("Chassis", m.group(1)))
    else:
        m = re.search(r"(?i)System serial number\s*:\s*([A-Z0-9]+)", sh_ver)
        if m: serials.append(("Chassis", m.group(1)))
    fail = safe_send(conn, "show failover | i (This|Other|Standby) addr|State|Monitored")
    m = re.search(r"(?i)Standby addr\s*(\d+\.\d+\.\d+\.\d+)", fail)
    if m: peer_ip = m.group(1)
    return hostname, model, serials or [("Chassis", "")], peer_ip


_classify_ios_nxos_device_type_cache = {}

def _classify_ios_nxos_device_type(model: str, netmiko_type: str) -> str:
    m = (model or "").upper()
    if netmiko_type == 'cisco_nxos':
        return 'Switch'
    if m in _classify_ios_nxos_device_type_cache:
        return _classify_ios_nxos_device_type_cache[m]
    switch_hints = ["WS-C","C9200","C9300","C9400","C9500","C9600","C2960","C3560","C3650","C3850","CATALYST","N9K","N3K"]
    router_hints = ["ISR","ASR","CSR","C8","C4"]
    if any(h in m for h in switch_hints): dev_kind = 'Switch'
    elif any(h in m for h in router_hints): dev_kind = 'Router'
    else: dev_kind = 'Switch'
    _classify_ios_nxos_device_type_cache[m] = dev_kind
    return dev_kind


# -------------------- Discovery (BFS) --------------------

def _enqueue(ctx: DiscoveryContext, ip: Optional[str]):
    if not ip: return
    try:
        ipaddress.ip_address(ip)
    except Exception:
        return
    if not _is_in_10_network(ip):
        return
    if ip in ctx.visited or ip in ctx.queue:
        return
    ctx.queue.append(ip)


def _collect_direct_aps_without_wlc(ctx: DiscoveryContext, seen_sigs: Set[str]):
    creds = _load_ap_env_credentials()
    if not creds:
        logging.info("AP .env não encontrado ou sem credenciais válidas; ignorando APs diretos.")
        return
    for sk, candidates in ctx.ap_candidates_by_site.items():
        if sk in ctx.wlc_sites:
            continue
        try:
            country, location = sk.split('|',1)
        except ValueError:
            country, location = "", ""
        for ap_ip, ap_name in sorted(candidates):
            if not ap_ip or not _is_in_10_network(ap_ip):
                continue
            if not is_ip_reachable(ap_ip):
                continue
            try:
                conn, dt = connect_any(ap_ip, creds)
                # --- NOVO: imprimir conexão a APs diretos, semelhante aos demais dispositivos ---
                try:
                    prompt = conn.find_prompt()
                    host_hint = prompt.strip('#>') if prompt else (ap_name or ap_ip)
                except Exception:
                    host_hint = ap_name or ap_ip
                print(f"Conectado: {host_hint} ({ap_ip})")
                # ------------------------------------------------------------------------------
                try:
                    host, model, serials = parse_ios_device_info(conn)
                    name_final = ap_name or host or ap_ip
                    if not country or not location:
                        country, location = infer_site_from_hostname(name_final)
                    sn_val = ""
                    for _lbl, sn in serials:
                        if sn: sn_val = sn; break
                    # Fallback mínimo para AP: se SN ou modelo não vieram, tentar no 'show inventory'
                    if not sn_val or not model:
                        inv_txt = safe_send(conn, "show inventory")
                        if inv_txt:
                            if not sn_val:
                                m_sn = re.search(r"(?i)\bSN\s*:\s*([A-Z0-9]+)", inv_txt)
                                if m_sn:
                                    sn_val = m_sn.group(1)
                            if not model:
                                m_pid = re.search(r"(?i)\bPID\s*:\s*([A-Z0-9\-]+)", inv_txt)
                                if m_pid:
                                    model = m_pid.group(1)
                    rec = DeviceRecord(device_name=name_final, country=country, location=location,
                                       device_type="Access Point", model=model or "", ip=ap_ip,
                                       serial_number=(sn_val or "").upper())
                    sig = record_signature(rec)
                    if sig not in seen_sigs:
                        ctx.found_records.append(rec); seen_sigs.add(sig)
                except Exception as e:
                    logging.info("Erro coletando AP direto %s (%s): %s", ap_name, ap_ip, e)
                finally:
                    try: conn.disconnect()
                    except Exception: pass
            except Exception as e:
                logging.info("Falha ao conectar no AP %s (%s): %s", ap_name, ap_ip, e)
                continue
           

# Descobre WLCs pares adicionais (mobilidade / HA) e enfileira

def _enqueue_additional_wlc_peers(conn, ctx: DiscoveryContext, self_ip: str):
    cmds = [
        "show mobility summary",               # AireOS
        "show wireless mobility summary",      # IOS-XE 9800
        "show redundancy summary",             # IOS-XE/AireOS
        "show redundancy peers",               # possíveis variantes
        "show redundancy detail",              # AireOS 5xxx: Peer Redundancy Management IP Address
    ]
    seen = set()
    for cmd in cmds:
        try:
            out = safe_send(conn, cmd)
        except Exception:
            continue
        if not out:
            continue
        for cand in set(RE_IP.findall(out)):
            if cand == self_ip:
                continue
            if cand in seen:
                continue
            seen.add(cand)
            _enqueue(ctx, cand)
            
def _load_ap_env_credentials() -> Optional[Credentials]:
    """Carrega credenciais para acesso direto aos APs a partir de arquivos .env.
    Procura em 'Inventário' e 'Inventario' e aceita múltiplas combinações de chaves.
    """
    candidate = [
        os.path.join(WORKDIR, 'Inventário', '.env'),
        os.path.join(WORKDIR, 'Inventário', 'Network.env'),
        os.path.join(WORKDIR, 'Inventario', '.env'),
        os.path.join(WORKDIR, 'Inventario', 'Network.env'),
    ]
    for p in candidate:
        if os.path.exists(p):
            try:
                load_dotenv(p)
            except Exception:
                pass
            for u_key, p_key in [("AP_USERNAME","AP_PASSWORD"),("AP_USER","AP_PASS"),("USERNAME","PASSWORD")]:
                u = os.getenv(u_key) or ""; pw = os.getenv(p_key) or ""
                if u and pw:
                    return Credentials(username=u, password=pw, secret=None)
    return None

if __name__ == "__main__":
    main()